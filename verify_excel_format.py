#!/usr/bin/env python3
"""
验证大塬QC报表Excel格式 - 确保完全遵循历史记录页面设置
"""
import os
import sys
import django
from datetime import date, timedelta

# 设置Django环境
sys.path.append('/var/www/yuantong')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'yuantong.settings')
django.setup()

from home.models import DayuanQCReport
from tasks.tasks import generate_dayuan_excel_report
from home.views import QC_REPORT_FIELD_MAPPING
from openpyxl import load_workbook

def verify_excel_format():
    """验证Excel格式完全遵循历史记录页面设置"""
    print("🔍 验证Excel格式完全遵循历史记录页面设置")
    print("=" * 60)
    
    # 获取昨日数据
    yesterday = date.today() - timedelta(days=1)
    reports = DayuanQCReport.objects.filter(date=yesterday)
    
    print(f"📅 昨日({yesterday})大塬QC报表数量: {reports.count()}")
    
    if not reports.exists():
        print("⚠️  昨日没有数据，无法验证Excel格式")
        return
    
    try:
        # 生成Excel文件
        print("📊 生成Excel报表...")
        excel_file_path = generate_dayuan_excel_report(reports, yesterday)
        print(f"✅ Excel文件生成成功: {excel_file_path}")
        
        # 验证Excel文件内容
        print("🔍 验证Excel文件格式...")
        wb = load_workbook(excel_file_path)
        ws = wb.active
        
        print(f"📋 工作表标题: {ws.title}")
        print(f"📊 数据维度: {ws.max_row}行 x {ws.max_column}列")
        
        # 验证列宽设置
        print("\n📏 列宽设置验证:")
        print("=" * 50)
        
        # 历史记录页面的列宽设置
        expected_widths = {
            'Date日期': 13.25,
            'IPKP CODE包装类型': 24,
            '操作人': 14,
            'LOT批号/日期': 13.5,
        }
        
        all_correct = True
        for col in range(1, ws.max_column + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            actual_width = ws.column_dimensions[column_letter].width
            header = ws.cell(row=1, column=col).value
            
            if header in expected_widths:
                expected_width = expected_widths[header]
                status = "✅" if actual_width == expected_width else "❌"
                print(f"{status} 列{col:2d} ({column_letter}): {header[:30]:<30} 期望: {expected_width:>6} 实际: {actual_width:>6}")
                if actual_width != expected_width:
                    all_correct = False
            else:
                expected_width = 8.3  # 默认宽度
                status = "✅" if actual_width == expected_width else "❌"
                print(f"{status} 列{col:2d} ({column_letter}): {header[:30]:<30} 期望: {expected_width:>6} 实际: {actual_width:>6}")
                if actual_width != expected_width:
                    all_correct = False
        
        print(f"\n📊 列宽设置验证结果: {'✅ 全部正确' if all_correct else '❌ 存在错误'}")
        
        # 验证边框设置
        print("\n🔲 边框设置验证:")
        print("=" * 50)
        
        border_correct = True
        for row in range(1, min(4, ws.max_row + 1)):  # 检查前3行
            for col in range(1, min(6, ws.max_column + 1)):  # 检查前5列
                cell = ws.cell(row=row, column=col)
                border = cell.border
                
                # 检查是否有边框
                has_border = (
                    border.left.style == 'thin' and
                    border.right.style == 'thin' and
                    border.top.style == 'thin' and
                    border.bottom.style == 'thin'
                )
                
                status = "✅" if has_border else "❌"
                print(f"{status} 行{row}列{col}: {str(cell.value)[:20]:<20} 边框: {'有' if has_border else '无'}")
                
                if not has_border:
                    border_correct = False
        
        print(f"\n📊 边框设置验证结果: {'✅ 全部正确' if border_correct else '❌ 存在错误'}")
        
        # 验证表头样式
        print("\n🎨 表头样式验证:")
        print("=" * 50)
        
        header_style_correct = True
        for col in range(1, min(6, ws.max_column + 1)):  # 检查前5列
            cell = ws.cell(row=1, column=col)
            
            # 检查字体样式
            font_bold = cell.font.bold
            font_size = cell.font.size
            font_color = cell.font.color.rgb if cell.font.color else None
            
            # 检查填充样式
            fill_color = cell.fill.start_color.rgb if cell.fill.start_color else None
            
            # 检查对齐方式
            alignment_center = (cell.alignment.horizontal == 'center' and 
                              cell.alignment.vertical == 'center')
            
            expected_font_color = "00FFFFFF"  # 白色
            expected_fill_color = "001976D2"  # 蓝色
            expected_font_size = 11.0
            expected_bold = True
            
            font_color_ok = font_color == expected_font_color
            fill_color_ok = fill_color == expected_fill_color
            font_size_ok = font_size == expected_font_size
            bold_ok = font_bold == expected_bold
            alignment_ok = alignment_center
            
            all_style_ok = font_color_ok and fill_color_ok and font_size_ok and bold_ok and alignment_ok
            status = "✅" if all_style_ok else "❌"
            
            print(f"{status} 列{col}: {str(cell.value)[:20]:<20}")
            print(f"     字体颜色: {font_color} {'✅' if font_color_ok else '❌'} (期望: {expected_font_color})")
            print(f"     填充颜色: {fill_color} {'✅' if fill_color_ok else '❌'} (期望: {expected_fill_color})")
            print(f"     字体大小: {font_size} {'✅' if font_size_ok else '❌'} (期望: {expected_font_size})")
            print(f"     是否加粗: {font_bold} {'✅' if bold_ok else '❌'} (期望: {expected_bold})")
            print(f"     居中对齐: {alignment_center} {'✅' if alignment_ok else '❌'}")
            print()
            
            if not all_style_ok:
                header_style_correct = False
        
        print(f"📊 表头样式验证结果: {'✅ 全部正确' if header_style_correct else '❌ 存在错误'}")
        
        # 验证数据样式
        print("\n📄 数据样式验证:")
        print("=" * 50)
        
        data_style_correct = True
        for row in range(2, min(4, ws.max_row + 1)):  # 检查前2行数据
            for col in range(1, min(6, ws.max_column + 1)):  # 检查前5列
                cell = ws.cell(row=row, column=col)
                
                # 检查字体样式
                font_size = cell.font.size
                font_bold = cell.font.bold
                
                # 检查对齐方式
                alignment_center = (cell.alignment.horizontal == 'center' and 
                                  cell.alignment.vertical == 'center')
                
                expected_font_size = 11.0
                expected_bold = False
                
                font_size_ok = font_size == expected_font_size
                bold_ok = font_bold == expected_bold
                alignment_ok = alignment_center
                
                all_style_ok = font_size_ok and bold_ok and alignment_ok
                status = "✅" if all_style_ok else "❌"
                
                print(f"{status} 行{row}列{col}: {str(cell.value)[:20]:<20}")
                print(f"     字体大小: {font_size} {'✅' if font_size_ok else '❌'} (期望: {expected_font_size})")
                print(f"     是否加粗: {font_bold} {'✅' if bold_ok else '❌'} (期望: {expected_bold})")
                print(f"     居中对齐: {alignment_center} {'✅' if alignment_ok else '❌'}")
                print()
                
                if not all_style_ok:
                    data_style_correct = False
        
        print(f"📊 数据样式验证结果: {'✅ 全部正确' if data_style_correct else '❌ 存在错误'}")
        
        # 验证数字格式
        print("\n🔢 数字格式验证:")
        print("=" * 50)
        
        # 查找吨数字段
        tons_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and 'Tons吨' in str(header):
                tons_col = col
                break
        
        if tons_col:
            print(f"吨数字段位置: 第{tons_col}列")
            number_format_correct = True
            
            for row in range(2, min(6, ws.max_row + 1)):
                cell = ws.cell(row=row, column=tons_col)
                if cell.value:
                    actual_format = cell.number_format
                    expected_format = '0.000'
                    format_ok = actual_format == expected_format
                    status = "✅" if format_ok else "❌"
                    
                    print(f"{status} 行{row}: {cell.value} 格式: {actual_format} {'✅' if format_ok else '❌'} (期望: {expected_format})")
                    
                    if not format_ok:
                        number_format_correct = False
        else:
            print("未找到吨数字段")
            number_format_correct = False
        
        print(f"📊 数字格式验证结果: {'✅ 全部正确' if number_format_correct else '❌ 存在错误'}")
        
        # 总体验证结果
        print("\n" + "=" * 60)
        print("🎯 总体验证结果:")
        print("=" * 60)
        
        all_verifications = [
            ("列宽设置", all_correct),
            ("边框设置", border_correct),
            ("表头样式", header_style_correct),
            ("数据样式", data_style_correct),
            ("数字格式", number_format_correct)
        ]
        
        all_passed = all(result for _, result in all_verifications)
        
        for name, result in all_verifications:
            status = "✅ 通过" if result else "❌ 失败"
            print(f"{status} {name}")
        
        print(f"\n🏆 总体结果: {'✅ 全部通过' if all_passed else '❌ 存在问题'}")
        
        # 清理临时文件
        try:
            os.unlink(excel_file_path)
            print("\n🗑️  临时文件已清理")
        except:
            pass
        
    except Exception as e:
        print(f"❌ Excel格式验证失败: {str(e)}")
        import traceback
        traceback.print_exc()

def main():
    """主函数"""
    verify_excel_format()
    
    print("\n" + "=" * 60)
    print("🎉 Excel格式验证完成！")
    print("\n📝 验证项目:")
    print("1. 列宽设置 - 关键列使用固定宽度，其他列使用8.3")
    print("2. 边框设置 - 所有单元格都有细边框")
    print("3. 表头样式 - 蓝色背景，白色字体，11号字，加粗，居中")
    print("4. 数据样式 - 11号字，不加粗，居中对齐")
    print("5. 数字格式 - 吨数字段保留3位小数")

if __name__ == '__main__':
    main()
