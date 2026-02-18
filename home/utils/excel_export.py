"""
Excel导出工具模块
提供各种Excel导出功能
"""

import logging
import urllib.parse
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def export_production_excel(request, model_class, report_name, period):
    """导出产量统计Excel"""
    logger.info(f"=== 开始导出{report_name}{period}产量统计 ===")
    logger.info(f"请求用户: {request.user.username}")
    logger.info(f"请求方法: {request.method}")
    logger.info(f"模型类: {model_class.__name__}")
    logger.info(f"报表名称: {report_name}")
    logger.info(f"统计周期: {period}")
    
    try:
        logger.info("✅ openpyxl库导入成功")
        
        # 确定日期范围
        if period == "昨日":
            target_date = (datetime.now() - timedelta(days=1)).date()
            start_date = target_date
            end_date = target_date
        else:  # 今日
            target_date = datetime.now().date()
            start_date = target_date
            end_date = target_date
        
        logger.info(f"📅 目标日期: {target_date}")
        logger.info(f"📅 查询开始日期: {start_date}")
        logger.info(f"📅 查询结束日期: {end_date}")
        
        # 构建查询
        query = Q(date__gte=start_date) & Q(date__lte=end_date)
        logger.info(f"🔍 构建查询条件: {query}")
        
        # 执行查询
        reports = model_class.objects.filter(query).order_by('shift', 'product_name', 'packaging', 'batch_number', 'remarks')
        logger.info(f"📊 查询到{reports.count()}条记录")
        
        if not reports.exists():
            logger.warning(f"⚠️ {report_name}{period}没有找到产量数据")
            return HttpResponse(f"{report_name}{period}没有找到产量数据", content_type='text/plain')
        
        # 按5个字段分组累加产量（与统计报表逻辑保持一致）
        logger.info("📊 开始按5个字段分组累加产量...")
        grouped_production = {}
        for report in reports:
            # 创建分组键
            group_key = (
                report.shift or '未设置',
                report.product_name or '未设置',
                report.packaging or '未设置',
                report.batch_number or '未设置',
                report.remarks or '未设置'
            )
            
            if group_key not in grouped_production:
                grouped_production[group_key] = {
                    'shift': group_key[0],
                    'product_name': group_key[1],
                    'packaging': group_key[2],
                    'batch_number': group_key[3],
                    'remarks': group_key[4],
                    'total_tons': 0,
                    'count': 0
                }
            
            # 累加产量
            try:
                if report.tons is not None:
                    grouped_production[group_key]['total_tons'] += float(report.tons)
                    grouped_production[group_key]['count'] += 1
            except (ValueError, TypeError):
                logger.warning(f"    ⚠️ 产量值转换失败: {report.tons}")
                continue
        
        logger.info(f"📊 分组累加完成，共{len(grouped_production)}个唯一组合")
        
        # 按班组分组数据（用于Excel显示）
        grouped_data = {}
        for group_key, production_data in grouped_production.items():
            shift = production_data['shift']
            if shift not in grouped_data:
                grouped_data[shift] = []
            grouped_data[shift].append(production_data)
        
        logger.info(f"📋 按班组分组完成，共{len(grouped_data)}个班组")
        for shift, shift_reports in grouped_data.items():
            logger.info(f"  - 班组 '{shift}': {len(shift_reports)}条记录")
        
        # 创建Excel工作簿
        logger.info("📊 开始创建Excel工作簿...")
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name}{period}产量统计"
        logger.info(f"✅ Excel工作表创建成功，标题: {ws.title}")
        
        # 设置表头
        headers = ['班组', '产品型号', '包装类型', '批号', '备注', '产量(吨)', '班组产量(吨)']
        ws.append(headers)
        logger.info(f"✅ 表头设置完成: {headers}")
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        logger.info("📝 开始写入数据到Excel...")
        row_num = 2
        total_tons = 0
        
        for shift, shift_reports in grouped_data.items():
            logger.info(f"  📊 处理班组 '{shift}'，共{len(shift_reports)}条记录")
            
            # 计算班组产量（使用累加后的数据）
            shift_total = sum(production_data['total_tons'] for production_data in shift_reports)
            total_tons += shift_total
            logger.info(f"    📈 班组总产量: {shift_total}吨")
            
            # 记录班组开始行，用于合并单元格
            shift_start_row = row_num
            
            # 为班组中的每个产品创建行
            for i, production_data in enumerate(shift_reports):
                # 使用累加后的产量数据
                tons = production_data['total_tons']
                count = production_data['count']
                
                # 设置所有列的数据
                ws.cell(row=row_num, column=2, value=production_data['product_name'])
                ws.cell(row=row_num, column=3, value=production_data['packaging'])
                ws.cell(row=row_num, column=4, value=production_data['batch_number'])
                ws.cell(row=row_num, column=5, value=production_data['remarks'])
                ws.cell(row=row_num, column=6, value=tons)
                
                # 设置产量列样式
                ws.cell(row=row_num, column=6).font = Font(bold=True, color="4CAF50")
                
                # 备注字段保持原始内容，不显示累加信息
                
                row_num += 1
            
            # 合并班组列和班组产量列
            if len(shift_reports) > 1:
                # 合并班组列
                ws.merge_cells(f'A{shift_start_row}:A{row_num-1}')
                # 合并班组产量列
                ws.merge_cells(f'G{shift_start_row}:G{row_num-1}')
            
            # 设置合并后的单元格内容和样式
            ws.cell(row=shift_start_row, column=1, value=shift)
            ws.cell(row=shift_start_row, column=7, value=shift_total)
            
            # 设置班组列样式
            ws.cell(row=shift_start_row, column=1).font = Font(bold=True, color="1976D2")
            ws.cell(row=shift_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置班组产量单元格样式
            ws.cell(row=shift_start_row, column=7).font = Font(bold=True, color="1976D2")
            ws.cell(row=shift_start_row, column=7).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            ws.cell(row=shift_start_row, column=7).alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"✅ 数据写入完成，总行数: {row_num-1}，总产量: {total_tons}吨")
        
        # 添加总计行
        ws.append(['总计', '', '', '', '', total_tons, ''])
        total_row = ws[ws.max_row]
        
        # 设置总计行样式
        total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF")
        total_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in total_row:
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = total_alignment
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
        
        # 自适应列宽
        logger.info("📏 开始设置列宽...")
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        content = str(cell.value)
                        chinese_chars = len([c for c in content if '\u4e00' <= c <= '\u9fff'])
                        length = len(content) + chinese_chars
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        logger.info("✅ 列宽设置完成")
        
        # 创建响应
        logger.info("🌐 开始创建HTTP响应...")
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 处理文件名编码，企业微信兼容性优化
        filename = f"{report_name}{period}产量统计_{target_date}.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        logger.info(f"📁 文件名: {filename}")
        logger.info(f"📁 编码后文件名: {encoded_filename}")
        
        # 使用RFC 5987标准编码，兼容企业微信
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
        
        # 添加额外的响应头，提高企业微信兼容性
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'must-revalidate'
        response['Pragma'] = 'public'
        
        # 添加企业微信兼容性头部
        response['X-Content-Type-Options'] = 'nosniff'
        response['Accept-Ranges'] = 'bytes'
        
        logger.info("📋 响应头设置完成:")
        logger.info(f"  - Content-Type: {response['Content-Type']}")
        logger.info(f"  - Content-Disposition: {response['Content-Disposition']}")
        
        # 保存到响应
        logger.info("💾 开始保存Excel到响应...")
        wb.save(response)
        logger.info("✅ Excel保存成功，准备返回响应")
        return response
        
    except Exception as e:
        import traceback
        error_msg = f"导出{report_name}{period}产量统计失败: {str(e)}"
        logger.error(error_msg)
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        logger.error(f"错误类型: {type(e).__name__}")
        if hasattr(e, '__traceback__') and e.__traceback__:
            logger.error(f"错误位置: {e.__traceback__.tb_frame.f_code.co_filename}:{e.__traceback__.tb_lineno}")
        return HttpResponse(f"导出失败: {str(e)}", content_type='text/plain')


def export_qc_report_excel(request, model_class, report_name, field_mapping):
    """通用的QC报表Excel导出函数 (智能隐藏空列版本)"""
    try:
        # 1. 获取筛选参数
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        product_name = request.GET.get('product_name')
        packaging = request.GET.get('packaging')
        squad = request.GET.get('squad')

        # 2. 构建查询
        query = Q()
        if start_date:
            query &= Q(date__gte=start_date)
        if end_date:
            query &= Q(date__lte=end_date)
        if start_time:
            query &= Q(time__gte=start_time)
        if end_time:
            query &= Q(time__lte=end_time)
        if product_name:
            query &= Q(product_name__icontains=product_name)
        if packaging:
            query &= Q(packaging__icontains=packaging)
        if squad:
            query &= Q(shift__icontains=squad)

        # 3. 执行查询
        reports = model_class.objects.filter(query).order_by('date', 'time')
        logger.info(f"导出{report_name}，查询到{reports.count()}条记录")

        # 4. 检查每个字段是否在所有行中均为空值
        field_values = {field: [] for field in field_mapping.keys()}
        
        for report in reports:
            # 获取用户真实姓名
            user_display_name = '-'
            if report.user:
                user_display_name = report.user.first_name or report.user.username
            elif report.username:
                try:
                    user = User.objects.get(username=report.username)
                    user_display_name = user.first_name or user.username
                except User.DoesNotExist:
                    user_display_name = report.username

            # 收集每个字段的值
            for field in field_mapping.keys():
                if field == 'username':
                    field_values[field].append(user_display_name)
                elif field == 'date':
                    field_values[field].append(report.date.strftime('%Y-%m-%d') if report.date else '')
                elif field == 'time':
                    field_values[field].append(report.time.strftime('%H:%M') if report.time else '')
                else:
                    value = getattr(report, field, None)
                    field_values[field].append(str(value) if value is not None else '')

        # 5. 过滤掉所有行均为空值的字段
        non_empty_fields = []
        for field, values in field_values.items():
            # 检查是否有任何非空值
            if any(str(value).strip() for value in values if value is not None and str(value).strip()):
                non_empty_fields.append(field)

        # 6. 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name} QC报表"

        # 7. 写入表头（只包含非空字段）
        headers = [field_mapping[field] for field in non_empty_fields]
        ws.append(headers)

        # 8. 写入数据（只包含非空字段）
        for i in range(len(reports)):
            row = [field_values[field][i] for field in non_empty_fields]
            ws.append(row)

        # 9. 自适应列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        content = str(cell.value)
                        # 简单计算：一个中文字符约等于2个英文字符宽度
                        chinese_chars = len([c for c in content if '\u4e00' <= c <= '\u9fff'])
                        length = len(content) + chinese_chars
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50) # 最小宽度10，最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width

        # 10. 设置响应头
        now_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'{report_name}QC报表_{now_str}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'{urllib.parse.quote(filename)}'
        # 保存到响应
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Excel export failed for {report_name}: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'导出失败: {str(e)}'
        }, status=500)


def export_qc_report_excel_universal(request, model_class, report_name, field_mapping, use_formatted_style=False):
    """通用的QC报表Excel导出函数 - 支持大塬格式和标准格式"""
    from django.db import models
    try:
        # 1. 获取筛选参数
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        product_name = request.GET.get('product_name')
        packaging = request.GET.get('packaging')
        squad = request.GET.get('squad')

        # 2. 构建查询
        query = Q()
        if start_date:
            query &= Q(date__gte=start_date)
        if end_date:
            query &= Q(date__lte=end_date)
        if start_time:
            query &= Q(time__gte=start_time)
        if end_time:
            query &= Q(time__lte=end_time)
        if product_name:
            query &= Q(product_name__icontains=product_name)
        if packaging:
            query &= Q(packaging__icontains=packaging)
        if squad:
            query &= Q(shift__icontains=squad)

        # 3. 执行查询 - 按时间从旧到新排序（与大塬保持一致）
        reports = model_class.objects.filter(query).order_by('date', 'time')
        logger.info(f"导出{report_name}QC历史记录，查询到{reports.count()}条记录")

        if not reports.exists():
            logger.warning(f"⚠️ {report_name}没有找到QC数据")
            return HttpResponse(f"{report_name}没有找到QC数据", content_type='text/plain')

        # 4. 检查每个字段是否在所有行中均为空值
        field_values = {field: [] for field in field_mapping.keys()}
        
        for report in reports:
            # 获取用户真实姓名
            user_display_name = '-'
            if report.user:
                user_display_name = report.user.first_name or report.user.username
            elif report.username:
                try:
                    user = User.objects.get(username=report.username)
                    user_display_name = user.first_name or user.username
                except User.DoesNotExist:
                    user_display_name = report.username

            # 收集每个字段的值
            for field in field_mapping.keys():
                if field == 'username':
                    field_values[field].append(user_display_name)
                elif field == 'date':
                    field_values[field].append(report.date.strftime('%Y-%m-%d') if report.date else '')
                elif field == 'time':
                    field_values[field].append(report.time.strftime('%H:%M') if report.time else '')
                else:
                    value = getattr(report, field, None)
                    # 对于数字类型字段，保持数字格式；对于其他字段，转换为字符串
                    if value is not None:
                        if use_formatted_style:
                            # 检查字段类型，如果是数字类型则保持数字格式
                            field_obj = report._meta.get_field(field)
                            if isinstance(field_obj, (models.FloatField, models.IntegerField, models.DecimalField)):
                                field_values[field].append(value)  # 保持数字格式
                                logger.debug(f"字段 {field} 保持数字格式: {value}")
                            else:
                                field_values[field].append(str(value))  # 转换为字符串
                        else:
                            field_values[field].append(str(value))  # 转换为字符串
                    else:
                        field_values[field].append('')  # 空值保持为空字符串

        # 5. 过滤掉所有行均为空值的字段
        non_empty_fields = []
        for field, values in field_values.items():
            # 检查是否有任何非空值（数字类型字段需要特殊处理）
            has_non_empty = False
            for value in values:
                if value is not None:
                    if isinstance(value, (int, float)):
                        has_non_empty = True
                        break
                    elif str(value).strip():
                        has_non_empty = True
                        break
            if has_non_empty:
                non_empty_fields.append(field)

        # 6. 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name} QC历史记录"

        # 7. 写入表头（只包含非空字段）
        headers = [field_mapping[field] for field in non_empty_fields]
        ws.append(headers)

        # 8. 写入数据（只包含非空字段）
        for i in range(len(reports)):
            row = [field_values[field][i] for field in non_empty_fields]
            ws.append(row)

        if use_formatted_style:
            # 9. 设置表头样式（大塬格式）
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 10. 设置数据行样式（大塬格式）
            data_font = Font(size=11)
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for row in ws.iter_rows(min_row=2):
                for col_idx, cell in enumerate(row):
                    cell.font = data_font
                    cell.alignment = data_alignment
                    
                    # 为吨数字段设置特殊格式（小数点后三位）
                    if col_idx < len(non_empty_fields):
                        field_name = non_empty_fields[col_idx]
                        if field_name == 'tons':
                            cell.number_format = '0.000'

            # 11. 设置列宽（大塬格式）
            column_widths = {
                'Date日期': 13.25,
                'IPKP CODE包装类型': 24,
                '操作人': 14,
                'LOT批号/日期': 13.5,
            }
            
            # 应用列宽设置
            for i, field in enumerate(non_empty_fields, 1):
                column_letter = get_column_letter(i)
                header_name = field_mapping[field]
                
                # 检查是否有特定的列宽设置
                if header_name in column_widths:
                    ws.column_dimensions[column_letter].width = column_widths[header_name]
                    logger.info(f"设置列 {header_name} 宽度为 {column_widths[header_name]}")
                else:
                    # 其余字段设置为8.3
                    ws.column_dimensions[column_letter].width = 8.3
                    logger.info(f"设置列 {header_name} 宽度为 8.3")

            # 12. 设置边框（大塬格式）
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
        else:
            # 9. 自适应列宽（标准格式）
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            content = str(cell.value)
                            # 简单计算：一个中文字符约等于2个英文字符宽度
                            chinese_chars = len([c for c in content if '\u4e00' <= c <= '\u9fff'])
                            length = len(content) + chinese_chars
                            if length > max_length:
                                max_length = length
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50) # 最小宽度10，最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width

        # 13. 设置响应头
        now_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'{report_name}QC历史记录_{now_str}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'{urllib.parse.quote(filename)}'
        
        # 保存到响应
        wb.save(response)
        logger.info(f"✅ {report_name}QC历史记录Excel导出完成，文件名: {filename}")
        return response

    except Exception as e:
        logger.error(f"{report_name}QC历史记录Excel导出失败: {str(e)}", exc_info=True)
        return JsonResponse({
            'status': 'error',
            'message': f'导出失败: {str(e)}'
        }, status=500)
