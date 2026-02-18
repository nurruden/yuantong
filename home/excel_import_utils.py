"""
Excel导入工具模块
提供通用的Excel导入功能，用于QC报表等模块的数据导入
"""
import logging
from datetime import datetime, date, time
from django.http import JsonResponse

logger = logging.getLogger(__name__)


# ==================== 兴辉报表列名映射配置 ====================

XINGHUI_COLUMN_MAPPING = {
    # 基本字段映射
    '日期': 'date', 'Date': 'date', '检测日期': 'date',
    '时间': 'time', 'Time': 'time', '检测时间': 'time',
    '班次': 'shift', 'Shift': 'shift', 'Squad': 'shift', '班组': 'shift', '班别': 'shift',
    '产品名称': 'product_name', 'Product Name': 'product_name', 'Grade': 'product_name', 
    '产品型号': 'product_name',
    '包装类型': 'packaging', 'Packaging': 'packaging', 'IPKP CODE': 'packaging',
    '批号': 'batch_number', 'Batch Number': 'batch_number', '批次号': 'batch_number', 
    'LOT': 'batch_number', 'LOT批号': 'batch_number', '批号/日期': 'batch_number', '批次': 'batch_number',
    
    # 检测数据字段映射
    '烘干后原土水分 (%)': 'moisture_after_drying',
    '烘干后原土水分（%）': 'moisture_after_drying',
    '烘干后原土水分(%)': 'moisture_after_drying',
    '干燥后原土水分(%)': 'moisture_after_drying',
    '干燥后原土水分（%）': 'moisture_after_drying',
    '烘干后原': 'moisture_after_drying',
    'Moisture after drying': 'moisture_after_drying', 
    '干燥后原土水分': 'moisture_after_drying',
    
    # 入窑前碱含量
    '入窑前碱含量(%)': 'alkali_content', 
    '入窑前碱含量': 'alkali_content',
    '入窑前': 'alkali_content',
    '含量 (%)': 'alkali_content',
    'Alkali content (%)': 'alkali_content',
    
    '助剂添加比例': 'flux',
    '助溶剂添加比例': 'flux', 
    '助溶剂': 'flux', 
    '助磨剂添加比例': 'flux',
    '*flux agent': 'flux', 
    'flux agent addition ratio': 'flux',
    
    # 渗透率 - 兴辉有三个渗透率字段
    '运通滤速率': 'permeability',
    '迈通渗透率': 'permeability',
    '远通渗透率(Darcy)': 'permeability', 
    '远通渗透率': 'permeability',
    '长高滤速率': 'permeability_long',
    '长富渗透率(Darcy)': 'permeability_long', 
    '长富渗透率': 'permeability_long',
    '兴桐渗透率': 'xinghui_permeability',
    '兴辉渗透率(Darcy)': 'xinghui_permeability',
    '兴辉渗透率': 'xinghui_permeability',
    
    # 可塑度可能是涡值
    '可塑度 (c/cm)': 'swirl',
    '可塑度': 'swirl',
    '涡值(cm)': 'swirl',
    '涡值（cm）': 'swirl',
    '滤值(cm)': 'swirl',
    '涡值': 'swirl', 
    'Swirl (cm)': 'swirl', 
    'Swirl': 'swirl',
    
    # 饼密度和振实密度
    '饼密度(g/cm3)': 'wet_cake_density', 
    '饼密度（g/cm3）': 'wet_cake_density',
    '饼密度': 'wet_cake_density', 
    '筛密度(g/cm3)': 'wet_cake_density',
    'Wet cake density': 'wet_cake_density',
    '振实密度(g/cm3)': 'bulk_density',
    '振实密度（g/cm3)': 'bulk_density',
    '振实密度（g/cm3）': 'bulk_density',
    '振实密度': 'bulk_density',
    '重度 (k) 14W': 'bulk_density',
    '灰值 (c/m)': 'bulk_density',
    
    # 白度
    '白度': 'brightness', 
    'Bri.': 'brightness', 
    'Brightness': 'brightness',
    
    '气味': 'odor', 
    'Odor': 'odor',
    
    # 电导值和pH
    '电导值 (as/c pH': 'conductance',
    '电导值(ms/cm)': 'conductance', 
    '电导值(ms/c pll': 'conductance',
    '电导值': 'conductance', 
    'Conductance (ms/c)': 'conductance', 
    'Conductance': 'conductance',
    'pH': 'ph', 
    'pH值': 'ph',
    
    '水分(%)': 'moisture', 
    '水分': 'moisture', 
    'Moisture (%)': 'moisture', 
    'Moisture': 'moisture',
    
    '批数': 'bags',
    '袋数': 'bags', 
    '烧数': 'bags',
    'Bags': 'bags',
    
    '吨': 'tons', 
    '吨数': 'tons',
    'Tons': 'tons', 
    '产量': 'tons',
    
    # 筛分数据字段映射
    '14W': 'sieving_14m',
    '+14M (%)': 'sieving_14m', 
    '+14M': 'sieving_14m', 
    '14M': 'sieving_14m',
    '+30M (%)': 'sieving_30m', 
    '+30M': 'sieving_30m', 
    '30M': 'sieving_30m',
    '+40M (%)': 'sieving_40m', 
    '+40M': 'sieving_40m', 
    '40M': 'sieving_40m',
    'M': 'sieving_40m',
    '+80M (%)': 'sieving_80m', 
    '+80M': 'sieving_80m', 
    '80M': 'sieving_80m',
    '+100M (%)': 'sieving_100m', 
    '+100M': 'sieving_100m', 
    '100M': 'sieving_100m',
    '+150M (%)': 'sieving_150m', 
    '+150M': 'sieving_150m',
    '150M': 'sieving_150m',
    '150M ': 'sieving_150m',
    '+200M (%)': 'sieving_200m', 
    '+200M': 'sieving_200m', 
    '200M': 'sieving_200m',
    '200M ': 'sieving_200m',
    '+325M (%)': 'sieving_325m', 
    '+325M': 'sieving_325m', 
    '325M': 'sieving_325m',
    '325M ': 'sieving_325m',
    
    # 离子数据字段映射
    '铁离子 (mg/钙离子 (mg/铝离子 (mg/白度': 'fe_ion',
    '铁离子（mg/kg）': 'fe_ion',
    '铁离子(mg/kg)': 'fe_ion',
    '铁离子(mg/': 'fe_ion',
    '铁离子': 'fe_ion', 
    'Fe离子': 'fe_ion', 
    'Fe': 'fe_ion',
    '钙离子（mg/kg）': 'ca_ion',
    '钙离子(mg/kg)': 'ca_ion',
    '钙离子(mg/': 'ca_ion',
    '钙离子': 'ca_ion',
    'Ca离子': 'ca_ion', 
    'Ca': 'ca_ion',
    '铝离子（mg/kg）': 'al_ion',
    '铝离子(mg/kg)': 'al_ion',
    '铝离子(mg/': 'al_ion',
    '铝离子': 'al_ion',
    'Al离子': 'al_ion', 
    'Al': 'al_ion',
    
    '吸油率 (%)': 'oil_absorption',
    '吸油率（%）': 'oil_absorption',
    '吸油量': 'oil_absorption', 
    '吸油率(%)': 'oil_absorption',
    '吸水率 (%)': 'water_absorption',
    '吸水率（%）': 'water_absorption',
    '吸水量': 'water_absorption', 
    '吸水率(%)': 'water_absorption',
    
    '备注': 'remarks', 
    'Remarks': 'remarks', 
    'Notes': 'remarks'
}

# 字段显示名称映射
XINGHUI_FIELD_DISPLAY_NAMES = {
    'date': '日期',
    'time': '时间',
    'shift': '班次',
    'product_name': '产品名称',
    'packaging': '包装类型',
    'batch_number': '批号',
    'moisture_after_drying': '干燥后原土水分(%)',
    'alkali_content': '入窑前碱含量(%)',
    'flux': '助溶剂添加比例',
    'permeability': '远通渗透率(Darcy)',
    'permeability_long': '长富渗透率(Darcy)',
    'xinghui_permeability': '兴辉渗透率(Darcy)',
    'wet_cake_density': '饼密度(g/cm3)',
    'filter_time': '过滤时间(秒)',
    'water_viscosity': '水黏度(mPa.s)',
    'cake_thickness': '饼厚(mm)',
    'bulk_density': '振实密度(g/cm3)',
    'brightness': '白度',
    'swirl': '涡值(cm)',
    'odor': '气味',
    'conductance': '电导值(ms/cm)',
    'ph': 'pH',
    'moisture': '水分(%)',
    'bags': '袋数',
    'tons': '吨',
    'sieving_14m': '+14M (%)',
    'sieving_30m': '+30M (%)',
    'sieving_40m': '+40M (%)',
    'sieving_80m': '+80M (%)',
    'sieving_100m': '+100M (%)',
    'sieving_150m': '+150M (%)',
    'sieving_200m': '+200M (%)',
    'sieving_325m': '+325M (%)',
    'fe_ion': 'Fe离子',
    'ca_ion': 'Ca离子',
    'al_ion': 'Al离子',
    'oil_absorption': '吸油量',
    'water_absorption': '吸水量',
    'remarks': '备注',
}


# ==================== 通用Excel读取函数 ====================

def read_excel_file(excel_file):
    """
    读取Excel文件，返回DataFrame或类似对象
    
    Args:
        excel_file: Django上传的文件对象
        
    Returns:
        tuple: (df, use_pandas) - DataFrame对象和是否使用pandas的标志
    """
    # 尝试导入pandas，如果失败则使用openpyxl
    try:
        import pandas as pd
        use_pandas = True
    except ImportError:
        use_pandas = False
    
    # 读取Excel文件
    if use_pandas:
        try:
            df = pd.read_excel(excel_file, sheet_name=0)
            # 删除完全空白的行
            df = df.dropna(how='all')
            
            # 修复：处理列名与pandas方法名冲突的问题（如"shift"）
            conflict_names = {'shift': '_shift_field_', 'date': '_date_field_', 'time': '_time_field_'}
            rename_conflicts = {}
            for col in df.columns:
                if col in conflict_names:
                    rename_conflicts[col] = conflict_names[col]
            
            if rename_conflicts:
                df = df.rename(columns=rename_conflicts)
                logger.info(f'⚠️ 重命名冲突列名: {rename_conflicts}')
            
            return df, use_pandas
        except Exception as e:
            raise Exception(f'读取Excel文件失败: {str(e)}')
    else:
        # 使用openpyxl读取
        from openpyxl import load_workbook
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value else '')
            
            # 处理特殊列名情况
            # 1. 处理重复的列名（如三个渗透率列）
            permeability_indices = []
            for i, header in enumerate(headers):
                if header and ('Permeability' in str(header) or '渗透率' in str(header) or '滤速率' in str(header)):
                    permeability_indices.append(i)
            
            # 如果有三个渗透率列，重命名它们
            if len(permeability_indices) >= 3:
                headers[permeability_indices[0]] = 'Permeability_1'
                headers[permeability_indices[1]] = 'Permeability_2'
                headers[permeability_indices[2]] = 'Permeability_3'
            elif len(permeability_indices) == 2:
                headers[permeability_indices[0]] = 'Permeability_1'
                headers[permeability_indices[1]] = 'Permeability_2'
            
            # 2. 处理"入窑前"和"含量 (%)"合并的情况
            for i, header in enumerate(headers):
                if header and '入窑前' in str(header):
                    if i + 1 < len(headers) and headers[i + 1] and '含量' in str(headers[i + 1]):
                        headers[i + 1] = '入窑前碱含量(%)_content'
                    break
            
            # 3. 处理合并列"铁离子 (mg/钙离子 (mg/铝离子 (mg/白度"
            for i, header in enumerate(headers):
                if header and '铁离子' in str(header) and '钙离子' in str(header):
                    headers[i] = '铁离子_钙离子_铝离子_白度_合并列'
                    break
            
            # 4. 处理合并列"电导值 (as/c pH"
            for i, header in enumerate(headers):
                if header and '电导值' in str(header) and 'pH' in str(header):
                    headers[i] = '电导值_pH_合并列'
                    break
            
            # 读取数据行，跳过完全空白的行
            df_data = []
            logger.info(f'📋 Excel原始列名: {headers}')
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            val = row[i]
                            if val is None or (isinstance(val, str) and val.strip() == ''):
                                row_dict[header] = None
                            else:
                                row_dict[header] = val
                        else:
                            row_dict[header] = None
                    df_data.append(row_dict)
            
            # 创建一个简单的DataFrame模拟对象
            class SimpleDF:
                def __init__(self, data, columns):
                    self.data = data
                    self.columns = columns
                
                def iterrows(self):
                    for idx, row_dict in enumerate(self.data):
                        class Row:
                            def __init__(self, data):
                                self._data = data
                            
                            def get(self, key, default=None):
                                return self._data.get(key, default)
                            
                            def __getitem__(self, key):
                                return self._data.get(key)
                        
                        yield idx, Row(row_dict)
                
                def __len__(self):
                    return len(self.data)
            
            df = SimpleDF(df_data, headers)
            return df, use_pandas
        except Exception as e:
            raise Exception(f'读取Excel文件失败: {str(e)}')


# ==================== 列名映射函数 ====================

def normalize_col_name(col_name):
    """规范化列名"""
    if col_name is None:
        return ''
    col_str = str(col_name).strip()
    col_str = col_str.replace('（', '(').replace('）', ')')
    col_str = ' '.join(col_str.split())
    return col_str


def map_excel_columns(df, column_mapping, use_pandas):
    """
    映射Excel列名到模型字段名
    
    Args:
        df: DataFrame对象
        column_mapping: 列名映射字典
        use_pandas: 是否使用pandas
        
    Returns:
        DataFrame: 映射后的DataFrame
    """
    if use_pandas and hasattr(df, 'rename'):
        # 处理重复的列名（如三个Permeability）
        permeability_cols = [col for col in df.columns if col and ('Permeability' in str(col) or '渗透率' in str(col) or '滤速率' in str(col))]
        if len(permeability_cols) >= 3:
            df.columns = [f'Permeability_1' if col == permeability_cols[0] else 
                         f'Permeability_2' if col == permeability_cols[1] else 
                         f'Permeability_3' if col == permeability_cols[2] else col 
                         for col in df.columns]
        elif len(permeability_cols) == 2:
            df.columns = [f'Permeability_1' if col == permeability_cols[0] else 
                         f'Permeability_2' if col == permeability_cols[1] else col 
                         for col in df.columns]
        
        # 添加Permeability映射
        column_mapping['Permeability_1'] = 'permeability'
        column_mapping['Permeability_2'] = 'permeability_long'
        column_mapping['Permeability_3'] = 'xinghui_permeability'
        
        # 修复：处理冲突列名的映射（如_shift_field_映射回shift）
        conflict_mapping = {'_shift_field_': 'shift', '_date_field_': 'date', '_time_field_': 'time'}
        for conflict_col, mapped_field in conflict_mapping.items():
            if conflict_col in df.columns:
                column_mapping[conflict_col] = mapped_field
        
        # 规范化映射
        normalized_mapping = {}
        for orig_key, mapped_value in column_mapping.items():
            normalized_key = normalize_col_name(orig_key)
            normalized_mapping[normalized_key] = mapped_value
        
        rename_dict = {}
        unmapped_cols = []
        
        for col in df.columns:
            col_normalized = normalize_col_name(col)
            if col in column_mapping:
                rename_dict[col] = column_mapping[col]
            elif col_normalized in normalized_mapping:
                rename_dict[col] = normalized_mapping[col_normalized]
            else:
                matched = False
                for orig_key, mapped_value in column_mapping.items():
                    orig_normalized = normalize_col_name(orig_key)
                    if orig_normalized and col_normalized:
                        orig_simple = orig_normalized.replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').lower()
                        col_simple = col_normalized.replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').lower()
                        if orig_simple == col_simple or (len(orig_simple) > 3 and orig_simple in col_simple):
                            rename_dict[col] = mapped_value
                            matched = True
                            break
                
                # 特殊处理"LOT批号"
                if not matched:
                    col_lower = col_normalized.lower()
                    if 'lot' in col_lower and '批号' in col_normalized:
                        rename_dict[col] = 'batch_number'
                        matched = True
                    elif '批号' in col_normalized and ('lot' in col_lower or col == 'LOT批号'):
                        rename_dict[col] = 'batch_number'
                        matched = True
                
                if not matched:
                    unmapped_cols.append(col)
        
        df_mapped = df.rename(columns=rename_dict)
        
        if unmapped_cols:
            logger.debug(f'⚠️ 未映射的列名: {unmapped_cols}')
        
        logger.info(f'📊 读取到 {len(df_mapped)} 行数据，原始列名: {list(df.columns)}, 映射后列名: {list(df_mapped.columns)}')
        return df_mapped
    else:
        # 对于openpyxl，列名映射在读取时已处理
        column_mapping['Permeability_1'] = 'permeability'
        column_mapping['Permeability_2'] = 'permeability_long'
        column_mapping['Permeability_3'] = 'xinghui_permeability'
        logger.info(f'📊 读取到 {len(df)} 行数据，列名: {df.columns if hasattr(df, "columns") else "未知"}')
        return df


# ==================== 数据提取函数 ====================

def get_row_value(row, key, use_pandas, column_mapping):
    """
    从行中获取值，支持pandas和openpyxl两种格式
    
    Args:
        row: 行对象
        key: 字段名
        use_pandas: 是否使用pandas
        column_mapping: 列名映射字典
        
    Returns:
        字段值或None
    """
    if use_pandas:
        try:
            import pandas as pd
            # 优先使用index访问
            if hasattr(row, 'index') and key in row.index:
                val = row[key]
                if callable(val):
                    logger.warning(f'字段 {key} 返回了方法对象，尝试其他方式获取')
                    val = None
                elif val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                    return val
            
            if hasattr(row, 'get'):
                val = row.get(key)
                if callable(val):
                    val = None
                elif val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                    return val
            
            if hasattr(row, key):
                val = getattr(row, key)
                if callable(val):
                    val = None
                elif val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                    return val
            
            # 模糊匹配逻辑
            field_keywords = {
                'date': ['日期', 'date', 'Date', '检测日期'],
                'time': ['时间', 'time', 'Time', '检测时间'],
                'shift': ['班次', 'shift', 'Shift', 'Squad', '班组', '班别', '_shift_field_'],
                'batch_number': ['批号', 'batch', 'Batch', 'LOT', 'lot', '批次号', '批次', 'LOT批号'],
                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                'permeability': ['远通', '迈通', '运通', '渗透率', 'permeability', 'Darcy', 'Permeability_1'],
                'permeability_long': ['长富', '长高', '渗透率', 'permeability', 'Permeability_2'],
                'xinghui_permeability': ['兴辉', '兴桐', 'Permeability_3'],
                'wet_cake_density': ['饼密度', '筛密度', 'cake', 'density'],
                'bulk_density': ['振实密度', 'bulk', 'density'],
                'brightness': ['白度', 'brightness', 'Bri'],
                'swirl': ['涡值', '滤值', '可塑度', 'swirl'],
                'odor': ['气味', 'odor'],
                'conductance': ['电导值', 'conductance'],
                'ph': ['pH', 'ph'],
                'moisture': ['水分', 'moisture'],
                'bags': ['袋数', '烧数', 'bags', '批数'],
                'tons': ['吨', 'tons', '产量'],
                'fe_ion': ['铁离子', 'Fe', 'fe'],
                'ca_ion': ['钙离子', 'Ca', 'ca'],
                'al_ion': ['铝离子', 'Al', 'al'],
                'oil_absorption': ['吸油', 'oil'],
                'water_absorption': ['吸水', 'water'],
                'sieving_14m': ['14M', '14', '+14'],
                'sieving_30m': ['30M', '30', '+30'],
                'sieving_40m': ['40M', '40', '+40'],
                'sieving_80m': ['80M', '80', '+80'],
                'sieving_100m': ['100M', '100', '+100'],
                'sieving_150m': ['150M', '150', '+150'],
                'sieving_200m': ['200M', '200', '+200'],
                'sieving_325m': ['325M', '325', '+325'],
            }
            
            keywords = field_keywords.get(key, [])
            
            if hasattr(row, 'index'):
                sorted_keywords = sorted(keywords, key=lambda k: (k.isascii(), -len(k)))
                
                for keyword in sorted_keywords:
                    best_match_col = None
                    best_match_val = None
                    
                    for col_name in row.index:
                        if col_name is None:
                            continue
                        col_str = str(col_name).strip()
                        
                        if keyword in col_str:
                            # 特殊处理：对于xinghui_permeability，必须确保列名包含"兴辉"或"兴桐"
                            if key == 'xinghui_permeability':
                                if '兴辉' not in col_str and '兴桐' not in col_str and 'Permeability_3' not in col_str:
                                    continue
                            
                            try:
                                val = row[col_name]
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    if best_match_col is None:
                                        best_match_col = col_name
                                        best_match_val = val
                                    elif len(col_str) < len(str(best_match_col)):
                                        best_match_col = col_name
                                        best_match_val = val
                            except (KeyError, IndexError):
                                continue
                    
                    if best_match_val is not None:
                        return best_match_val
                    
                    if best_match_col is not None:
                        return None
            
            return None
        except Exception as e:
            logger.debug(f'获取字段 {key} 失败: {str(e)}')
            return None
    else:
        # openpyxl格式处理
        if isinstance(row, dict):
            if key in row:
                val = row[key]
                if val is not None:
                    return val
        
        matched_cols = []
        for orig_col, mapped_col in column_mapping.items():
            if mapped_col == key:
                matched_cols.append(orig_col)
        
        for orig_col in matched_cols:
            if isinstance(row, dict) and orig_col in row:
                val = row[orig_col]
                if val is not None:
                    return val
        
        if isinstance(row, dict):
            key_keywords = {
                'batch_number': ['批号', 'batch', 'Batch', 'LOT', 'lot', '批次号', '批次', 'LOT批号'],
                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                'permeability': ['远通', '迈通', '运通', '渗透率', 'permeability', 'Darcy'],
                'permeability_long': ['长富', '长高', '渗透率', 'permeability'],
                'xinghui_permeability': ['兴辉', '兴桐'],
                'swirl': ['涡值', '滤值', '可塑度', 'swirl'],
                'sieving_150m': ['150M', '150', '+150'],
                'sieving_200m': ['200M', '200', '+200'],
                'sieving_325m': ['325M', '325', '+325'],
                'fe_ion': ['铁离子', 'Fe', 'fe'],
                'ca_ion': ['钙离子', 'Ca', 'ca'],
                'al_ion': ['铝离子', 'Al', 'al'],
                'oil_absorption': ['吸油', 'oil'],
                'water_absorption': ['吸水', 'water'],
                'moisture': ['水分', 'moisture'],
            }
            
            keywords = key_keywords.get(key, [])
            
            for orig_col in row.keys():
                if orig_col is None:
                    continue
                orig_col_str = str(orig_col).strip()
                
                if keywords:
                    for keyword in keywords:
                        if keyword in orig_col_str:
                            if key == 'xinghui_permeability':
                                if '兴辉' not in orig_col_str and '兴桐' not in orig_col_str:
                                    continue
                            
                            val = row[orig_col]
                            if val is not None:
                                return val
                else:
                    key_normalized = str(key).strip().replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                    orig_col_normalized = orig_col_str.replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                    if key_normalized in orig_col_normalized or orig_col_normalized in key_normalized:
                        val = row[orig_col]
                        if val is not None:
                            return val
        
        return None


def is_notna(val, use_pandas):
    """检查值是否不为空"""
    if use_pandas:
        import pandas as pd
        return pd.notna(val)
    else:
        return val is not None and val != '' and str(val).strip() != ''


# ==================== 数据验证函数 ====================

def is_hint_row(date_val, use_pandas):
    """检查是否是提示信息行"""
    hint_keywords = ['说明', '提示', '注意', '请删除', '示例', 'hint', 'note', '说明：']
    if date_val and is_notna(date_val, use_pandas):
        date_str = str(date_val).strip()
        for keyword in hint_keywords:
            if keyword in date_str:
                return True
    return False


def has_valid_data(row, get_row_value_func, use_pandas):
    """检查行是否有有效数据"""
    date_val = get_row_value_func('date')
    product_name_val = get_row_value_func('product_name')
    
    has_any_data = False
    if date_val and is_notna(date_val, use_pandas):
        has_any_data = True
    if product_name_val and is_notna(product_name_val, use_pandas):
        has_any_data = True
    
    if not has_any_data:
        for key in ['shift', 'packaging', 'bags', 'batch_number', 'moisture_after_drying', 
                   'alkali_content', 'permeability', 'permeability_long', 'xinghui_permeability', 'wet_cake_density']:
            val = get_row_value_func(key)
            if val and is_notna(val, use_pandas):
                has_any_data = True
                break
    
    return has_any_data


# ==================== 数据处理函数 ====================

def process_date_value(date_val, use_pandas):
    """处理日期值"""
    if date_val and is_notna(date_val, use_pandas):
        if isinstance(date_val, str):
            try:
                return datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
            except:
                try:
                    return datetime.strptime(date_val.strip(), '%Y/%m/%d').date()
                except:
                    return date.today()
        elif hasattr(date_val, 'date'):
            return date_val.date()
        else:
            return date.today()
    return None


def process_time_value(time_val, use_pandas):
    """处理时间值"""
    if time_val and is_notna(time_val, use_pandas):
        try:
            if isinstance(time_val, str):
                time_str = str(time_val).strip()
                if not time_str:
                    return time(0, 0)
                else:
                    try:
                        return datetime.strptime(time_str, '%H:%M').time()
                    except ValueError:
                        try:
                            return datetime.strptime(time_str, '%H:%M:%S').time()
                        except ValueError:
                            try:
                                if ':' not in time_str and ('.' in time_str or time_str.replace('.', '').replace('-', '').isdigit()):
                                    time_float = float(time_str)
                                    total_seconds = int(time_float * 24 * 3600)
                                    hours = total_seconds // 3600
                                    minutes = (total_seconds % 3600) // 60
                                    seconds = total_seconds % 60
                                    return time(hours, minutes, seconds)
                                else:
                                    if time_str.isdigit() and len(time_str) <= 4:
                                        hours = int(time_str[:2]) if len(time_str) >= 2 else int(time_str[0])
                                        minutes = int(time_str[-2:]) if len(time_str) >= 2 else 0
                                        return time(hours, minutes)
                                    else:
                                        return time(0, 0)
                            except (ValueError, TypeError):
                                return time(0, 0)
            elif hasattr(time_val, 'time'):
                return time_val.time()
            elif isinstance(time_val, time):
                return time_val
            elif hasattr(time_val, 'hour') and hasattr(time_val, 'minute'):
                return time(time_val.hour, time_val.minute, getattr(time_val, 'second', 0))
            elif isinstance(time_val, (int, float)):
                total_seconds = int(time_val * 24 * 3600)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return time(hours, minutes, seconds)
            else:
                return time(0, 0)
        except Exception as e:
            return time(0, 0)
    return time(0, 0)


def process_numeric_value(val, use_pandas):
    """处理数字值"""
    if val is not None and is_notna(val, use_pandas):
        try:
            import pandas as pd
            if isinstance(val, str):
                val = val.strip()
                if val == '' or val.isspace():
                    return None
                else:
                    float_val = float(val)
                    if pd.isna(float_val) if hasattr(pd, 'isna') else (float_val != float_val or abs(float_val) == float('inf')):
                        return None
                    else:
                        return float_val
            elif isinstance(val, (int, float)):
                if pd.isna(val) if hasattr(pd, 'isna') else (val != val or abs(val) == float('inf')):
                    return None
                else:
                    return float(val)
            else:
                return float(val)
        except (ValueError, TypeError):
            return None
    return None


# ==================== 通用导入函数 ====================

def import_xinghui_report_data(request, model_class, module_name, log_module_code):
    """
    通用的兴辉报表导入函数
    
    Args:
        request: Django请求对象
        model_class: 模型类（XinghuiQCReport或Xinghui2QCReport）
        module_name: 模块名称（用于错误提示）
        log_module_code: 日志模块代码（'xinghui'或'xinghui2'）
        
    Returns:
        JsonResponse: 导入结果
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 检查是否有上传的文件
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': '请选择要导入的Excel文件'}, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': '仅支持Excel文件格式(.xlsx, .xls)'}, status=400)
        
        # 读取Excel文件
        df, use_pandas = read_excel_file(excel_file)
        
        # 映射列名
        column_mapping = XINGHUI_COLUMN_MAPPING.copy()
        df_mapped = map_excel_columns(df, column_mapping, use_pandas)
        
        # 处理数据并导入
        imported_count = 0
        error_count = 0
        error_messages = []
        skipped_count = 0
        
        # 导入validate_field_by_model函数
        from home.utils.validators import validate_field_by_model
        
        # 处理每一行数据
        for index, row_obj in df_mapped.iterrows():
            try:
                row = row_obj
                
                # 创建get_row_value的闭包函数
                def get_row_value_func(key):
                    return get_row_value(row, key, use_pandas, column_mapping)
                
                # 检查是否是提示信息行
                date_val = get_row_value_func('date')
                if is_hint_row(date_val, use_pandas):
                    skipped_count += 1
                    logger.debug(f'跳过第 {index + 2} 行：提示信息行')
                    continue
                
                # 检查必填字段：date和product_name
                product_name_val = get_row_value_func('product_name')
                if not (date_val and is_notna(date_val, use_pandas)):
                    error_count += 1
                    error_msg = f'第 {index + 2} 行缺少必填字段: 日期'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                if not (product_name_val and is_notna(product_name_val, use_pandas)):
                    error_count += 1
                    error_msg = f'第 {index + 2} 行缺少必填字段: 产品名称'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                # 处理日期和时间
                date_obj = process_date_value(date_val, use_pandas)
                if not date_obj:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行日期格式错误'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                time_val = get_row_value_func('time')
                time_obj = process_time_value(time_val, use_pandas)
                
                # 构建数据字典
                data = {
                    'date': date_obj,
                    'time': time_obj,
                    'product_name': str(product_name_val).strip(),
                }
                
                # 处理字符串字段
                string_fields = ['shift', 'packaging', 'batch_number', 'flux', 'remarks']
                for field in string_fields:
                    val = get_row_value_func(field)
                    if val and is_notna(val, use_pandas):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 处理入窑前碱含量
                alkali_val = get_row_value_func('alkali_content')
                data['alkali_content'] = process_numeric_value(alkali_val, use_pandas)
                
                # 处理数字字段
                numeric_fields = [
                    'moisture_after_drying', 'permeability', 'permeability_long', 'xinghui_permeability',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor',
                    'conductance', 'ph', 'moisture', 'bags', 'tons', 'fe_ion', 'ca_ion',
                    'al_ion', 'oil_absorption', 'water_absorption', 'sieving_14m', 'sieving_30m',
                    'sieving_40m', 'sieving_80m'
                ]
                
                for field in numeric_fields:
                    val = get_row_value_func(field)
                    data[field] = process_numeric_value(val, use_pandas)
                
                # 处理合并列（仅openpyxl）
                if not use_pandas:
                    # 处理铁离子/钙离子/铝离子/白度合并列
                    merged_ion_col = None
                    if hasattr(row, 'get'):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    elif isinstance(row, dict):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    
                    if merged_ion_col and is_notna(merged_ion_col, use_pandas):
                        merged_str = str(merged_ion_col)
                        if '/' in merged_str:
                            parts = merged_str.split('/')
                            if len(parts) >= 3:
                                try:
                                    if parts[0].strip():
                                        data['fe_ion'] = float(parts[0].strip())
                                    if parts[1].strip():
                                        data['ca_ion'] = float(parts[1].strip())
                                    if parts[2].strip():
                                        data['al_ion'] = float(parts[2].strip())
                                    if len(parts) >= 4 and parts[3].strip():
                                        data['brightness'] = float(parts[3].strip())
                                except:
                                    pass
                    
                    # 处理电导值/pH合并列
                    merged_conductance_col = None
                    if hasattr(row, 'get'):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    elif isinstance(row, dict):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    
                    if merged_conductance_col and is_notna(merged_conductance_col, use_pandas):
                        merged_str = str(merged_conductance_col)
                        if '/' in merged_str or 'pH' in merged_str:
                            parts = merged_str.replace('pH', '').split('/')
                            if len(parts) >= 1 and parts[0].strip():
                                try:
                                    data['conductance'] = float(parts[0].strip())
                                except:
                                    pass
                            import re
                            ph_match = re.search(r'pH[:\s]*([0-9.]+)', merged_str, re.IGNORECASE)
                            if ph_match:
                                try:
                                    data['ph'] = float(ph_match.group(1))
                                except:
                                    pass
                
                # 处理筛分字段（可能是字符串）
                sieving_fields = ['sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m']
                for field in sieving_fields:
                    val = get_row_value_func(field)
                    if val and is_notna(val, use_pandas):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 数据校验
                validation_errors = []
                field_display_names = XINGHUI_FIELD_DISPLAY_NAMES.copy()
                
                for field_name, field_value in data.items():
                    if field_name in ['user', 'username']:
                        continue
                    
                    field_display_name = field_display_names.get(field_name, field_name)
                    is_valid, error_msg = validate_field_by_model(
                        model_class, 
                        field_name, 
                        field_value, 
                        field_display_name
                    )
                    
                    if not is_valid:
                        validation_errors.append(f'{field_display_name}: {error_msg}')
                
                if validation_errors:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行数据校验失败: {"; ".join(validation_errors)}'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}数据校验失败: {error_msg}')
                    continue
                
                # 设置用户信息
                data['user'] = request.user
                data['username'] = request.user.username
                
                # 创建记录
                model_class.objects.create(**data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                error_msg = f'第 {index + 2} 行导入失败: {str(e)}'
                error_messages.append(error_msg)
                logger.error(f'导入{module_name}失败: {error_msg}', exc_info=True)
        
        # 记录操作日志
        from home.models import UserOperationLog
        UserOperationLog.log_operation(
            request, 'CREATE', log_module_code, None,
            f'批量导入Excel数据: 成功{imported_count}条, 失败{error_count}条'
        )
        
        result = {
            'status': 'success',
            'message': f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条空行，失败 {error_count} 条',
            'imported_count': imported_count,
            'error_count': error_count,
            'skipped_count': skipped_count
        }
        
        logger.info(f'📊 导入统计: 成功 {imported_count} 条，跳过 {skipped_count} 条，失败 {error_count} 条')
        
        if error_messages:
            result['error_messages'] = error_messages[:10]
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f'导入{module_name}失败: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'导入失败: {str(e)}'}, status=500)


# ==================== 远通报表列名映射配置 ====================

YUANTONG_COLUMN_MAPPING = {
    # 基本字段映射
    '日期': 'date', 'Date': 'date', '检测日期': 'date',
    '时间': 'time', 'Time': 'time', '检测时间': 'time',
    '班次': 'shift', 'Shift': 'shift', 'Squad': 'shift', '班组': 'shift', '班别': 'shift',
    '产品名称': 'product_name', 'Product Name': 'product_name', 'Grade': 'product_name', 
    '产品型号': 'product_name',
    '包装类型': 'packaging', 'Packaging': 'packaging', 'IPKP CODE': 'packaging',
    '批号': 'batch_number', 'Batch Number': 'batch_number', '批次号': 'batch_number', 
    'LOT': 'batch_number', 'LOT批号': 'batch_number', '批号/日期': 'batch_number', '批次': 'batch_number',
    '物料类型': 'material_type', 'Material Type': 'material_type',
    
    # 检测数据字段映射
    '烘干后原土水分 (%)': 'moisture_after_drying',
    '烘干后原土水分（%）': 'moisture_after_drying',
    '烘干后原土水分(%)': 'moisture_after_drying',
    '干燥后原土水分(%)': 'moisture_after_drying',
    '干燥后原土水分（%）': 'moisture_after_drying',
    '烘干后原': 'moisture_after_drying',
    'Moisture after drying': 'moisture_after_drying', 
    '干燥后原土水分': 'moisture_after_drying',
    
    # 入窑前碱含量
    '入窑前碱含量(%)': 'alkali_content', 
    '入窑前碱含量': 'alkali_content',
    '入窑前': 'alkali_content',
    '含量 (%)': 'alkali_content',
    'Alkali content (%)': 'alkali_content',
    
    '助剂添加比例': 'flux',
    '助溶剂添加比例': 'flux', 
    '助溶剂': 'flux', 
    '助磨剂添加比例': 'flux',
    '*flux agent': 'flux', 
    'flux agent addition ratio': 'flux',
    
    # 远通特有字段
    '远通渗透率系数': 'yuantong_permeability_coefficient',
    '远通样品重量': 'yuantong_sample_weight',
    '远通过滤面积': 'yuantong_filter_area',
    
    # 渗透率
    '运通滤速率': 'permeability',
    '迈通渗透率': 'permeability',
    '远通渗透率(Darcy)': 'permeability', 
    '远通渗透率': 'permeability',
    'Permeability_1': 'permeability',
    '长高滤速率': 'permeability_long',
    '长富渗透率(Darcy)': 'permeability_long', 
    '长富渗透率': 'permeability_long',
    'Permeability_2': 'permeability_long',
    
    # 过滤相关
    '过滤时间(秒)': 'filter_time', '过滤时间': 'filter_time', 'Filter Time': 'filter_time',
    '水黏度(mPa.s)': 'water_viscosity', '水黏度': 'water_viscosity', 'Water Viscosity': 'water_viscosity',
    '饼厚(mm)': 'cake_thickness', '饼厚': 'cake_thickness', 'Cake Thickness': 'cake_thickness',
    
    # 饼密度
    '饼密度(g/cm3)': 'wet_cake_density', 
    '饼密度（g/cm3）': 'wet_cake_density',
    '饼密度': 'wet_cake_density', 
    '筛密度(g/cm3)': 'wet_cake_density',
    'Wet cake density': 'wet_cake_density',
    '远通饼密度(g/cm3)': 'yuantong_cake_density', 
    '远通饼密度（g/cm3）': 'yuantong_cake_density',
    '远通饼密度 (g/cm3)': 'yuantong_cake_density',
    '远通饼密度(g/cm³)': 'yuantong_cake_density',
    '远通饼密度（g/cm³）': 'yuantong_cake_density',
    '远通饼密度': 'yuantong_cake_density',
    '长富饼密度(g/cm3)': 'changfu_cake_density', 
    '长富饼密度（g/cm3）': 'changfu_cake_density',
    '长富饼密度 (g/cm3)': 'changfu_cake_density',
    '长富饼密度(g/cm³)': 'changfu_cake_density',
    '长富饼密度（g/cm³）': 'changfu_cake_density',
    '长富饼密度': 'changfu_cake_density',
    
    '振实密度(g/cm3)': 'bulk_density',
    '振实密度（g/cm3)': 'bulk_density',
    '振实密度（g/cm3）': 'bulk_density',
    '振实密度': 'bulk_density',
    '重度 (k) 14W': 'bulk_density',
    '灰值 (c/m)': 'bulk_density',
    
    # 白度
    '白度': 'brightness', 
    'Bri.': 'brightness', 
    'Brightness': 'brightness',
    
    # 可塑度可能是涡值
    '可塑度 (c/cm)': 'swirl',
    '可塑度': 'swirl',
    '涡值(cm)': 'swirl',
    '涡值（cm）': 'swirl',
    '滤值(cm)': 'swirl',
    '涡值': 'swirl', 
    'Swirl (cm)': 'swirl', 
    'Swirl': 'swirl',
    
    '气味': 'odor', 
    'Odor': 'odor',
    
    # 电导值和pH
    '电导值 (as/c pH': 'conductance',
    '电导值(ms/cm)': 'conductance', 
    '电导值(ms/c pll': 'conductance',
    '电导值': 'conductance', 
    'Conductance (ms/c)': 'conductance', 
    'Conductance': 'conductance',
    'pH': 'ph', 
    'pH值': 'ph',
    
    '水分(%)': 'moisture', 
    '水分': 'moisture', 
    'Moisture (%)': 'moisture', 
    'Moisture': 'moisture',
    
    '批数': 'bags',
    '袋数': 'bags', 
    '烧数': 'bags',
    'Bags': 'bags',
    
    '吨': 'tons', 
    '吨数': 'tons',
    'Tons': 'tons', 
    '产量': 'tons',
    
    # 筛分数据字段映射
    '14W': 'sieving_14m',
    '+14M (%)': 'sieving_14m', 
    '+14M': 'sieving_14m', 
    '14M': 'sieving_14m',
    '+30M (%)': 'sieving_30m', 
    '+30M': 'sieving_30m', 
    '30M': 'sieving_30m',
    '+40M (%)': 'sieving_40m', 
    '+40M': 'sieving_40m', 
    '40M': 'sieving_40m',
    'M': 'sieving_40m',
    '+80M (%)': 'sieving_80m', 
    '+80M': 'sieving_80m', 
    '80M': 'sieving_80m',
    '+100M (%)': 'sieving_100m', 
    '+100M': 'sieving_100m', 
    '100M': 'sieving_100m',
    '+150M (%)': 'sieving_150m', 
    '+150M': 'sieving_150m',
    '150M': 'sieving_150m',
    '150M ': 'sieving_150m',
    '+200M (%)': 'sieving_200m', 
    '+200M': 'sieving_200m', 
    '200M': 'sieving_200m',
    '200M ': 'sieving_200m',
    '+325M (%)': 'sieving_325m', 
    '+325M': 'sieving_325m', 
    '325M': 'sieving_325m',
    '325M ': 'sieving_325m',
    
    # 离子数据字段映射
    '铁离子 (mg/钙离子 (mg/铝离子 (mg/白度': 'fe_ion',
    '铁离子（mg/kg）': 'fe_ion',
    '铁离子(mg/kg)': 'fe_ion',
    '铁离子(mg/': 'fe_ion',
    '铁离子': 'fe_ion', 
    'Fe离子': 'fe_ion', 
    'Fe': 'fe_ion',
    '钙离子（mg/kg）': 'ca_ion',
    '钙离子(mg/kg)': 'ca_ion',
    '钙离子(mg/': 'ca_ion',
    '钙离子': 'ca_ion',
    'Ca离子': 'ca_ion', 
    'Ca': 'ca_ion',
    '铝离子（mg/kg）': 'al_ion',
    '铝离子(mg/kg)': 'al_ion',
    '铝离子(mg/': 'al_ion',
    '铝离子': 'al_ion',
    'Al离子': 'al_ion', 
    'Al': 'al_ion',
    
    '吸油率 (%)': 'oil_absorption',
    '吸油率（%）': 'oil_absorption',
    '吸油量': 'oil_absorption', 
    '吸油率(%)': 'oil_absorption',
    '吸水率 (%)': 'water_absorption',
    '吸水率（%）': 'water_absorption',
    '吸水量': 'water_absorption', 
    '吸水率(%)': 'water_absorption',
    
    '备注': 'remarks', 
    'Remarks': 'remarks', 
    'Notes': 'remarks'
}

# 远通字段显示名称映射
YUANTONG_FIELD_DISPLAY_NAMES = {
    'date': '日期',
    'time': '时间',
    'shift': '班次',
    'product_name': '产品名称',
    'packaging': '包装类型',
    'batch_number': '批号',
    'material_type': '物料类型',
    'moisture_after_drying': '干燥后原土水分(%)',
    'alkali_content': '入窑前碱含量(%)',
    'flux': '助溶剂添加比例',
    'yuantong_permeability_coefficient': '远通渗透率系数',
    'yuantong_sample_weight': '远通样品重量',
    'yuantong_filter_area': '远通过滤面积',
    'permeability': '远通渗透率(Darcy)',
    'permeability_long': '长富渗透率(Darcy)',
    'filter_time': '过滤时间(秒)',
    'water_viscosity': '水黏度(mPa.s)',
    'cake_thickness': '饼厚(mm)',
    'wet_cake_density': '饼密度(g/cm3)',
    'yuantong_cake_density': '远通饼密度(g/cm3)',
    'changfu_cake_density': '长富饼密度(g/cm3)',
    'bulk_density': '振实密度(g/cm3)',
    'brightness': '白度',
    'swirl': '涡值(cm)',
    'odor': '气味',
    'conductance': '电导值(ms/cm)',
    'ph': 'pH',
    'moisture': '水分(%)',
    'bags': '袋数',
    'tons': '吨数',
    'sieving_14m': '+14M',
    'sieving_30m': '+30M',
    'sieving_40m': '+40M',
    'sieving_80m': '+80M',
    'sieving_100m': '+100M',
    'sieving_150m': '+150M',
    'sieving_200m': '+200M',
    'sieving_325m': '+325M',
    'fe_ion': 'Fe离子',
    'ca_ion': 'Ca离子',
    'al_ion': 'Al离子',
    'oil_absorption': '吸油率(%)',
    'water_absorption': '吸水率(%)',
    'remarks': '备注',
}


def import_yuantong_report_data(request, model_class, module_name, log_module_code):
    """
    通用的远通报表导入函数
    
    Args:
        request: Django请求对象
        model_class: 模型类（YuantongQCReport或Yuantong2QCReport）
        module_name: 模块名称（用于错误提示）
        log_module_code: 日志模块代码（'yuantong'或'yuantong2'）
        
    Returns:
        JsonResponse: 导入结果
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 检查是否有上传的文件
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': '请选择要导入的Excel文件'}, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': '仅支持Excel文件格式(.xlsx, .xls)'}, status=400)
        
        # 读取Excel文件
        df, use_pandas = read_excel_file(excel_file)
        
        # 映射列名
        column_mapping = YUANTONG_COLUMN_MAPPING.copy()
        df_mapped = map_excel_columns(df, column_mapping, use_pandas)
        
        # 处理数据并导入
        imported_count = 0
        error_count = 0
        error_messages = []
        skipped_count = 0
        
        # 导入validate_field_by_model函数
        from home.utils.validators import validate_field_by_model
        
        # 处理每一行数据
        for index, row_obj in df_mapped.iterrows():
            try:
                row = row_obj
                
                # 创建get_row_value的闭包函数
                def get_row_value_func(key):
                    return get_row_value(row, key, use_pandas, column_mapping)
                
                # 检查是否是提示信息行
                date_val = get_row_value_func('date')
                if is_hint_row(date_val, use_pandas):
                    skipped_count += 1
                    logger.debug(f'跳过第 {index + 2} 行：提示信息行')
                    continue
                
                # 检查必填字段：date和product_name
                product_name_val = get_row_value_func('product_name')
                if not (date_val and is_notna(date_val, use_pandas)):
                    error_count += 1
                    error_msg = f'第 {index + 2} 行缺少必填字段: 日期'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                if not (product_name_val and is_notna(product_name_val, use_pandas)):
                    error_count += 1
                    error_msg = f'第 {index + 2} 行缺少必填字段: 产品名称'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                # 处理日期和时间
                date_obj = process_date_value(date_val, use_pandas)
                if not date_obj:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行日期格式错误'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}失败: {error_msg}')
                    continue
                
                time_val = get_row_value_func('time')
                time_obj = process_time_value(time_val, use_pandas)
                
                # 构建数据字典
                data = {
                    'date': date_obj,
                    'time': time_obj,
                    'product_name': str(product_name_val).strip(),
                }
                
                # 处理字符串字段
                string_fields = ['shift', 'packaging', 'batch_number', 'flux', 'remarks', 'material_type']
                for field in string_fields:
                    val = get_row_value_func(field)
                    if val and is_notna(val, use_pandas):
                        data[field] = str(val)
                    else:
                        if field == 'material_type':
                            data[field] = '助熔煅烧品'
                        else:
                            data[field] = ''
                
                # 处理入窑前碱含量
                alkali_val = get_row_value_func('alkali_content')
                data['alkali_content'] = process_numeric_value(alkali_val, use_pandas)
                
                # 处理数字字段
                numeric_fields = [
                    'moisture_after_drying', 'yuantong_permeability_coefficient', 'yuantong_sample_weight', 
                    'yuantong_filter_area', 'permeability', 'permeability_long', 'filter_time', 
                    'water_viscosity', 'cake_thickness', 'wet_cake_density', 'yuantong_cake_density', 
                    'changfu_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor',
                    'conductance', 'ph', 'moisture', 'bags', 'tons', 'fe_ion', 'ca_ion',
                    'al_ion', 'oil_absorption', 'water_absorption', 'sieving_14m', 'sieving_30m',
                    'sieving_40m', 'sieving_80m'
                ]
                
                for field in numeric_fields:
                    val = get_row_value_func(field)
                    data[field] = process_numeric_value(val, use_pandas)
                
                # 处理筛分字段（可能是字符串）
                sieving_fields = ['sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m']
                for field in sieving_fields:
                    val = get_row_value_func(field)
                    if val and is_notna(val, use_pandas):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 数据校验：检查必填字段
                validation_errors = []
                field_display_names = YUANTONG_FIELD_DISPLAY_NAMES.copy()
                
                # 检查所有必填字段（根据model定义）
                for field_name, field_value in data.items():
                    if field_name in ['user', 'username']:
                        continue
                    
                    field_display_name = field_display_names.get(field_name, field_name)
                    is_valid, error_msg = validate_field_by_model(
                        model_class, 
                        field_name, 
                        field_value, 
                        field_display_name
                    )
                    
                    if not is_valid:
                        validation_errors.append(f'{field_display_name}: {error_msg}')
                
                if validation_errors:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行数据校验失败: {"; ".join(validation_errors)}'
                    error_messages.append(error_msg)
                    logger.error(f'导入{module_name}数据校验失败: {error_msg}')
                    continue
                
                # 设置用户信息
                data['user'] = request.user
                data['username'] = request.user.username
                
                # 创建记录
                model_class.objects.create(**data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                error_msg = f'第 {index + 2} 行导入失败: {str(e)}'
                error_messages.append(error_msg)
                logger.error(f'导入{module_name}失败: {error_msg}', exc_info=True)
        
        # 记录操作日志
        from home.models import UserOperationLog
        UserOperationLog.log_operation(
            request, 'CREATE', log_module_code, None,
            f'批量导入Excel数据: 成功{imported_count}条, 失败{error_count}条'
        )
        
        result = {
            'status': 'success',
            'message': f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条空行，失败 {error_count} 条',
            'imported_count': imported_count,
            'error_count': error_count,
            'skipped_count': skipped_count
        }
        
        logger.info(f'📊 导入统计: 成功 {imported_count} 条，跳过 {skipped_count} 条，失败 {error_count} 条')
        
        if error_messages:
            result['error_messages'] = error_messages[:10]
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f'导入{module_name}失败: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'导入失败: {str(e)}'}, status=500)
