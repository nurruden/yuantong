"""
QC报表相关视图
包含所有QC报表的API类和导出函数
"""

# 导入必要的模块
from django.shortcuts import render, redirect
from django.views import View
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import os
import json
import logging
import urllib.parse
from datetime import datetime, date, time, timedelta

# 导入配置
from home.config import (
    MENU_ITEMS,
    QC_REPORT_FIELD_MAPPING,
    get_report_module_code,
    get_report_permission_code,
)

# 导入工具函数
from home.utils.excel_export import (
    export_production_excel,
    export_qc_report_excel,
    export_qc_report_excel_universal,
)
from home.utils.permissions import (
    user_has_permission,
    permission_required,
    filter_menu_by_permission,
)
from home.utils.user_helpers import (
    get_user_info,
)
from home.utils.validators import (
    validate_field_by_model,
)

# 导入模型
from home.models import (
    DongtaiQCReport,
    DayuanQCReport,
    XinghuiQCReport,
    ChangfuQCReport,
    YuantongQCReport,
    Yuantong2QCReport,
    Xinghui2QCReport,
    Parameter,
    UserOperationLog,
)

logger = logging.getLogger(__name__)

# ==================== QC报表通用基类和工具函数 ===================

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class BaseQCReportAPI(View):
    """QC报表API基类，提供通用的CRUD操作"""
    
    # 子类需要重写这些属性
    model_class = None  # 具体的模型类
    report_name = None  # 报表名称，用于日志和错误信息
    history_template = None  # 历史记录页面模板
    field_mapping = None  # Excel导出字段映射
    
    def get(self, request, report_id=None):
        """获取报表数据"""
        logger = logging.getLogger(__name__)
        
        # 检查用户是否有QC报表查看权限
        if not user_has_permission(request.user, 'qc_report_view'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_view'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            # 检查是否是昨日产量统计请求
            if request.GET.get('action') == 'yesterday_production':
                return self.calculate_yesterday_production(request)
            
            # 检查是否是今日产量统计请求
            if request.GET.get('action') == 'today_production':
                return self.calculate_today_production(request)
            
            # 只允许明确的页面渲染视图去渲染模板，API接口始终返回JSON
            if report_id:
                # 获取单个报表
                report = self.model_class.objects.get(id=report_id)
                data = self._serialize_report(report, request.user)
                return JsonResponse({'status': 'success', 'data': data})
            else:
                # 获取报表列表
                reports_query = self.model_class.objects.all().order_by('-date', '-time')
                reports_query = self._apply_filters(reports_query, request)
                
                # 分页处理
                page_number = request.GET.get('page', 1)
                page_size = request.GET.get('page_size', 10)
                paginator = Paginator(reports_query, page_size)
                page_obj = paginator.get_page(page_number)
                
                data = [self._serialize_report(report, request.user) for report in page_obj.object_list]
                
                return JsonResponse({
                    'status': 'success',
                    'data': data,
                    'total_pages': paginator.num_pages,
                    'total_count': paginator.count,
                    'current_page': page_obj.number
                })
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            logger.error(f'Error getting QC reports for {self.report_name}: {str(e)}', exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    def post(self, request, report_id=None):
        """创建新报表"""
        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            data = json.loads(request.body)
            data = self._process_input_data(data, request)
            report = self.model_class.objects.create(**data)
            
            # 记录操作日志
            self._log_operation(request, 'CREATE', report, data)
            return JsonResponse({
                'status': 'success',
                'message': '创建成功',
                'data': {
                    'id': report.id,
                    'date': report.date.strftime('%Y-%m-%d'),
                    'time': report.time.strftime('%H:%M'),
                    'username': report.username
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    def put(self, request, report_id):
        """更新报表"""
        logger = logging.getLogger('home')
        logger.info(f'开始处理 PUT 请求 - report_id: {report_id}')

        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)

        try:
            report = self.model_class.objects.get(id=report_id)
            data = json.loads(request.body)
            data = self._process_input_data(data, request)
            
            # 更新报表
            for key, value in data.items():
                if hasattr(report, key):
                    setattr(report, key, value)
            report.save()
            # 记录操作日志
            self._log_operation(request, 'UPDATE', report, data)
            
            return JsonResponse({
                'status': 'success',
                'message': '更新成功',
                'data': {
                    'id': report.id,
                    'date': report.date.strftime('%Y-%m-%d') if report.date else None,
                    'time': report.time.strftime('%H:%M') if report.time else None,
                    'username': report.username
                }
            })
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            logger.error(f'处理 PUT 请求时发生错误: {str(e)}', exc_info=True)
            return JsonResponse({'status': 'error', 'message': f'服务器错误: {str(e)}'}, status=500)

    def delete(self, request, report_id):
        """删除报表"""
        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            report = self.model_class.objects.get(id=report_id)
            
            # 检查是否是记录的操作人
            if report.username != request.user.username:
                return JsonResponse({
                    'status': 'error',
                    'message': '只能删除自己创建的记录'
                }, status=403)
            
            # 检查删除权限（基于日期限制）
            from datetime import datetime, timedelta
            
            # 获取编辑时间限制参数
            edit_limit = 7  # 默认7天
            try:
                param = Parameter.objects.filter(id='report_edit_limit').first()
                if param and param.value:
                    edit_limit = int(param.value)
            except:
                pass
            
            # 检查是否在编辑期限内
            if report.date:
                report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                days_diff = (datetime.now() - report_date).days
                
                if days_diff > edit_limit:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'该记录已超过{edit_limit}天编辑期限，无法删除'
                    }, status=403)
            
            # 记录操作日志
            self._log_operation(request, 'DELETE', report)
            report.delete()
            return JsonResponse({'status': 'success', 'message': '删除成功'})
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    def _serialize_report(self, report, current_user=None):
        """序列化报表数据"""
        user_info = get_user_info(report.username)
        
        # 计算权限状态
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d'),
            'time': report.time.strftime('%H:%M'),
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'xinghui_permeability': getattr(report, 'xinghui_permeability', None),
            'wet_cake_density': report.wet_cake_density,
            'filter_time': getattr(report, 'filter_time', None),
            'water_viscosity': getattr(report, 'water_viscosity', None),
            'cake_thickness': getattr(report, 'cake_thickness', None),
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'username': user_info.get('name', report.username),
            'original_username': report.username,  # 原始用户名，用于权限检查
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _apply_filters(self, queryset, request):
        """应用筛选条件"""
        from home.utils import get_user_data_filter_by_company_department
        
        # 应用数据权限过滤
        data_filter = get_user_data_filter_by_company_department(request.user, self.report_name, 'username')
        if data_filter:
            queryset = queryset.filter(**data_filter)
        
        # 应用其他筛选条件
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        product_name = request.GET.get('product_name')
        packaging = request.GET.get('packaging')
        shift = request.GET.get('shift')
        squad = request.GET.get('squad')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if start_time:
            queryset = queryset.filter(time__gte=start_time)
        if end_time:
            queryset = queryset.filter(time__lte=end_time)
        if product_name:
            queryset = queryset.filter(product_name__icontains=product_name)
        if packaging:
            queryset = queryset.filter(packaging__icontains=packaging)
        # 处理班组筛选：优先使用squad参数（前端表单使用），其次使用shift参数（兼容性）
        if squad:
            queryset = queryset.filter(shift__icontains=squad)
        elif shift:
            queryset = queryset.filter(shift__icontains=shift)

        return queryset

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用细粒度权限控制"""
        if not current_user:
            return False
        
        # 使用细粒度权限控制，支持跨用户编辑
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, self.report_name)
    
    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用细粒度权限控制"""
        if not current_user:
            return False
        
        # 使用细粒度权限控制，支持跨用户删除
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, self.report_name)
    
    def _check_date_limit(self, report):
        """检查日期限制"""
        from datetime import datetime
        
        # 获取编辑时间限制参数
        edit_limit = 7  # 默认7天
        try:
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            return days_diff <= edit_limit
        
        return True
    
    def _get_permission_reason(self, report, current_user):
        """获取权限限制原因"""
        if not current_user:
            return '未登录'
        
        # 检查编辑权限
        can_edit = self._check_edit_permission(report, current_user)
        if not can_edit:
            # 详细分析权限限制原因
            if report.username != current_user.username:
                # 检查是否启用跨用户编辑
                try:
                    from home.models import Parameter
                    cross_edit_param = Parameter.objects.filter(id='enable_cross_user_edit').first()
                    if not cross_edit_param or cross_edit_param.value != 'true':
                        return '系统未启用跨用户编辑功能'
                    
                    # 检查是否有跨用户编辑权限
                    from home.utils import has_hierarchical_permission
                    module_code = get_report_module_code(self.report_name)
                    
                    if not has_hierarchical_permission(current_user, f"{module_code}_edit_others"):
                        return '无跨用户编辑权限'
                    
                    return '无权限编辑他人数据'
                except:
                    return '无权限编辑他人数据'
            
            if not self._check_date_limit(report):
                from datetime import datetime
                edit_limit = 7
                try:
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                return f'超过{edit_limit}天编辑期限'
        
        return ''

    def _process_input_data(self, data, request):
        """处理输入数据"""
        # 处理日期和时间
        if 'date' in data:
            data['date'] = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if 'time' in data:
            data['time'] = datetime.strptime(data['time'], '%H:%M').time()

        # 添加用户信息
        data['user'] = request.user
        data['username'] = request.user.username

        return data

    def calculate_yesterday_production(self, request):
        """统计昨日产量 - 按班组、产品型号、包装类型、批号、备注分组统计吨数"""
        from datetime import date, timedelta
        
        try:
            # 获取昨天的日期
            yesterday = date.today() - timedelta(days=1)
            
            # 查询昨日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=yesterday,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'remarks',  # 备注
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number', 'remarks')
            
            # 按5个字段分组累加产量（与Excel导出逻辑保持一致）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置',
                    report['remarks'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'remarks': group_key[4],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'remarks': stat['remarks'] if stat['remarks'] is not None and stat['remarks'] != '' else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': yesterday.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': yesterday.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def calculate_today_production(self, request):
        """统计今日产量 - 按班组、产品型号、包装类型、批号、备注分组统计吨数"""
        from datetime import date
        
        try:
            # 获取今天的日期
            today = date.today()
            
            # 查询今日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=today,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'remarks',  # 备注
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number', 'remarks')
            
            # 按5个字段分组累加产量（与Excel导出逻辑保持一致）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置',
                    report['remarks'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'remarks': group_key[4],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'remarks': stat['remarks'] if stat['remarks'] is not None and stat['remarks'] != '' else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': today.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': today.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def _log_operation(self, request, operation_type, report, new_data=None):
        """记录操作日志"""
        import logging
        from home.models import UserOperationLog
        
        try:
            # 获取报表类型
            report_type = self._get_report_type()
            
            # 处理 new_data，确保所有对象都可以JSON序列化
            if new_data:
                # 将 new_data 中的日期/时间对象转换为字符串
                processed_data = {}
                for key, value in new_data.items():
                    if hasattr(value, 'strftime'):  # 日期或时间对象
                        processed_data[key] = str(value)
                    elif hasattr(value, '__dict__'):  # 复杂对象
                        processed_data[key] = str(value)
                    else:
                        processed_data[key] = value
                serialized_data = processed_data
            else:
                serialized_data = self._serialize_for_log(report)
            
            # 记录操作日志
            log_entry = UserOperationLog.log_operation(
                request=request,
                operation_type=operation_type,
                report_type=report_type,
                report_id=report.id,
                operation_detail=f'{operation_type} {self.report_name}，日期: {report.date}，班次: {report.shift}',
                new_data=serialized_data
            )
            
            # 如果日志记录成功，记录更多调试信息
            if log_entry:
                logger = logging.getLogger('user_operations')
                logger.info(f"操作日志记录成功 - ID: {log_entry.id}")
                
                # 记录客户端环境信息
                client_info = request.META.get('HTTP_USER_AGENT', '未知')
                logger.info(f"客户端环境: {client_info}")
                
                # 记录页面状态信息
                page_state = {
                    'currentPath': request.path,
                    'title': '未知',
                    'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                logger.info(f"页面状态: {json.dumps(page_state, ensure_ascii=False, indent=2)}")
                
                # 记录会话信息
                session_info = {
                    'hasCsrfToken': 'csrftoken' in request.COOKIES,
                    'hasSessionCookie': 'sessionid' in request.COOKIES,
                    'sessionCookieLength': len(request.COOKIES.get('sessionid', ''))
                }
                logger.info(f"会话信息: {json.dumps(session_info, ensure_ascii=False, indent=2)}")
                
                # 记录页面元素状态
                element_state = {
                    'hasFilterForm': False,
                    'hasDataContainer': False,
                    'filterFormElements': []
                }
                logger.info(f"页面元素状态: {json.dumps(element_state, ensure_ascii=False, indent=2)}")
                
                # 记录错误信息（如果有）
                if 'error' in request.GET:
                    logger.warning(f"操作过程中出现错误: {request.GET.get('error', '未知错误')}")
                    if 'errorDetail' in request.GET:
                        logger.warning(f"错误详情: {request.GET.get('errorDetail', '')}")
                    if 'errorStack' in request.GET:
                        logger.warning(f"错误堆栈: {request.GET.get('errorStack', '')}")
        
        except Exception as e:
            # 如果日志记录失败，记录到Django日志
            logger = logging.getLogger('home')
            logger.error(f'记录操作日志失败: {str(e)}', exc_info=True)

    def _get_report_type(self):
        """获取报表类型标识"""
        report_name = getattr(self, 'report_name', '')
        if '东泰' in report_name: return 'dongtai'
        elif '远通' in report_name and '二线' not in report_name: return 'yuantong'
        elif '远通二线' in report_name or '远通2号' in report_name: return 'yuantong2'
        elif '大塬' in report_name: return 'dayuan'
        elif '长富' in report_name: return 'changfu'
        elif '兴辉' in report_name and '二线' not in report_name: return 'xinghui'
        elif '兴辉二线' in report_name or '兴辉2号' in report_name: return 'xinghui2'
        else: return None

    def _serialize_for_log(self, report):
        """序列化报表数据用于日志记录"""
        try:
            return {
                'id': report.id, 'date': str(report.date) if report.date else None,
                'time': str(report.time) if report.time else None, 'shift': report.shift,
                'product_name': report.product_name, 'packaging': report.packaging,
                'batch_number': report.batch_number, 'moisture_after_drying': report.moisture_after_drying,
                'alkali_content': report.alkali_content, 'flux': report.flux,
                'permeability': report.permeability, 'permeability_long': report.permeability_long,
                'wet_cake_density': report.wet_cake_density, 'filter_time': getattr(report, 'filter_time', None),
                'water_viscosity': getattr(report, 'water_viscosity', None), 'cake_thickness': getattr(report, 'cake_thickness', None),
                'bulk_density': report.bulk_density, 'brightness': report.brightness,
                'swirl': report.swirl, 'odor': report.odor, 'conductance': report.conductance,
                'ph': report.ph, 'moisture': report.moisture, 'bags': report.bags, 'tons': report.tons,
                'fe_ion': report.fe_ion, 'ca_ion': report.ca_ion, 'al_ion': report.al_ion,
                'oil_absorption': report.oil_absorption, 'water_absorption': report.water_absorption,
                'remarks': report.remarks, 'username': report.username,
                'created_at': str(report.created_at) if report.created_at else None,
                'updated_at': str(report.updated_at) if report.updated_at else None,
            }
        except Exception as e:
            return {'error': f'序列化失败: {str(e)}'}



# ==================== QC报表API类和导出函数 ===================

class BaseQCReportAPI(View):
    """QC报表API基类，提供通用的CRUD操作"""
    
    # 子类需要重写这些属性
    model_class = None  # 具体的模型类
    report_name = None  # 报表名称，用于日志和错误信息
    history_template = None  # 历史记录页面模板
    field_mapping = None  # Excel导出字段映射
    
    def get(self, request, report_id=None):
        """获取报表数据"""
        logger = logging.getLogger(__name__)
        
        # 检查用户是否有QC报表查看权限
        if not user_has_permission(request.user, 'qc_report_view'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_view'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            # 检查是否是昨日产量统计请求
            if request.GET.get('action') == 'yesterday_production':
                return self.calculate_yesterday_production(request)
            
            # 检查是否是今日产量统计请求
            if request.GET.get('action') == 'today_production':
                return self.calculate_today_production(request)
            
            # 只允许明确的页面渲染视图去渲染模板，API接口始终返回JSON
            if report_id:
                # 获取单个报表
                report = self.model_class.objects.get(id=report_id)
                data = self._serialize_report(report, request.user)
                return JsonResponse({'status': 'success', 'data': data})
            else:
                # 获取报表列表
                reports_query = self.model_class.objects.all().order_by('-date', '-time')
                reports_query = self._apply_filters(reports_query, request)
                
                # 分页处理
                page_number = request.GET.get('page', 1)
                page_size = request.GET.get('page_size', 10)
                paginator = Paginator(reports_query, page_size)
                page_obj = paginator.get_page(page_number)
                
                data = [self._serialize_report(report, request.user) for report in page_obj.object_list]
                
                return JsonResponse({
                    'status': 'success',
                    'data': data,
                    'total_pages': paginator.num_pages,
                    'total_count': paginator.count,
                    'current_page': page_obj.number
                })
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            logger.error(f'Error getting QC reports for {self.report_name}: {str(e)}', exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    def post(self, request, report_id=None):
        """创建新报表"""
        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            data = json.loads(request.body)
            data = self._process_input_data(data, request)
            report = self.model_class.objects.create(**data)
            
            # 记录操作日志
            self._log_operation(request, 'CREATE', report, data)
            return JsonResponse({
                'status': 'success',
                'message': '创建成功',
                'data': {
                    'id': report.id,
                    'date': report.date.strftime('%Y-%m-%d'),
                    'time': report.time.strftime('%H:%M'),
                    'username': report.username
                }
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    def put(self, request, report_id):
        """更新报表"""
        logger = logging.getLogger('home')
        logger.info(f'开始处理 PUT 请求 - report_id: {report_id}')

        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)

        try:
            report = self.model_class.objects.get(id=report_id)
            data = json.loads(request.body)
            data = self._process_input_data(data, request)
            
            # 更新报表
            for key, value in data.items():
                if hasattr(report, key):
                    setattr(report, key, value)
            report.save()
            # 记录操作日志
            self._log_operation(request, 'UPDATE', report, data)
            
            return JsonResponse({
                'status': 'success',
                'message': '更新成功',
                'data': {
                    'id': report.id,
                    'date': report.date.strftime('%Y-%m-%d') if report.date else None,
                    'time': report.time.strftime('%H:%M') if report.time else None,
                    'username': report.username
                }
            })
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            logger.error(f'处理 PUT 请求时发生错误: {str(e)}', exc_info=True)
            return JsonResponse({'status': 'error', 'message': f'服务器错误: {str(e)}'}, status=500)

    def delete(self, request, report_id):
        """删除报表"""
        # 检查用户是否有QC报表编辑权限
        if not user_has_permission(request.user, 'qc_report_edit'):
            # 渲染美观的403错误页面
            from django.shortcuts import render
            context = {
                'user': request.user,
                'request_path': request.path,
                'permission_code': 'qc_report_edit'
            }
            return render(request, '403.html', context, status=403)
        
        try:
            report = self.model_class.objects.get(id=report_id)
            
            # 检查是否是记录的操作人
            if report.username != request.user.username:
                return JsonResponse({
                    'status': 'error',
                    'message': '只能删除自己创建的记录'
                }, status=403)
            
            # 检查删除权限（基于日期限制）
            from datetime import datetime, timedelta
            
            # 获取编辑时间限制参数
            edit_limit = 7  # 默认7天
            try:
                param = Parameter.objects.filter(id='report_edit_limit').first()
                if param and param.value:
                    edit_limit = int(param.value)
            except:
                pass
            
            # 检查是否在编辑期限内
            if report.date:
                report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                days_diff = (datetime.now() - report_date).days
                
                if days_diff > edit_limit:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'该记录已超过{edit_limit}天编辑期限，无法删除'
                    }, status=403)
            
            # 记录操作日志
            self._log_operation(request, 'DELETE', report)
            report.delete()
            return JsonResponse({'status': 'success', 'message': '删除成功'})
        except self.model_class.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '报表不存在'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    def _serialize_report(self, report, current_user=None):
        """序列化报表数据"""
        user_info = get_user_info(report.username)
        
        # 计算权限状态
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d'),
            'time': report.time.strftime('%H:%M'),
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'xinghui_permeability': getattr(report, 'xinghui_permeability', None),
            'wet_cake_density': report.wet_cake_density,
            'filter_time': getattr(report, 'filter_time', None),
            'water_viscosity': getattr(report, 'water_viscosity', None),
            'cake_thickness': getattr(report, 'cake_thickness', None),
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'username': user_info.get('name', report.username),
            'original_username': report.username,  # 原始用户名，用于权限检查
            'can_edit': can_edit,
                    'can_delete': can_delete,
        'permission_reason': self._get_permission_reason(report, current_user)
    }

    def _apply_filters(self, queryset, request):
        """应用筛选条件"""
        from home.utils import get_user_data_filter_by_company_department
        
        # 应用数据权限过滤
        data_filter = get_user_data_filter_by_company_department(request.user, self.report_name, 'username')
        if data_filter:
            queryset = queryset.filter(**data_filter)
        
        # 应用其他筛选条件
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        start_time = request.GET.get('start_time')
        end_time = request.GET.get('end_time')
        product_name = request.GET.get('product_name')
        packaging = request.GET.get('packaging')
        shift = request.GET.get('shift')
        squad = request.GET.get('squad')

        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        if start_time:
            queryset = queryset.filter(time__gte=start_time)
        if end_time:
            queryset = queryset.filter(time__lte=end_time)
        if product_name:
            queryset = queryset.filter(product_name__icontains=product_name)
        if packaging:
            queryset = queryset.filter(packaging__icontains=packaging)
        # 处理班组筛选：优先使用squad参数（前端表单使用），其次使用shift参数（兼容性）
        if squad:
            queryset = queryset.filter(shift__icontains=squad)
        elif shift:
            queryset = queryset.filter(shift__icontains=shift)

        return queryset

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用细粒度权限控制"""
        if not current_user:
            return False
        
        # 使用细粒度权限控制，支持跨用户编辑
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, self.report_name)
    
    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用细粒度权限控制"""
        if not current_user:
            return False
        
        # 使用细粒度权限控制，支持跨用户删除
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, self.report_name)
    
    def _check_date_limit(self, report):
        """检查日期限制"""
        from datetime import datetime
        
        # 获取编辑时间限制参数
        edit_limit = 7  # 默认7天
        try:
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            return days_diff <= edit_limit
        
        return True
    
    def _get_permission_reason(self, report, current_user):
        """获取权限限制原因"""
        if not current_user:
            return '未登录'
        
        # 检查编辑权限
        can_edit = self._check_edit_permission(report, current_user)
        if not can_edit:
            # 详细分析权限限制原因
            if report.username != current_user.username:
                # 检查是否启用跨用户编辑
                try:
                    from home.models import Parameter
                    cross_edit_param = Parameter.objects.filter(id='enable_cross_user_edit').first()
                    if not cross_edit_param or cross_edit_param.value != 'true':
                        return '系统未启用跨用户编辑功能'
                    
                    # 检查是否有跨用户编辑权限
                    from home.utils import has_hierarchical_permission
                    module_code = get_report_module_code(self.report_name)
                    
                    if not has_hierarchical_permission(current_user, f"{module_code}_edit_others"):
                        return '无跨用户编辑权限'
                    
                    return '无权限编辑他人数据'
                except:
                    return '无权限编辑他人数据'
            
            if not self._check_date_limit(report):
                from datetime import datetime
                edit_limit = 7
                try:
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                return f'超过{edit_limit}天编辑期限'
        
        return ''

    def _process_input_data(self, data, request):
        """处理输入数据"""
        # 处理日期和时间
        if 'date' in data:
            data['date'] = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if 'time' in data:
            data['time'] = datetime.strptime(data['time'], '%H:%M').time()

        # 添加用户信息
        data['user'] = request.user
        data['username'] = request.user.username

        return data

    def calculate_yesterday_production(self, request):
        """统计昨日产量 - 按班组、产品型号、包装类型、批号、备注分组统计吨数"""
        from datetime import date, timedelta
        
        try:
            # 获取昨天的日期
            yesterday = date.today() - timedelta(days=1)
            
            # 查询昨日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=yesterday,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'remarks',  # 备注
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number', 'remarks')
            
            # 按5个字段分组累加产量（与Excel导出逻辑保持一致）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置',
                    report['remarks'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'remarks': group_key[4],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'remarks': stat['remarks'] if stat['remarks'] is not None and stat['remarks'] != '' else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': yesterday.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': yesterday.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def calculate_today_production(self, request):
        """统计今日产量 - 按班组、产品型号、包装类型、批号、备注分组统计吨数"""
        from datetime import date
        
        try:
            # 获取今天的日期
            today = date.today()
            
            # 查询今日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=today,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'remarks',  # 备注
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number', 'remarks')
            
            # 按5个字段分组累加产量（与Excel导出逻辑保持一致）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置',
                    report['remarks'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'remarks': group_key[4],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'remarks': stat['remarks'] if stat['remarks'] is not None and stat['remarks'] != '' else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': today.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': today.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def _log_operation(self, request, operation_type, report, new_data=None):
        """记录操作日志"""
        import logging
        from home.models import UserOperationLog
        
        try:
            # 获取报表类型
            report_type = self._get_report_type()
            
            # 处理 new_data，确保所有对象都可以JSON序列化
            if new_data:
                # 将 new_data 中的日期/时间对象转换为字符串
                processed_data = {}
                for key, value in new_data.items():
                    if hasattr(value, 'strftime'):  # 日期或时间对象
                        processed_data[key] = str(value)
                    elif hasattr(value, '__dict__'):  # 复杂对象
                        processed_data[key] = str(value)
                    else:
                        processed_data[key] = value
                serialized_data = processed_data
            else:
                serialized_data = self._serialize_for_log(report)
            
            # 记录操作日志
            log_entry = UserOperationLog.log_operation(
                request=request,
                operation_type=operation_type,
                report_type=report_type,
                report_id=report.id,
                operation_detail=f'{operation_type} {self.report_name}，日期: {report.date}，班次: {report.shift}',
                new_data=serialized_data
            )
            
            # 如果日志记录成功，记录更多调试信息
            if log_entry:
                logger = logging.getLogger('user_operations')
                logger.info(f"操作日志记录成功 - ID: {log_entry.id}")
                
                # 记录客户端环境信息
                client_info = request.META.get('HTTP_USER_AGENT', '未知')
                logger.info(f"客户端环境: {client_info}")
                
                # 记录页面状态信息
                page_state = {
                    'currentPath': request.path,
                    'title': '未知',
                    'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                logger.info(f"页面状态: {json.dumps(page_state, ensure_ascii=False, indent=2)}")
                
                # 记录会话信息
                session_info = {
                    'hasCsrfToken': 'csrftoken' in request.COOKIES,
                    'hasSessionCookie': 'sessionid' in request.COOKIES,
                    'sessionCookieLength': len(request.COOKIES.get('sessionid', ''))
                }
                logger.info(f"会话信息: {json.dumps(session_info, ensure_ascii=False, indent=2)}")
                
                # 记录页面元素状态
                element_state = {
                    'hasFilterForm': False,
                    'hasDataContainer': False,
                    'filterFormElements': []
                }
                logger.info(f"页面元素状态: {json.dumps(element_state, ensure_ascii=False, indent=2)}")
                
                # 记录错误信息（如果有）
                if 'error' in request.GET:
                    logger.warning(f"操作过程中出现错误: {request.GET.get('error', '未知错误')}")
                    if 'errorDetail' in request.GET:
                        logger.warning(f"错误详情: {request.GET.get('errorDetail', '')}")
                    if 'errorStack' in request.GET:
                        logger.warning(f"错误堆栈: {request.GET.get('errorStack', '')}")
        
        except Exception as e:
            # 如果日志记录失败，记录到Django日志
            logger = logging.getLogger('home')
            logger.error(f'记录操作日志失败: {str(e)}', exc_info=True)

    def _get_report_type(self):
        """获取报表类型标识"""
        report_name = getattr(self, 'report_name', '')
        if '东泰' in report_name: return 'dongtai'
        elif '远通' in report_name and '二线' not in report_name: return 'yuantong'
        elif '远通二线' in report_name or '远通2号' in report_name: return 'yuantong2'
        elif '大塬' in report_name: return 'dayuan'
        elif '长富' in report_name: return 'changfu'
        elif '兴辉' in report_name and '二线' not in report_name: return 'xinghui'
        elif '兴辉二线' in report_name or '兴辉2号' in report_name: return 'xinghui2'
        else: return None

    def _serialize_for_log(self, report):
        """序列化报表数据用于日志记录"""
        try:
            return {
                'id': report.id, 'date': str(report.date) if report.date else None,
                'time': str(report.time) if report.time else None, 'shift': report.shift,
                'product_name': report.product_name, 'packaging': report.packaging,
                'batch_number': report.batch_number, 'moisture_after_drying': report.moisture_after_drying,
                'alkali_content': report.alkali_content, 'flux': report.flux,
                'permeability': report.permeability, 'permeability_long': report.permeability_long,
                'wet_cake_density': report.wet_cake_density, 'filter_time': getattr(report, 'filter_time', None),
                'water_viscosity': getattr(report, 'water_viscosity', None), 'cake_thickness': getattr(report, 'cake_thickness', None),
                'bulk_density': report.bulk_density, 'brightness': report.brightness,
                'swirl': report.swirl, 'odor': report.odor, 'conductance': report.conductance,
                'ph': report.ph, 'moisture': report.moisture, 'bags': report.bags, 'tons': report.tons,
                'fe_ion': report.fe_ion, 'ca_ion': report.ca_ion, 'al_ion': report.al_ion,
                'oil_absorption': report.oil_absorption, 'water_absorption': report.water_absorption,
                'remarks': report.remarks, 'username': report.username,
                'created_at': str(report.created_at) if report.created_at else None,
                'updated_at': str(report.updated_at) if report.updated_at else None,
            }
        except Exception as e:
            return {'error': f'序列化失败: {str(e)}'}


# 原土入库删除接口
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.shortcuts import render
import requests




# QC报表视图
@login_required
@csrf_exempt
def qc_report(request):
    """QC报表页面"""
    # 根据用户权限过滤菜单
    filtered_menu_items = filter_menu_by_permission(MENU_ITEMS, request.user.username)
    
    return render(request, 'production/qc_report.html', {
        'user': request.user,
        'menu_items': filtered_menu_items,
    })

@login_required
@permission_required('qc_report_view')
def dayuan_report(request):
    """大塬报表页面"""
    return render(request, 'production/dayuan_report.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def dayuan_report_import_excel(request):
    """大塬QC报表Excel导入功能"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 检查是否有上传的文件
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': '请选择要导入的Excel文件'}, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': '仅支持Excel文件格式(.xlsx, .xls)'}, status=400)
        
        # 尝试导入pandas，如果失败则使用openpyxl
        try:
            import pandas as pd
            use_pandas = True
        except ImportError:
            use_pandas = False
        
        # 读取Excel文件
        if use_pandas:
            try:
                df = pd.read_excel(excel_file, sheet_name=0)
                # 删除完全空白的行
                df = df.dropna(how='all')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'读取Excel文件失败: {str(e)}'}, status=400)
        else:
            # 使用openpyxl读取
            from openpyxl import load_workbook
            try:
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                # 读取表头
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value if cell.value else '')
                
                # 处理特殊列名情况
                # 1. 处理重复的列名（如两个Permeability或两个渗透率）
                permeability_indices = []
                for i, header in enumerate(headers):
                    if header and ('Permeability' in str(header) or '渗透率' in str(header) or '滤速率' in str(header)):
                        permeability_indices.append(i)
                
                # 如果有两个渗透率列，重命名它们
                if len(permeability_indices) == 2:
                    headers[permeability_indices[0]] = 'Permeability_1'  # 第一个是远通渗透率
                    headers[permeability_indices[1]] = 'Permeability_2'  # 第二个是长富渗透率
                
                # 2. 处理"入窑前"和"含量 (%)"合并的情况
                # 如果找到"入窑前"列，检查下一列是否是"含量 (%)"
                for i, header in enumerate(headers):
                    if header and '入窑前' in str(header):
                        if i + 1 < len(headers) and headers[i + 1] and '含量' in str(headers[i + 1]):
                            # 将"含量 (%)"列也映射到alkali_content
                            headers[i + 1] = '入窑前碱含量(%)_content'
                        break
                
                # 3. 处理合并列"铁离子 (mg/钙离子 (mg/铝离子 (mg/白度"
                # 这种合并列需要在读取数据时特殊处理
                for i, header in enumerate(headers):
                    if header and '铁离子' in str(header) and '钙离子' in str(header):
                        # 这是一个合并列，需要拆分
                        headers[i] = '铁离子_钙离子_铝离子_白度_合并列'
                        break
                
                # 4. 处理合并列"电导值 (as/c pH"
                for i, header in enumerate(headers):
                    if header and '电导值' in str(header) and 'pH' in str(header):
                        # 这是一个合并列
                        headers[i] = '电导值_pH_合并列'
                        break
                
                # 读取数据行，跳过完全空白的行
                df_data = []
                # 记录原始列名，用于调试
                logger.info(f'📋 Excel原始列名: {headers}')
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 检查是否是完全空白的行
                    if any(cell is not None and str(cell).strip() != '' for cell in row):
                        row_dict = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                val = row[i]
                                # 处理空值和空字符串
                                if val is None or (isinstance(val, str) and val.strip() == ''):
                                    row_dict[header] = None
                                else:
                                    row_dict[header] = val
                            else:
                                row_dict[header] = None
                        df_data.append(row_dict)
                
                # 创建一个简单的DataFrame模拟对象
                class SimpleDF:
                    def __init__(self, data, columns):
                        self.data = data
                        self.columns = columns
                    
                    def iterrows(self):
                        for idx, row_dict in enumerate(self.data):
                            # 创建一个简单的行对象，支持get方法
                            class Row:
                                def __init__(self, data):
                                    self._data = data
                                
                                def get(self, key, default=None):
                                    return self._data.get(key, default)
                                
                                def __getitem__(self, key):
                                    return self._data.get(key)
                            
                            yield idx, Row(row_dict)
                    
                    def __len__(self):
                        return len(self.data)
                
                df = SimpleDF(df_data, headers)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'读取Excel文件失败: {str(e)}'}, status=400)
        
        # 定义列名映射关系（支持中英文）
        # 注意：根据新的Excel表头格式进行映射，字段顺序与QC_REPORT_FIELD_MAPPING保持一致
        column_mapping = {
            # 基本字段映射
            '日期': 'date', 'Date': 'date', '检测日期': 'date',
            '时间': 'time', 'Time': 'time', '检测时间': 'time',
            '班次': 'shift', 'Shift': 'shift', 'Squad': 'shift', '班组': 'shift', '班别': 'shift',
            '产品名称': 'product_name', 'Product Name': 'product_name', 'Grade': 'product_name', 
            '产品型号': 'product_name',
            '包装类型': 'packaging', 'Packaging': 'packaging', 'IPKP CODE': 'packaging',
            '批号': 'batch_number', 'Batch Number': 'batch_number', '批次号': 'batch_number', 
            'LOT': 'batch_number', '批号/日期': 'batch_number', '批次': 'batch_number',
            
            # 检测数据字段映射 - 新表头格式（按照QC_REPORT_FIELD_MAPPING的顺序）
            '烘干后原土水分 (%)': 'moisture_after_drying',
            '烘干后原土水分（%）': 'moisture_after_drying',  # 中文括号
            '烘干后原土水分(%)': 'moisture_after_drying',
            '干燥后原土水分(%)': 'moisture_after_drying',
            '干燥后原土水分（%）': 'moisture_after_drying',  # 中文括号
            '烘干后原': 'moisture_after_drying',
            'Moisture after drying': 'moisture_after_drying', 
            '干燥后原土水分': 'moisture_after_drying',
            
            # 入窑前碱含量 - 新表头可能是分开的两列"入窑前"和"含量 (%)"，需要合并处理
            '入窑前碱含量(%)': 'alkali_content', 
            '入窑前碱含量': 'alkali_content',
            '入窑前': 'alkali_content',  # 如果只有"入窑前"列，也映射到alkali_content
            '含量 (%)': 'alkali_content',  # 如果"含量 (%)"单独存在
            'Alkali content (%)': 'alkali_content',
            
            '助剂添加比例': 'flux',
            '助溶剂添加比例': 'flux', 
            '助溶剂': 'flux', 
            '*flux agent': 'flux', 
            'flux agent addition ratio': 'flux',
            
            # 渗透率 - 新表头格式（按照QC_REPORT_FIELD_MAPPING的顺序）
            '运通滤速率': 'permeability',  # 运通 = 远通
            '远通渗透率(Darcy)': 'permeability', 
            '远通渗透率': 'permeability',
            '长高滤速率': 'permeability_long',  # 长高 = 长富
            '长富渗透率(Darcy)': 'permeability_long', 
            '长富渗透率': 'permeability_long',
            # 注意：如果Excel中有两个"Permeability"列，需要通过列位置区分
            # 第一个Permeability对应远通渗透率，第二个对应长富渗透率
            # 注意：大塬不使用兴辉渗透率字段
            
            # 可塑度可能是涡值
            '可塑度 (c/cm)': 'swirl',
            '可塑度': 'swirl',
            '涡值(cm)': 'swirl',
            '涡值（cm）': 'swirl',  # 中文括号
            '涡值': 'swirl', 
            'Swirl (cm)': 'swirl', 
            'Swirl': 'swirl',
            
            # 饼密度和振实密度（按照QC_REPORT_FIELD_MAPPING的顺序）
            '饼密度(g/cm3)': 'wet_cake_density', 
            '饼密度（g/cm3）': 'wet_cake_density',  # 中文括号
            '饼密度': 'wet_cake_density', 
            'Wet cake density': 'wet_cake_density',
            # 注意：大塬不使用以下字段：远通饼密度、长富饼密度、过滤时间、水黏度、饼厚
            '振实密度(g/cm3)': 'bulk_density',
            '振实密度（g/cm3)': 'bulk_density',  # 中文括号
            '振实密度（g/cm3）': 'bulk_density',  # 中文括号
            '振实密度': 'bulk_density',
            '重度 (k) 14W': 'bulk_density',  # 可能是振实密度的另一种表示
            '灰值 (c/m)': 'bulk_density',  # 灰值可能是振实密度
            
            # 白度
            '白度': 'brightness', 
            'Bri.': 'brightness', 
            'Brightness': 'brightness',
            
            '气味': 'odor', 
            'Odor': 'odor',
            
            # 电导值和pH - 新表头可能是合并的"电导值 (as/c pH"
            '电导值 (as/c pH': 'conductance',  # 如果合并列，优先映射到电导值
            '电导值(ms/cm)': 'conductance', 
            '电导值': 'conductance', 
            'Conductance (ms/c)': 'conductance', 
            'Conductance': 'conductance',
            'pH': 'ph', 
            'pH值': 'ph',
            
            '水分(%)': 'moisture', 
            '水分': 'moisture', 
            'Moisture (%)': 'moisture', 
            'Moisture': 'moisture',
            
            '批数': 'bags',  # 新表头中的"批数"可能是"袋数"
            '袋数': 'bags', 
            'Bags': 'bags',
            
            '吨': 'tons', 
            'Tons': 'tons', 
            '产量': 'tons',
            
            # 筛分数据字段映射 - 新表头格式（没有+号）
            '14W': 'sieving_14m',  # 重度 (k) 14W 中的14W
            '+14M (%)': 'sieving_14m', 
            '+14M': 'sieving_14m', 
            '14M': 'sieving_14m',
            '+30M (%)': 'sieving_30m', 
            '+30M': 'sieving_30m', 
            '30M': 'sieving_30m',
            '+40M (%)': 'sieving_40m', 
            '+40M': 'sieving_40m', 
            '40M': 'sieving_40m',
            'M': 'sieving_40m',  # 单独的M可能是40M
            '+80M (%)': 'sieving_80m', 
            '+80M': 'sieving_80m', 
            '80M': 'sieving_80m',
            '+100M (%)': 'sieving_100m', 
            '+100M': 'sieving_100m', 
            '100M': 'sieving_100m',
            '+150M (%)': 'sieving_150m', 
            '+150M': 'sieving_150m',
            '150M': 'sieving_150m',
            '150M ': 'sieving_150m',  # 带尾随空格
            '+200M (%)': 'sieving_200m', 
            '+200M': 'sieving_200m', 
            '200M': 'sieving_200m',
            '200M ': 'sieving_200m',  # 带尾随空格
            '+325M (%)': 'sieving_325m', 
            '+325M': 'sieving_325m', 
            '325M': 'sieving_325m',
            '325M ': 'sieving_325m',  # 带尾随空格
            
            # 离子数据字段映射 - 新表头可能是合并列"铁离子 (mg/钙离子 (mg/铝离子 (mg/白度"
            # 这种情况需要在读取时特殊处理，这里先提供单独列的映射
            '铁离子 (mg/钙离子 (mg/铝离子 (mg/白度': 'fe_ion',  # 合并列，优先取第一个
            '铁离子（mg/kg）': 'fe_ion',  # 中文括号
            '铁离子(mg/kg)': 'fe_ion',
            '铁离子': 'fe_ion', 
            'Fe离子': 'fe_ion', 
            'Fe': 'fe_ion',
            '钙离子（mg/kg）': 'ca_ion',  # 中文括号
            '钙离子(mg/kg)': 'ca_ion',
            '钙离子': 'ca_ion',
            'Ca离子': 'ca_ion', 
            'Ca': 'ca_ion',
            '铝离子（mg/kg）': 'al_ion',  # 中文括号
            '铝离子(mg/kg)': 'al_ion',
            '铝离子': 'al_ion',
            'Al离子': 'al_ion', 
            'Al': 'al_ion',
            
            '吸油率 (%)': 'oil_absorption',
            '吸油率（%）': 'oil_absorption',  # 中文括号
            '吸油量': 'oil_absorption', 
            '吸油率(%)': 'oil_absorption',
            '吸水率 (%)': 'water_absorption',
            '吸水率（%）': 'water_absorption',  # 中文括号
            '吸水量': 'water_absorption', 
            '吸水率(%)': 'water_absorption',
            
            '水分(%)': 'moisture',
            '水分（%）': 'moisture',  # 中文括号
            '水分   （%）': 'moisture',  # 多个空格和中文括号
            '水分': 'moisture', 
            'Moisture (%)': 'moisture', 
            'Moisture': 'moisture',
            
            '备注': 'remarks', 
            'Remarks': 'remarks', 
            'Notes': 'remarks'
        }
        
        # 处理数据并导入
        imported_count = 0
        error_count = 0
        error_messages = []
        skipped_count = 0  # 记录跳过的行数
        
        # 统一处理逻辑：先映射列名，再处理数据
        # 如果是pandas DataFrame，先重命名列
        if use_pandas and hasattr(df, 'rename'):
            # 处理重复的列名（如两个Permeability）
            # 找到所有Permeability列
            permeability_cols = [col for col in df.columns if col and ('Permeability' in str(col) or '渗透率' in str(col) or '滤速率' in str(col))]
            if len(permeability_cols) == 2:
                # 重命名第一个和第二个
                df.columns = [f'Permeability_1' if col == permeability_cols[0] else 
                             f'Permeability_2' if col == permeability_cols[1] else col 
                             for col in df.columns]
            
            # 添加Permeability_1和Permeability_2到映射
            column_mapping['Permeability_1'] = 'permeability'
            column_mapping['Permeability_2'] = 'permeability_long'
            
            # 改进的列名映射：先尝试精确匹配，再尝试模糊匹配
            def normalize_col_name(col_name):
                """标准化列名，去除空格、统一括号等"""
                if col_name is None:
                    return ''
                col_str = str(col_name).strip()
                # 统一括号：将中文括号转换为英文括号
                col_str = col_str.replace('（', '(').replace('）', ')')
                # 去除多余空格
                col_str = ' '.join(col_str.split())
                return col_str
            
            # 创建标准化后的映射字典
            normalized_mapping = {}
            for orig_key, mapped_value in column_mapping.items():
                normalized_key = normalize_col_name(orig_key)
                normalized_mapping[normalized_key] = mapped_value
            
            # 先进行精确匹配映射
            rename_dict = {}
            unmapped_cols = []
            
            for col in df.columns:
                col_normalized = normalize_col_name(col)
                # 先尝试精确匹配
                if col in column_mapping:
                    rename_dict[col] = column_mapping[col]
                elif col_normalized in normalized_mapping:
                    rename_dict[col] = normalized_mapping[col_normalized]
                else:
                    # 尝试模糊匹配
                    matched = False
                    for orig_key, mapped_value in column_mapping.items():
                        orig_normalized = normalize_col_name(orig_key)
                        # 检查是否包含关键部分
                        if orig_normalized and col_normalized:
                            # 去除括号和空格后比较
                            orig_simple = orig_normalized.replace(' ', '').replace('(', '').replace(')', '').lower()
                            col_simple = col_normalized.replace(' ', '').replace('(', '').replace(')', '').lower()
                            if orig_simple == col_simple or (len(orig_simple) > 3 and orig_simple in col_simple):
                                rename_dict[col] = mapped_value
                                matched = True
                                break
                    
                    if not matched:
                        unmapped_cols.append(col)
            
            # 执行重命名
            df_mapped = df.rename(columns=rename_dict)
            
            # 记录未映射的列名（用于调试）
            if unmapped_cols:
                logger.debug(f'⚠️ 未映射的列名: {unmapped_cols}')
            
            # 特别检查swirl字段的映射情况（大塬导入专用）
            if 'swirl' not in df_mapped.columns:
                logger.warning(f'⚠️ swirl字段未映射！原始列名: {list(df.columns)}, 映射字典: {rename_dict}')
                # 尝试查找包含"涡值"的列
                for col in df.columns:
                    if '涡值' in str(col) or 'swirl' in str(col).lower():
                        logger.info(f'🔍 发现疑似swirl字段: {col}，尝试强制映射')
                        if col not in rename_dict:
                            rename_dict[col] = 'swirl'
                            df_mapped = df_mapped.rename(columns={col: 'swirl'})
                            logger.info(f'✅ 强制映射: {col} -> swirl')
                            break
            
            logger.info(f'📊 读取到 {len(df_mapped)} 行数据，原始列名: {list(df.columns)}, 映射后列名: {list(df_mapped.columns)}')
        else:
            # 对于openpyxl，需要手动映射
            # 添加Permeability_1和Permeability_2到映射
            column_mapping['Permeability_1'] = 'permeability'
            column_mapping['Permeability_2'] = 'permeability_long'
            
            df_mapped = df
            logger.info(f'📊 读取到 {len(df_mapped)} 行数据，列名: {df_mapped.columns if hasattr(df_mapped, "columns") else "未知"}')
        
        # 处理每一行数据
        for index, row_obj in df_mapped.iterrows():
            try:
                # 统一获取行数据的方法
                if use_pandas:
                    row = row_obj
                else:
                    row = row_obj
                
                # 获取行数据的辅助函数
                def get_row_value(key):
                    """从行中获取值，支持pandas和openpyxl两种格式"""
                    # 添加调试日志（仅对swirl字段和前几行）
                    if key == 'swirl' and index < 3:
                        logger.debug(f'第{index+2}行，开始获取swirl字段值，use_pandas={use_pandas}')
                    if use_pandas:
                        # pandas格式 - 列名已经通过rename映射过了，直接用key（数据库字段名）获取
                        try:
                            # 首先尝试直接通过映射后的列名获取
                            if hasattr(row, 'get'):
                                val = row.get(key)
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    if key == 'swirl' and index < 3:
                                        logger.debug(f'第{index+2}行，swirl字段通过row.get获取到值: {val}')
                                    return val
                            if hasattr(row, key):
                                val = getattr(row, key)
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    if key == 'swirl' and index < 3:
                                        logger.debug(f'第{index+2}行，swirl字段通过getattr获取到值: {val}')
                                    return val
                            if hasattr(row, 'index') and key in row.index:
                                val = row[key]
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    if key == 'swirl' and index < 3:
                                        logger.debug(f'第{index+2}行，swirl字段通过row[key]获取到值: {val}')
                                    return val
                            
                            # 如果直接获取不到，尝试通过原始列名获取（模糊匹配）
                            # 定义字段的关键词映射
                            field_keywords = {
                                'date': ['日期', 'date', 'Date', '检测日期'],
                                'time': ['时间', 'time', 'Time', '检测时间'],
                                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                                'permeability': ['远通', '渗透率', 'permeability', 'Darcy', 'Permeability_1'],
                                'permeability_long': ['长富', '长高', '渗透率', 'permeability', 'Permeability_2'],
                                'wet_cake_density': ['饼密度', 'cake', 'density'],
                                'bulk_density': ['振实密度', 'bulk', 'density'],
                                'brightness': ['白度', 'brightness', 'Bri'],
                                'swirl': ['涡值', '可塑度', 'swirl', 'Swirl'],
                                'odor': ['气味', 'odor'],
                                'conductance': ['电导值', 'conductance'],
                                'ph': ['pH', 'ph'],
                                'moisture': ['水分', 'moisture'],
                                'bags': ['袋数', 'bags', '批数'],
                                'tons': ['吨', 'tons', '产量'],
                                'fe_ion': ['铁离子', 'Fe', 'fe'],
                                'ca_ion': ['钙离子', 'Ca', 'ca'],
                                'al_ion': ['铝离子', 'Al', 'al'],
                                'oil_absorption': ['吸油', 'oil'],
                                'water_absorption': ['吸水', 'water'],
                                'sieving_14m': ['14M', '14', '+14'],
                                'sieving_30m': ['30M', '30', '+30'],
                                'sieving_40m': ['40M', '40', '+40'],
                                'sieving_80m': ['80M', '80', '+80'],
                                'sieving_100m': ['100M', '100', '+100'],
                                'sieving_150m': ['150M', '150', '+150'],
                                'sieving_200m': ['200M', '200', '+200'],
                                'sieving_325m': ['325M', '325', '+325'],
                            }
                            
                            # 获取当前字段的关键词
                            keywords = field_keywords.get(key, [])
                            
                            # 遍历行的所有列（使用row.index获取列名）
                            if hasattr(row, 'index'):
                                # 优先匹配更精确的关键词（中文关键词优先，更长的关键词优先）
                                # 对关键词按优先级排序：中文关键词 > 英文关键词，长关键词 > 短关键词
                                sorted_keywords = sorted(keywords, key=lambda k: (k.isascii(), -len(k)))
                                
                                for keyword in sorted_keywords:
                                    best_match_col = None
                                    best_match_val = None
                                    
                                    for col_name in row.index:
                                        if col_name is None:
                                            continue
                                        col_str = str(col_name).strip()
                                        
                                        # 检查是否包含关键词
                                        if keyword in col_str:
                                            try:
                                                val = row[col_name]
                                                # 检查值是否为空
                                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                                    # 找到匹配的列，记录最佳匹配
                                                    if best_match_col is None:
                                                        best_match_col = col_name
                                                        best_match_val = val
                                                    # 如果当前列名更精确（包含更多关键词），优先选择
                                                    elif len(col_str) < len(str(best_match_col)):
                                                        best_match_col = col_name
                                                        best_match_val = val
                                            except (KeyError, IndexError):
                                                continue
                                    
                                    # 如果找到最佳匹配，返回该值
                                    if best_match_val is not None:
                                        if index < 3 and key in ['bulk_density', 'swirl']:
                                            logger.debug(f'第{index+2}行，字段{key}通过关键词"{keyword}"匹配到列"{best_match_col}"，值: {best_match_val}, 类型: {type(best_match_val)}')
                                        return best_match_val
                                    
                                    # 如果当前关键词找到了列但值为空，记录但不继续匹配其他关键词
                                    # （因为已经找到了正确的列，只是值为空）
                                    if best_match_col is not None:
                                        if index < 3 and key in ['bulk_density', 'swirl']:
                                            logger.debug(f'第{index+2}行，字段{key}通过关键词"{keyword}"匹配到列"{best_match_col}"但值为空')
                                        return None
                            
                            # 如果没有找到匹配的值，返回None
                            if index < 3 and key in ['bulk_density', 'swirl']:
                                logger.debug(f'第{index+2}行，字段{key}未找到匹配的列')
                            return None
                        except Exception as e:
                            if key == 'swirl' and index < 3:
                                logger.debug(f'获取字段 {key} 失败: {str(e)}')
                            return None
                    else:
                        # openpyxl格式，row是字典，key是Excel原始列名
                        # 先尝试直接用key获取（如果列名就是数据库字段名）
                        if isinstance(row, dict):
                            if key in row:
                                val = row[key]
                                if val is not None:
                                    return val
                        
                        # 通过column_mapping查找所有可能映射到key的原始列名
                        # 遍历所有映射，找到mapped_col == key的orig_col
                        matched_cols = []
                        for orig_col, mapped_col in column_mapping.items():
                            if mapped_col == key:
                                matched_cols.append(orig_col)
                        
                        # 尝试从所有匹配的列名中获取值
                        for orig_col in matched_cols:
                            if isinstance(row, dict) and orig_col in row:
                                val = row[orig_col]
                                if val is not None:
                                    if key == 'swirl' and index < 3:
                                        logger.debug(f'第{index+2}行，swirl字段通过列名"{orig_col}"获取到值: {val}')
                                    return val
                        
                        # 如果还是找不到，尝试模糊匹配（去除空格、括号等）
                        if isinstance(row, dict):
                            # 定义关键字段的匹配关键词
                            key_keywords = {
                                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                                'permeability': ['远通', '渗透率', 'permeability', 'Darcy'],
                                'permeability_long': ['长富', '长高', '渗透率', 'permeability'],
                                'swirl': ['涡值', '可塑度', 'swirl', 'Swirl'],
                                'sieving_150m': ['150M', '150', '+150'],
                                'sieving_200m': ['200M', '200', '+200'],
                                'sieving_325m': ['325M', '325', '+325'],
                                'fe_ion': ['铁离子', 'Fe', 'fe'],
                                'ca_ion': ['钙离子', 'Ca', 'ca'],
                                'al_ion': ['铝离子', 'Al', 'al'],
                                'oil_absorption': ['吸油', 'oil'],
                                'water_absorption': ['吸水', 'water'],
                                'moisture': ['水分', 'moisture'],
                            }
                            
                            # 获取当前字段的关键词
                            keywords = key_keywords.get(key, [])
                            
                            # 遍历所有列名，查找匹配的
                            for orig_col in row.keys():
                                if orig_col is None:
                                    continue
                                orig_col_str = str(orig_col).strip()
                                
                                # 如果有关键词，检查是否包含关键词
                                if keywords:
                                    for keyword in keywords:
                                        if keyword in orig_col_str:
                                            val = row[orig_col]
                                            if val is not None:
                                                if key == 'swirl' and index < 3:
                                                    logger.debug(f'第{index+2}行，swirl字段通过关键词"{keyword}"匹配到列"{orig_col}"获取到值: {val}')
                                                return val
                                else:
                                    # 没有关键词，使用标准化匹配
                                    key_normalized = str(key).strip().replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                                    orig_col_normalized = orig_col_str.replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                                    # 检查是否包含关键部分
                                    if key_normalized in orig_col_normalized or orig_col_normalized in key_normalized:
                                        val = row[orig_col]
                                        if val is not None:
                                            return val
                        
                        return None
                
                def is_notna(val):
                    """检查值是否不为空"""
                    if use_pandas:
                        return pd.notna(val)
                    else:
                        return val is not None and val != '' and str(val).strip() != ''
                
                # 检查是否为空行 - 检查关键字段
                date_val = get_row_value('date')
                product_name_val = get_row_value('product_name')
                
                # 放宽空行检查：只要日期或产品名称有一个不为空，就继续处理
                # 或者有其他任何字段有值
                has_any_data = False
                if date_val and is_notna(date_val):
                    has_any_data = True
                if product_name_val and is_notna(product_name_val):
                    has_any_data = True
                
                # 如果日期和产品名称都为空，再检查其他字段
                if not has_any_data:
                    for key in ['shift', 'packaging', 'bags', 'batch_number', 'moisture_after_drying', 
                               'alkali_content', 'permeability', 'permeability_long', 'wet_cake_density']:
                        val = get_row_value(key)
                        if val and is_notna(val):
                            has_any_data = True
                            break
                    
                    # 如果没有任何数据，跳过这一行
                    if not has_any_data:
                        skipped_count += 1
                        continue
                
                data = {}
                
                # 处理日期
                if date_val and is_notna(date_val):
                    if isinstance(date_val, str):
                        try:
                            data['date'] = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                        except:
                            try:
                                # 尝试其他日期格式
                                data['date'] = datetime.strptime(date_val.strip(), '%Y/%m/%d').date()
                            except:
                                data['date'] = date.today()
                    elif hasattr(date_val, 'date'):
                        data['date'] = date_val.date()
                    else:
                        data['date'] = date.today()
                else:
                    data['date'] = date.today()
                
                # 处理时间
                time_val = get_row_value('time')
                # 添加调试日志（仅在前几行）
                if index < 3:
                    logger.debug(f'第{index+2}行，时间字段time，获取到的值: {time_val}, 类型: {type(time_val)}')
                
                if time_val and is_notna(time_val):
                    try:
                        # 如果是字符串类型
                        if isinstance(time_val, str):
                            time_str = str(time_val).strip()
                            if not time_str:
                                data['time'] = time(0, 0)
                            else:
                                try:
                                    # 尝试 HH:MM 格式
                                    data['time'] = datetime.strptime(time_str, '%H:%M').time()
                                except ValueError:
                                    try:
                                        # 尝试 HH:MM:SS 格式
                                        data['time'] = datetime.strptime(time_str, '%H:%M:%S').time()
                                    except ValueError:
                                        try:
                                            # 尝试处理Excel时间格式（小数，如0.4166666666666667表示10:00）
                                            if ':' not in time_str and ('.' in time_str or time_str.replace('.', '').replace('-', '').isdigit()):
                                                time_float = float(time_str)
                                                # Excel时间格式：0.0 = 00:00:00, 0.5 = 12:00:00, 1.0 = 24:00:00
                                                total_seconds = int(time_float * 24 * 3600)
                                                hours = total_seconds // 3600
                                                minutes = (total_seconds % 3600) // 60
                                                seconds = total_seconds % 60
                                                data['time'] = time(hours, minutes, seconds)
                                            else:
                                                # 如果都不匹配，尝试直接解析数字（如1000表示10:00）
                                                if time_str.isdigit() and len(time_str) <= 4:
                                                    # 假设是HHMM格式
                                                    hours = int(time_str[:2]) if len(time_str) >= 2 else int(time_str[0])
                                                    minutes = int(time_str[-2:]) if len(time_str) >= 2 else 0
                                                    data['time'] = time(hours, minutes)
                                                else:
                                                    if index < 3:
                                                        logger.debug(f'第{index+2}行，时间字符串无法解析: {time_str}')
                                                    data['time'] = time(0, 0)
                                        except (ValueError, TypeError) as e:
                                            if index < 3:
                                                logger.debug(f'第{index+2}行，时间解析失败: {str(e)}, 原始值: {time_str}')
                                            data['time'] = time(0, 0)
                        # 如果是datetime对象
                        elif hasattr(time_val, 'time'):
                            data['time'] = time_val.time()
                        # 如果是time对象
                        elif isinstance(time_val, time):
                            data['time'] = time_val
                        # 如果是datetime对象（pandas可能返回）
                        elif hasattr(time_val, 'hour') and hasattr(time_val, 'minute'):
                            # 可能是pandas的Timestamp对象
                            data['time'] = time(time_val.hour, time_val.minute, getattr(time_val, 'second', 0))
                        # 如果是数字（Excel时间格式）
                        elif isinstance(time_val, (int, float)):
                            # Excel时间格式：0.0 = 00:00:00, 0.5 = 12:00:00
                            total_seconds = int(time_val * 24 * 3600)
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            seconds = total_seconds % 60
                            data['time'] = time(hours, minutes, seconds)
                        else:
                            if index < 3:
                                logger.debug(f'第{index+2}行，时间值类型不支持: {type(time_val)}, 值: {time_val}')
                            data['time'] = time(0, 0)
                    except Exception as e:
                        if index < 3:
                            logger.debug(f'第{index+2}行，时间处理异常: {str(e)}, 原始值: {time_val}, 类型: {type(time_val)}')
                        data['time'] = time(0, 0)
                else:
                    data['time'] = time(0, 0)
                
                # 处理字符串字段
                string_fields = ['shift', 'product_name', 'packaging', 'batch_number', 'flux', 'remarks']
                for field in string_fields:
                    val = get_row_value(field)
                    if val and is_notna(val):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 特殊处理：处理"入窑前碱含量"
                alkali_val = get_row_value('alkali_content')
                if alkali_val and is_notna(alkali_val):
                    try:
                        data['alkali_content'] = float(alkali_val)
                    except:
                        data['alkali_content'] = None
                else:
                    data['alkali_content'] = None
                
                # 处理数字字段
                numeric_fields = [
                    'moisture_after_drying', 'permeability', 'permeability_long',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor',
                    'conductance', 'ph', 'moisture', 'bags', 'tons', 'fe_ion', 'ca_ion',
                    'al_ion', 'oil_absorption', 'water_absorption', 'sieving_14m', 'sieving_30m',
                    'sieving_40m', 'sieving_80m'
                ]
                # 注意：alkali_content已经在上面单独处理了
                
                for field in numeric_fields:
                    val = get_row_value(field)
                    # 添加调试日志（仅在前几行，特别是bulk_density和swirl字段）
                    if index < 3 or field in ['bulk_density', 'swirl']:
                        logger.debug(f'第{index+2}行，字段{field}，获取到的值: {val}, 类型: {type(val)}')
                        # 对于swirl字段，额外记录列名信息
                        if field == 'swirl' and use_pandas and hasattr(row, 'index'):
                            logger.debug(f'第{index+2}行，swirl字段，可用列名: {list(row.index)}')
                    
                    # 严格检查值是否有效
                    if val is not None and is_notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.strip()
                                # 空字符串或只包含空白字符的字符串视为空值
                                if val == '' or val.isspace():
                                    data[field] = None
                                # 对于swirl字段，如果值是"涡值示例"等示例文本，跳过
                                elif field == 'swirl' and ('示例' in val or 'example' in val.lower()):
                                    if index < 3:
                                        logger.debug(f'第{index+2}行，字段{field}是示例文本，跳过: {val}')
                                    data[field] = None
                                else:
                                    # 尝试转换为浮点数
                                    float_val = float(val)
                                    # 检查是否是NaN或Inf
                                    if pd.isna(float_val) if hasattr(pd, 'isna') else (float_val != float_val or abs(float_val) == float('inf')):
                                        data[field] = None
                                    else:
                                        data[field] = float_val
                            elif isinstance(val, (int, float)):
                                # 检查是否是NaN或Inf
                                if pd.isna(val) if hasattr(pd, 'isna') else (val != val or abs(val) == float('inf')):
                                    data[field] = None
                                else:
                                    data[field] = float(val)
                            else:
                                # 其他类型，尝试转换
                                data[field] = float(val)
                        except (ValueError, TypeError) as e:
                            if index < 3 or field in ['bulk_density', 'swirl']:
                                logger.debug(f'第{index+2}行，字段{field}转换失败: {str(e)}, 原始值: {val}, 类型: {type(val)}')
                            data[field] = None
                    else:
                        # 值为None或空，设置为None
                        if index < 3 or field in ['bulk_density', 'swirl']:
                            logger.debug(f'第{index+2}行，字段{field}值为空，设置为None')
                        data[field] = None
                
                # 特殊处理：处理合并列"铁离子/钙离子/铝离子/白度"
                # 如果找到合并列，尝试拆分
                if not use_pandas:
                    merged_ion_col = None
                    if hasattr(row, 'get'):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    elif isinstance(row, dict):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    
                    if merged_ion_col and is_notna(merged_ion_col):
                        # 尝试从合并列中提取值（如果值是字符串，可能需要解析）
                        # 这里假设合并列的值可能是用斜杠分隔的
                        merged_str = str(merged_ion_col)
                        if '/' in merged_str:
                            parts = merged_str.split('/')
                            if len(parts) >= 3:
                                try:
                                    if parts[0].strip():
                                        data['fe_ion'] = float(parts[0].strip())
                                    if parts[1].strip():
                                        data['ca_ion'] = float(parts[1].strip())
                                    if parts[2].strip():
                                        data['al_ion'] = float(parts[2].strip())
                                    if len(parts) >= 4 and parts[3].strip():
                                        data['brightness'] = float(parts[3].strip())
                                except:
                                    pass
                
                # 特殊处理：处理合并列"电导值/pH"
                if not use_pandas:
                    merged_conductance_col = None
                    if hasattr(row, 'get'):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    elif isinstance(row, dict):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    
                    if merged_conductance_col and is_notna(merged_conductance_col):
                        # 尝试从合并列中提取值
                        merged_str = str(merged_conductance_col)
                        if '/' in merged_str or 'pH' in merged_str:
                            # 尝试解析
                            parts = merged_str.replace('pH', '').split('/')
                            if len(parts) >= 1 and parts[0].strip():
                                try:
                                    data['conductance'] = float(parts[0].strip())
                                except:
                                    pass
                            # pH值可能在字符串的其他位置
                            import re
                            ph_match = re.search(r'pH[:\s]*([0-9.]+)', merged_str, re.IGNORECASE)
                            if ph_match:
                                try:
                                    data['ph'] = float(ph_match.group(1))
                                except:
                                    pass
                
                # 处理筛分字段（可能是字符串）
                sieving_fields = ['sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m']
                for field in sieving_fields:
                    val = get_row_value(field)
                    # 添加调试日志（仅在前几行）
                    if index < 3:
                        logger.debug(f'第{index+2}行，筛分字段{field}，获取到的值: {val}')
                    
                    if val and is_notna(val):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                    
                # 再次检查是否为空行 - 放宽条件，只要有日期或任何字段有值就保留
                has_valid_data = False
                # 检查日期（必须有）
                if data.get('date'):
                    has_valid_data = True
                # 检查其他字段
                if data.get('product_name') and str(data['product_name']).strip():
                    has_valid_data = True
                if data.get('shift') and str(data['shift']).strip():
                    has_valid_data = True
                if data.get('packaging') and str(data['packaging']).strip():
                    has_valid_data = True
                # 检查数字字段是否有值
                for field in ['bags', 'moisture_after_drying', 'alkali_content', 'permeability', 
                             'permeability_long', 'wet_cake_density', 'brightness', 'swirl']:
                    if data.get(field) is not None:
                        has_valid_data = True
                        break
                
                # 如果没有有效数据，跳过这一行
                if not has_valid_data:
                    skipped_count += 1
                    continue
                
                # 设置用户信息
                data['user'] = request.user
                data['username'] = request.user.username
                
                # 记录swirl字段的值（用于调试）
                if index < 3:
                    logger.debug(f'第{index+2}行，准备创建记录，swirl字段值: {data.get("swirl")}, 类型: {type(data.get("swirl"))}')
                
                # 创建记录
                DayuanQCReport.objects.create(**data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                error_msg = f'第 {index + 2} 行导入失败: {str(e)}'
                error_messages.append(error_msg)
                logger.error(f'导入大塬QC报表失败: {error_msg}', exc_info=True)
        
        # 记录操作日志
        from home.models import UserOperationLog
        UserOperationLog.log_operation(
            request, 'CREATE', 'dayuan', None,
            f'批量导入Excel数据: 成功{imported_count}条, 失败{error_count}条'
        )
        
        result = {
            'status': 'success',
            'message': f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条空行，失败 {error_count} 条',
            'imported_count': imported_count,
            'error_count': error_count,
            'skipped_count': skipped_count
        }
        
        logger.info(f'📊 导入统计: 成功 {imported_count} 条，跳过 {skipped_count} 条，失败 {error_count} 条')
        
        if error_messages:
            result['error_messages'] = error_messages[:10]  # 只返回前10条错误信息
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f'导入大塬QC报表Excel失败: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'导入失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def changfu_report_import_excel(request):
    """长富QC报表Excel导入功能"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 检查是否有上传的文件
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': '请选择要导入的Excel文件'}, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': '仅支持Excel文件格式(.xlsx, .xls)'}, status=400)
        
        # 尝试导入pandas，如果失败则使用openpyxl
        try:
            import pandas as pd
            use_pandas = True
        except ImportError:
            use_pandas = False
        
        # 读取Excel文件
        if use_pandas:
            try:
                df = pd.read_excel(excel_file, sheet_name=0)
                # 删除完全空白的行
                df = df.dropna(how='all')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'读取Excel文件失败: {str(e)}'}, status=400)
        else:
            # 使用openpyxl读取
            from openpyxl import load_workbook
            try:
                wb = load_workbook(excel_file, data_only=True)
                ws = wb.active
                # 读取表头
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value if cell.value else '')
                
                # 处理特殊列名情况
                # 1. 处理重复的列名（如两个Permeability或两个渗透率）
                permeability_indices = []
                for i, header in enumerate(headers):
                    if header and ('Permeability' in str(header) or '渗透率' in str(header) or '滤速率' in str(header)):
                        permeability_indices.append(i)
                
                # 如果有两个渗透率列，重命名它们
                if len(permeability_indices) == 2:
                    headers[permeability_indices[0]] = 'Permeability_1'  # 第一个是远通渗透率
                    headers[permeability_indices[1]] = 'Permeability_2'  # 第二个是长富渗透率
                
                # 2. 处理"入窑前"和"含量 (%)"合并的情况
                # 如果找到"入窑前"列，检查下一列是否是"含量 (%)"
                for i, header in enumerate(headers):
                    if header and '入窑前' in str(header):
                        if i + 1 < len(headers) and headers[i + 1] and '含量' in str(headers[i + 1]):
                            # 将"含量 (%)"列也映射到alkali_content
                            headers[i + 1] = '入窑前碱含量(%)_content'
                        break
                
                # 3. 处理合并列"铁离子 (mg/钙离子 (mg/铝离子 (mg/白度"
                # 这种合并列需要在读取数据时特殊处理
                for i, header in enumerate(headers):
                    if header and '铁离子' in str(header) and '钙离子' in str(header):
                        # 这是一个合并列，需要拆分
                        headers[i] = '铁离子_钙离子_铝离子_白度_合并列'
                        break
                
                # 4. 处理合并列"电导值 (as/c pH"
                for i, header in enumerate(headers):
                    if header and '电导值' in str(header) and 'pH' in str(header):
                        # 这是一个合并列
                        headers[i] = '电导值_pH_合并列'
                        break
                
                # 读取数据行，跳过完全空白的行
                df_data = []
                # 记录原始列名，用于调试
                logger.info(f'📋 Excel原始列名: {headers}')
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # 检查是否是完全空白的行
                    if any(cell is not None and str(cell).strip() != '' for cell in row):
                        row_dict = {}
                        for i, header in enumerate(headers):
                            if i < len(row):
                                val = row[i]
                                # 处理空值和空字符串
                                if val is None or (isinstance(val, str) and val.strip() == ''):
                                    row_dict[header] = None
                                else:
                                    row_dict[header] = val
                            else:
                                row_dict[header] = None
                        df_data.append(row_dict)
                
                # 创建一个简单的DataFrame模拟对象
                class SimpleDF:
                    def __init__(self, data, columns):
                        self.data = data
                        self.columns = columns
                    
                    def iterrows(self):
                        for idx, row_dict in enumerate(self.data):
                            # 创建一个简单的行对象，支持get方法
                            class Row:
                                def __init__(self, data):
                                    self._data = data
                                
                                def get(self, key, default=None):
                                    return self._data.get(key, default)
                                
                                def __getitem__(self, key):
                                    return self._data.get(key)
                            
                            yield idx, Row(row_dict)
                    
                    def __len__(self):
                        return len(self.data)
                
                df = SimpleDF(df_data, headers)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'读取Excel文件失败: {str(e)}'}, status=400)
        
        # 定义列名映射关系（支持中英文）
        # 注意：根据新的Excel表头格式进行映射
        column_mapping = {
            # 基本字段映射
            '日期': 'date', 'Date': 'date', '检测日期': 'date',
            '时间': 'time', 'Time': 'time', '检测时间': 'time',
            '班次': 'shift', 'Shift': 'shift', 'Squad': 'shift', '班组': 'shift', '班别': 'shift',
            '产品名称': 'product_name', 'Product Name': 'product_name', 'Grade': 'product_name', 
            '产品型号': 'product_name',
            '包装类型': 'packaging', 'Packaging': 'packaging', 'IPKP CODE': 'packaging',
            '批号': 'batch_number', 'Batch Number': 'batch_number', '批次号': 'batch_number', 
            'LOT': 'batch_number', '批号/日期': 'batch_number', '批次': 'batch_number',
            
            # 检测数据字段映射 - 新表头格式
            '烘干后原土水分 (%)': 'moisture_after_drying',
            '烘干后原土水分（%）': 'moisture_after_drying',  # 中文括号
            '烘干后原土水分(%)': 'moisture_after_drying',
            '干燥后原土水分(%)': 'moisture_after_drying',
            '干燥后原土水分（%）': 'moisture_after_drying',  # 中文括号
            '烘干后原': 'moisture_after_drying',
            'Moisture after drying': 'moisture_after_drying', 
            '干燥后原土水分': 'moisture_after_drying',
            
            # 入窑前碱含量 - 新表头可能是分开的两列"入窑前"和"含量 (%)"，需要合并处理
            '入窑前碱含量(%)': 'alkali_content', 
            '入窑前碱含量': 'alkali_content',
            '入窑前': 'alkali_content',  # 如果只有"入窑前"列，也映射到alkali_content
            '含量 (%)': 'alkali_content',  # 如果"含量 (%)"单独存在
            'Alkali content (%)': 'alkali_content',
            
            '助剂添加比例': 'flux',
            '助溶剂添加比例': 'flux', 
            '助溶剂': 'flux', 
            '*flux agent': 'flux', 
            'flux agent addition ratio': 'flux',
            
            # 渗透率 - 新表头格式
            '运通滤速率': 'permeability',  # 运通 = 远通
            '远通渗透率(Darcy)': 'permeability', 
            '远通渗透率': 'permeability',
            '长高滤速率': 'permeability_long',  # 长高 = 长富
            '长富渗透率(Darcy)': 'permeability_long', 
            '长富渗透率': 'permeability_long',
            # 注意：如果Excel中有两个"Permeability"列，需要通过列位置区分
            # 第一个Permeability对应远通渗透率，第二个对应长富渗透率
            
            # 可塑度可能是涡值
            '可塑度 (c/cm)': 'swirl',
            '可塑度': 'swirl',
            '涡值(cm)': 'swirl',
            '涡值（cm）': 'swirl',  # 中文括号
            '涡值': 'swirl', 
            'Swirl (cm)': 'swirl', 
            'Swirl': 'swirl',
            
            # 饼密度和振实密度
            '饼密度(g/cm3)': 'wet_cake_density', 
            '饼密度（g/cm3）': 'wet_cake_density',  # 中文括号
            '饼密度': 'wet_cake_density', 
            'Wet cake density': 'wet_cake_density',
            '振实密度(g/cm3)': 'bulk_density',
            '振实密度（g/cm3)': 'bulk_density',  # 中文括号
            '振实密度（g/cm3）': 'bulk_density',  # 中文括号
            '振实密度': 'bulk_density',
            '重度 (k) 14W': 'bulk_density',  # 可能是振实密度的另一种表示
            '灰值 (c/m)': 'bulk_density',  # 灰值可能是振实密度
            
            # 白度
            '白度': 'brightness', 
            'Bri.': 'brightness', 
            'Brightness': 'brightness',
            
            '气味': 'odor', 
            'Odor': 'odor',
            
            # 电导值和pH - 新表头可能是合并的"电导值 (as/c pH"
            '电导值 (as/c pH': 'conductance',  # 如果合并列，优先映射到电导值
            '电导值(ms/cm)': 'conductance', 
            '电导值': 'conductance', 
            'Conductance (ms/c)': 'conductance', 
            'Conductance': 'conductance',
            'pH': 'ph', 
            'pH值': 'ph',
            
            '水分(%)': 'moisture', 
            '水分': 'moisture', 
            'Moisture (%)': 'moisture', 
            'Moisture': 'moisture',
            
            '批数': 'bags',  # 新表头中的"批数"可能是"袋数"
            '袋数': 'bags', 
            'Bags': 'bags',
            
            '吨': 'tons', 
            'Tons': 'tons', 
            '产量': 'tons',
            
            # 筛分数据字段映射 - 新表头格式（没有+号）
            '14W': 'sieving_14m',  # 重度 (k) 14W 中的14W
            '+14M (%)': 'sieving_14m', 
            '+14M': 'sieving_14m', 
            '14M': 'sieving_14m',
            '+30M (%)': 'sieving_30m', 
            '+30M': 'sieving_30m', 
            '30M': 'sieving_30m',
            '+40M (%)': 'sieving_40m', 
            '+40M': 'sieving_40m', 
            '40M': 'sieving_40m',
            'M': 'sieving_40m',  # 单独的M可能是40M
            '+80M (%)': 'sieving_80m', 
            '+80M': 'sieving_80m', 
            '80M': 'sieving_80m',
            '+100M (%)': 'sieving_100m', 
            '+100M': 'sieving_100m', 
            '100M': 'sieving_100m',
            '+150M (%)': 'sieving_150m', 
            '+150M': 'sieving_150m',
            '150M': 'sieving_150m',
            '150M ': 'sieving_150m',  # 带尾随空格
            '+200M (%)': 'sieving_200m', 
            '+200M': 'sieving_200m', 
            '200M': 'sieving_200m',
            '200M ': 'sieving_200m',  # 带尾随空格
            '+325M (%)': 'sieving_325m', 
            '+325M': 'sieving_325m', 
            '325M': 'sieving_325m',
            '325M ': 'sieving_325m',  # 带尾随空格
            
            # 离子数据字段映射 - 新表头可能是合并列"铁离子 (mg/钙离子 (mg/铝离子 (mg/白度"
            # 这种情况需要在读取时特殊处理，这里先提供单独列的映射
            '铁离子 (mg/钙离子 (mg/铝离子 (mg/白度': 'fe_ion',  # 合并列，优先取第一个
            '铁离子（mg/kg）': 'fe_ion',  # 中文括号
            '铁离子(mg/kg)': 'fe_ion',
            '铁离子': 'fe_ion', 
            'Fe离子': 'fe_ion', 
            'Fe': 'fe_ion',
            '钙离子（mg/kg）': 'ca_ion',  # 中文括号
            '钙离子(mg/kg)': 'ca_ion',
            '钙离子': 'ca_ion',
            'Ca离子': 'ca_ion', 
            'Ca': 'ca_ion',
            '铝离子（mg/kg）': 'al_ion',  # 中文括号
            '铝离子(mg/kg)': 'al_ion',
            '铝离子': 'al_ion',
            'Al离子': 'al_ion', 
            'Al': 'al_ion',
            
            '吸油率 (%)': 'oil_absorption',
            '吸油率（%）': 'oil_absorption',  # 中文括号
            '吸油量': 'oil_absorption', 
            '吸油率(%)': 'oil_absorption',
            '吸水率 (%)': 'water_absorption',
            '吸水率（%）': 'water_absorption',  # 中文括号
            '吸水量': 'water_absorption', 
            '吸水率(%)': 'water_absorption',
            
            '水分(%)': 'moisture',
            '水分（%）': 'moisture',  # 中文括号
            '水分   （%）': 'moisture',  # 多个空格和中文括号
            '水分': 'moisture', 
            'Moisture (%)': 'moisture', 
            'Moisture': 'moisture',
            
            '备注': 'remarks', 
            'Remarks': 'remarks', 
            'Notes': 'remarks'
        }
        
        # 处理数据并导入
        imported_count = 0
        error_count = 0
        error_messages = []
        skipped_count = 0  # 记录跳过的行数
        
        # 统一处理逻辑：先映射列名，再处理数据
        # 如果是pandas DataFrame，先重命名列
        if use_pandas and hasattr(df, 'rename'):
            # 处理重复的列名（如两个Permeability）
            # 找到所有Permeability列
            permeability_cols = [col for col in df.columns if col and ('Permeability' in str(col) or '渗透率' in str(col) or '滤速率' in str(col))]
            if len(permeability_cols) == 2:
                # 重命名第一个和第二个
                df.columns = [f'Permeability_1' if col == permeability_cols[0] else 
                             f'Permeability_2' if col == permeability_cols[1] else col 
                             for col in df.columns]
            
            # 添加Permeability_1和Permeability_2到映射
            column_mapping['Permeability_1'] = 'permeability'
            column_mapping['Permeability_2'] = 'permeability_long'
            
            # 改进的列名映射：先尝试精确匹配，再尝试模糊匹配
            def normalize_col_name(col_name):
                """标准化列名，去除空格、统一括号等"""
                if col_name is None:
                    return ''
                col_str = str(col_name).strip()
                # 统一括号：将中文括号转换为英文括号
                col_str = col_str.replace('（', '(').replace('）', ')')
                # 去除多余空格
                col_str = ' '.join(col_str.split())
                return col_str
            
            # 创建标准化后的映射字典
            normalized_mapping = {}
            for orig_key, mapped_value in column_mapping.items():
                normalized_key = normalize_col_name(orig_key)
                normalized_mapping[normalized_key] = mapped_value
            
            # 先进行精确匹配映射
            rename_dict = {}
            unmapped_cols = []
            
            for col in df.columns:
                col_normalized = normalize_col_name(col)
                # 先尝试精确匹配
                if col in column_mapping:
                    rename_dict[col] = column_mapping[col]
                elif col_normalized in normalized_mapping:
                    rename_dict[col] = normalized_mapping[col_normalized]
                else:
                    # 尝试模糊匹配
                    matched = False
                    for orig_key, mapped_value in column_mapping.items():
                        orig_normalized = normalize_col_name(orig_key)
                        # 检查是否包含关键部分
                        if orig_normalized and col_normalized:
                            # 去除括号和空格后比较
                            orig_simple = orig_normalized.replace(' ', '').replace('(', '').replace(')', '').lower()
                            col_simple = col_normalized.replace(' ', '').replace('(', '').replace(')', '').lower()
                            if orig_simple == col_simple or (len(orig_simple) > 3 and orig_simple in col_simple):
                                rename_dict[col] = mapped_value
                                matched = True
                                break
                    
                    if not matched:
                        unmapped_cols.append(col)
            
            # 执行重命名
            df_mapped = df.rename(columns=rename_dict)
            
            # 记录未映射的列名（用于调试）
            if unmapped_cols:
                logger.debug(f'⚠️ 未映射的列名: {unmapped_cols}')
            
            logger.info(f'📊 读取到 {len(df_mapped)} 行数据，原始列名: {list(df.columns)}, 映射后列名: {list(df_mapped.columns)}')
        else:
            # 对于openpyxl，需要手动映射
            # 添加Permeability_1和Permeability_2到映射
            column_mapping['Permeability_1'] = 'permeability'
            column_mapping['Permeability_2'] = 'permeability_long'
            
            df_mapped = df
            logger.info(f'📊 读取到 {len(df_mapped)} 行数据，列名: {df_mapped.columns if hasattr(df_mapped, "columns") else "未知"}')
        
        # 处理每一行数据
        for index, row_obj in df_mapped.iterrows():
            try:
                # 统一获取行数据的方法
                if use_pandas:
                    row = row_obj
                else:
                    row = row_obj
                
                # 获取行数据的辅助函数
                def get_row_value(key):
                    """从行中获取值，支持pandas和openpyxl两种格式"""
                    if use_pandas:
                        # pandas格式 - 列名已经通过rename映射过了，直接用key（数据库字段名）获取
                        try:
                            # 首先尝试直接通过映射后的列名获取
                            if hasattr(row, 'get'):
                                val = row.get(key)
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    return val
                            if hasattr(row, key):
                                val = getattr(row, key)
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    return val
                            if hasattr(row, 'index') and key in row.index:
                                val = row[key]
                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                    return val
                            
                            # 如果直接获取不到，尝试通过原始列名获取（模糊匹配）
                            # 定义字段的关键词映射
                            field_keywords = {
                                'date': ['日期', 'date', 'Date', '检测日期'],
                                'time': ['时间', 'time', 'Time', '检测时间'],
                                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                                'permeability': ['远通', '渗透率', 'permeability', 'Darcy', 'Permeability_1'],
                                'permeability_long': ['长富', '长高', '渗透率', 'permeability', 'Permeability_2'],
                                'wet_cake_density': ['饼密度', 'cake', 'density'],
                                'bulk_density': ['振实密度', 'bulk', 'density'],
                                'brightness': ['白度', 'brightness', 'Bri'],
                                'swirl': ['涡值', '可塑度', 'swirl', 'Swirl'],
                                'odor': ['气味', 'odor'],
                                'conductance': ['电导值', 'conductance'],
                                'ph': ['pH', 'ph'],
                                'moisture': ['水分', 'moisture'],
                                'bags': ['袋数', 'bags', '批数'],
                                'tons': ['吨', 'tons', '产量'],
                                'fe_ion': ['铁离子', 'Fe', 'fe'],
                                'ca_ion': ['钙离子', 'Ca', 'ca'],
                                'al_ion': ['铝离子', 'Al', 'al'],
                                'oil_absorption': ['吸油', 'oil'],
                                'water_absorption': ['吸水', 'water'],
                                'sieving_14m': ['14M', '14', '+14'],
                                'sieving_30m': ['30M', '30', '+30'],
                                'sieving_40m': ['40M', '40', '+40'],
                                'sieving_80m': ['80M', '80', '+80'],
                                'sieving_100m': ['100M', '100', '+100'],
                                'sieving_150m': ['150M', '150', '+150'],
                                'sieving_200m': ['200M', '200', '+200'],
                                'sieving_325m': ['325M', '325', '+325'],
                            }
                            
                            # 获取当前字段的关键词
                            keywords = field_keywords.get(key, [])
                            
                            # 遍历行的所有列（使用row.index获取列名）
                            if hasattr(row, 'index'):
                                # 优先匹配更精确的关键词（中文关键词优先，更长的关键词优先）
                                # 对关键词按优先级排序：中文关键词 > 英文关键词，长关键词 > 短关键词
                                sorted_keywords = sorted(keywords, key=lambda k: (k.isascii(), -len(k)))
                                
                                for keyword in sorted_keywords:
                                    best_match_col = None
                                    best_match_val = None
                                    
                                    for col_name in row.index:
                                        if col_name is None:
                                            continue
                                        col_str = str(col_name).strip()
                                        
                                        # 检查是否包含关键词
                                        if keyword in col_str:
                                            try:
                                                val = row[col_name]
                                                # 检查值是否为空
                                                if val is not None and (not pd.isna(val) if hasattr(pd, 'isna') else True):
                                                    # 找到匹配的列，记录最佳匹配
                                                    if best_match_col is None:
                                                        best_match_col = col_name
                                                        best_match_val = val
                                                    # 如果当前列名更精确（包含更多关键词），优先选择
                                                    elif len(col_str) < len(str(best_match_col)):
                                                        best_match_col = col_name
                                                        best_match_val = val
                                            except (KeyError, IndexError):
                                                continue
                                    
                                    # 如果找到最佳匹配，返回该值
                                    if best_match_val is not None:
                                        return best_match_val
                                    
                                    # 如果当前关键词找到了列但值为空，记录但不继续匹配其他关键词
                                    # （因为已经找到了正确的列，只是值为空）
                                    if best_match_col is not None:
                                        return None
                            
                            # 如果没有找到匹配的值，返回None
                            return None
                        except Exception as e:
                            logger.debug(f'获取字段 {key} 失败: {str(e)}')
                            return None
                    else:
                        # openpyxl格式，row是字典，key是Excel原始列名
                        # 先尝试直接用key获取（如果列名就是数据库字段名）
                        if isinstance(row, dict):
                            if key in row:
                                val = row[key]
                                if val is not None:
                                    return val
                        
                        # 通过column_mapping查找所有可能映射到key的原始列名
                        # 遍历所有映射，找到mapped_col == key的orig_col
                        matched_cols = []
                        for orig_col, mapped_col in column_mapping.items():
                            if mapped_col == key:
                                matched_cols.append(orig_col)
                        
                        # 尝试从所有匹配的列名中获取值
                        for orig_col in matched_cols:
                            if isinstance(row, dict) and orig_col in row:
                                val = row[orig_col]
                                if val is not None:
                                    return val
                        
                        # 如果还是找不到，尝试模糊匹配（去除空格、括号等）
                        if isinstance(row, dict):
                            # 定义关键字段的匹配关键词
                            key_keywords = {
                                'moisture_after_drying': ['烘干后', '干燥后', '原土水分', 'moisture', 'after', 'drying'],
                                'permeability': ['远通', '渗透率', 'permeability', 'Darcy'],
                                'permeability_long': ['长富', '长高', '渗透率', 'permeability'],
                                'swirl': ['涡值', '可塑度', 'swirl', 'Swirl'],
                                'sieving_150m': ['150M', '150', '+150'],
                                'sieving_200m': ['200M', '200', '+200'],
                                'sieving_325m': ['325M', '325', '+325'],
                                'fe_ion': ['铁离子', 'Fe', 'fe'],
                                'ca_ion': ['钙离子', 'Ca', 'ca'],
                                'al_ion': ['铝离子', 'Al', 'al'],
                                'oil_absorption': ['吸油', 'oil'],
                                'water_absorption': ['吸水', 'water'],
                                'moisture': ['水分', 'moisture'],
                            }
                            
                            # 获取当前字段的关键词
                            keywords = key_keywords.get(key, [])
                            
                            # 遍历所有列名，查找匹配的
                            for orig_col in row.keys():
                                if orig_col is None:
                                    continue
                                orig_col_str = str(orig_col).strip()
                                
                                # 如果有关键词，检查是否包含关键词
                                if keywords:
                                    for keyword in keywords:
                                        if keyword in orig_col_str:
                                            val = row[orig_col]
                                            if val is not None:
                                                return val
                                else:
                                    # 没有关键词，使用标准化匹配
                                    key_normalized = str(key).strip().replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                                    orig_col_normalized = orig_col_str.replace(' ', '').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace('+', '').lower()
                                    # 检查是否包含关键部分
                                    if key_normalized in orig_col_normalized or orig_col_normalized in key_normalized:
                                        val = row[orig_col]
                                        if val is not None:
                                            return val
                        
                        return None
                
                def is_notna(val):
                    """检查值是否不为空"""
                    if use_pandas:
                        return pd.notna(val)
                    else:
                        return val is not None and val != '' and str(val).strip() != ''
                
                # 检查是否为空行 - 检查关键字段
                date_val = get_row_value('date')
                product_name_val = get_row_value('product_name')
                
                # 放宽空行检查：只要日期或产品名称有一个不为空，就继续处理
                # 或者有其他任何字段有值
                has_any_data = False
                if date_val and is_notna(date_val):
                    has_any_data = True
                if product_name_val and is_notna(product_name_val):
                    has_any_data = True
                
                # 如果日期和产品名称都为空，再检查其他字段
                if not has_any_data:
                    for key in ['shift', 'packaging', 'bags', 'batch_number', 'moisture_after_drying', 
                               'alkali_content', 'permeability', 'permeability_long', 'wet_cake_density']:
                        val = get_row_value(key)
                        if val and is_notna(val):
                            has_any_data = True
                            break
                    
                    # 如果没有任何数据，跳过这一行
                    if not has_any_data:
                        skipped_count += 1
                        continue
                
                data = {}
                
                # 处理日期
                if date_val and is_notna(date_val):
                    if isinstance(date_val, str):
                        try:
                            data['date'] = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                        except:
                            try:
                                # 尝试其他日期格式
                                data['date'] = datetime.strptime(date_val.strip(), '%Y/%m/%d').date()
                            except:
                                data['date'] = date.today()
                    elif hasattr(date_val, 'date'):
                        data['date'] = date_val.date()
                    else:
                        data['date'] = date.today()
                else:
                    data['date'] = date.today()
                
                # 处理时间
                time_val = get_row_value('time')
                if time_val and is_notna(time_val):
                    try:
                        # 如果是字符串类型
                        if isinstance(time_val, str):
                            time_str = str(time_val).strip()
                            if not time_str:
                                data['time'] = time(0, 0)
                            else:
                                try:
                                    # 尝试 HH:MM 格式
                                    data['time'] = datetime.strptime(time_str, '%H:%M').time()
                                except ValueError:
                                    try:
                                        # 尝试 HH:MM:SS 格式
                                        data['time'] = datetime.strptime(time_str, '%H:%M:%S').time()
                                    except ValueError:
                                        try:
                                            # 尝试处理Excel时间格式（小数，如0.4166666666666667表示10:00）
                                            if ':' not in time_str and ('.' in time_str or time_str.replace('.', '').replace('-', '').isdigit()):
                                                time_float = float(time_str)
                                                # Excel时间格式：0.0 = 00:00:00, 0.5 = 12:00:00, 1.0 = 24:00:00
                                                total_seconds = int(time_float * 24 * 3600)
                                                hours = total_seconds // 3600
                                                minutes = (total_seconds % 3600) // 60
                                                seconds = total_seconds % 60
                                                data['time'] = time(hours, minutes, seconds)
                                            else:
                                                # 如果都不匹配，尝试直接解析数字（如1000表示10:00）
                                                if time_str.isdigit() and len(time_str) <= 4:
                                                    # 假设是HHMM格式
                                                    hours = int(time_str[:2]) if len(time_str) >= 2 else int(time_str[0])
                                                    minutes = int(time_str[-2:]) if len(time_str) >= 2 else 0
                                                    data['time'] = time(hours, minutes)
                                                else:
                                                    data['time'] = time(0, 0)
                                        except (ValueError, TypeError) as e:
                                            data['time'] = time(0, 0)
                        # 如果是datetime对象
                        elif hasattr(time_val, 'time'):
                            data['time'] = time_val.time()
                        # 如果是time对象
                        elif isinstance(time_val, time):
                            data['time'] = time_val
                        # 如果是datetime对象（pandas可能返回）
                        elif hasattr(time_val, 'hour') and hasattr(time_val, 'minute'):
                            # 可能是pandas的Timestamp对象
                            data['time'] = time(time_val.hour, time_val.minute, getattr(time_val, 'second', 0))
                        # 如果是数字（Excel时间格式）
                        elif isinstance(time_val, (int, float)):
                            # Excel时间格式：0.0 = 00:00:00, 0.5 = 12:00:00
                            total_seconds = int(time_val * 24 * 3600)
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            seconds = total_seconds % 60
                            data['time'] = time(hours, minutes, seconds)
                        else:
                            data['time'] = time(0, 0)
                    except Exception as e:
                        data['time'] = time(0, 0)
                else:
                    data['time'] = time(0, 0)
                
                # 处理字符串字段
                string_fields = ['shift', 'product_name', 'packaging', 'batch_number', 'flux', 'remarks']
                for field in string_fields:
                    val = get_row_value(field)
                    if val and is_notna(val):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 特殊处理：处理"入窑前碱含量"
                alkali_val = get_row_value('alkali_content')
                if alkali_val and is_notna(alkali_val):
                    try:
                        data['alkali_content'] = float(alkali_val)
                    except:
                        data['alkali_content'] = None
                else:
                    data['alkali_content'] = None
                
                # 处理数字字段
                numeric_fields = [
                    'moisture_after_drying', 'permeability', 'permeability_long',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor',
                    'conductance', 'ph', 'moisture', 'bags', 'tons', 'fe_ion', 'ca_ion',
                    'al_ion', 'oil_absorption', 'water_absorption', 'sieving_14m', 'sieving_30m',
                    'sieving_40m', 'sieving_80m'
                ]
                # 注意：alkali_content已经在上面单独处理了
                
                for field in numeric_fields:
                    val = get_row_value(field)
                    # 严格检查值是否有效
                    if val is not None and is_notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.strip()
                                # 空字符串或只包含空白字符的字符串视为空值
                                if val == '' or val.isspace():
                                    data[field] = None
                                else:
                                    # 尝试转换为浮点数
                                    float_val = float(val)
                                    # 检查是否是NaN或Inf
                                    if pd.isna(float_val) if hasattr(pd, 'isna') else (float_val != float_val or abs(float_val) == float('inf')):
                                        data[field] = None
                                    else:
                                        data[field] = float_val
                            elif isinstance(val, (int, float)):
                                # 检查是否是NaN或Inf
                                if pd.isna(val) if hasattr(pd, 'isna') else (val != val or abs(val) == float('inf')):
                                    data[field] = None
                                else:
                                    data[field] = float(val)
                            else:
                                # 其他类型，尝试转换
                                data[field] = float(val)
                        except (ValueError, TypeError) as e:
                            data[field] = None
                    else:
                        # 值为None或空，设置为None
                        data[field] = None
                
                # 特殊处理：处理合并列"铁离子/钙离子/铝离子/白度"
                # 如果找到合并列，尝试拆分
                if not use_pandas:
                    merged_ion_col = None
                    if hasattr(row, 'get'):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    elif isinstance(row, dict):
                        merged_ion_col = row.get('铁离子_钙离子_铝离子_白度_合并列')
                    
                    if merged_ion_col and is_notna(merged_ion_col):
                        # 尝试从合并列中提取值（如果值是字符串，可能需要解析）
                        # 这里假设合并列的值可能是用斜杠分隔的
                        merged_str = str(merged_ion_col)
                        if '/' in merged_str:
                            parts = merged_str.split('/')
                            if len(parts) >= 3:
                                try:
                                    if parts[0].strip():
                                        data['fe_ion'] = float(parts[0].strip())
                                    if parts[1].strip():
                                        data['ca_ion'] = float(parts[1].strip())
                                    if parts[2].strip():
                                        data['al_ion'] = float(parts[2].strip())
                                    if len(parts) >= 4 and parts[3].strip():
                                        data['brightness'] = float(parts[3].strip())
                                except:
                                    pass
                
                # 特殊处理：处理合并列"电导值/pH"
                if not use_pandas:
                    merged_conductance_col = None
                    if hasattr(row, 'get'):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    elif isinstance(row, dict):
                        merged_conductance_col = row.get('电导值_pH_合并列')
                    
                    if merged_conductance_col and is_notna(merged_conductance_col):
                        # 尝试从合并列中提取值
                        merged_str = str(merged_conductance_col)
                        if '/' in merged_str or 'pH' in merged_str:
                            # 尝试解析
                            parts = merged_str.replace('pH', '').split('/')
                            if len(parts) >= 1 and parts[0].strip():
                                try:
                                    data['conductance'] = float(parts[0].strip())
                                except:
                                    pass
                            # pH值可能在字符串的其他位置
                            import re
                            ph_match = re.search(r'pH[:\s]*([0-9.]+)', merged_str, re.IGNORECASE)
                            if ph_match:
                                try:
                                    data['ph'] = float(ph_match.group(1))
                                except:
                                    pass
                
                # 处理筛分字段（可能是字符串）
                sieving_fields = ['sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m']
                for field in sieving_fields:
                    val = get_row_value(field)
                    if val and is_notna(val):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                    
                # 再次检查是否为空行 - 放宽条件，只要有日期或任何字段有值就保留
                has_valid_data = False
                # 检查日期（必须有）
                if data.get('date'):
                    has_valid_data = True
                # 检查其他字段
                if data.get('product_name') and str(data['product_name']).strip():
                    has_valid_data = True
                if data.get('shift') and str(data['shift']).strip():
                    has_valid_data = True
                if data.get('packaging') and str(data['packaging']).strip():
                    has_valid_data = True
                # 检查数字字段是否有值
                for field in ['bags', 'moisture_after_drying', 'alkali_content', 'permeability', 
                             'permeability_long', 'wet_cake_density', 'brightness', 'swirl']:
                    if data.get(field) is not None:
                        has_valid_data = True
                        break
                
                # 如果没有有效数据，跳过这一行
                if not has_valid_data:
                    skipped_count += 1
                    continue
                
                # 数据校验：对每个字段进行校验
                validation_errors = []
                from home.models import ChangfuQCReport
                
                # 定义字段显示名称映射
                field_display_names = {
                    'date': '日期',
                    'time': '时间',
                    'shift': '班次',
                    'product_name': '产品名称',
                    'packaging': '包装类型',
                    'batch_number': '批号',
                    'moisture_after_drying': '干燥后原土水分(%)',
                    'alkali_content': '入窑前碱含量(%)',
                    'flux': '助溶剂添加比例',
                    'permeability': '远通渗透率(Darcy)',
                    'permeability_long': '长富渗透率(Darcy)',
                    'wet_cake_density': '饼密度(g/cm3)',
                    'filter_time': '过滤时间(秒)',
                    'water_viscosity': '水黏度(mPa.s)',
                    'cake_thickness': '饼厚(mm)',
                    'bulk_density': '振实密度(g/cm3)',
                    'brightness': '白度',
                    'swirl': '涡值(cm)',
                    'odor': '气味',
                    'conductance': '电导值(ms/cm)',
                    'ph': 'pH',
                    'moisture': '水分(%)',
                    'bags': '袋数',
                    'tons': '吨',
                    'sieving_14m': '+14M (%)',
                    'sieving_30m': '+30M (%)',
                    'sieving_40m': '+40M (%)',
                    'sieving_80m': '+80M (%)',
                    'sieving_100m': '+100M (%)',
                    'sieving_150m': '+150M (%)',
                    'sieving_200m': '+200M (%)',
                    'sieving_325m': '+325M (%)',
                    'fe_ion': 'Fe离子',
                    'ca_ion': 'Ca离子',
                    'al_ion': 'Al离子',
                    'oil_absorption': '吸油量',
                    'water_absorption': '吸水量',
                    'remarks': '备注',
                }
                
                # 对每个字段进行校验
                for field_name, field_value in data.items():
                    # 跳过用户相关字段
                    if field_name in ['user', 'username']:
                        continue
                    
                    # 获取字段显示名称
                    field_display_name = field_display_names.get(field_name, field_name)
                    
                    # 使用validate_field_by_model进行校验
                    is_valid, error_msg = validate_field_by_model(
                        ChangfuQCReport, 
                        field_name, 
                        field_value, 
                        field_display_name
                    )
                    
                    if not is_valid:
                        validation_errors.append(f'{field_display_name}: {error_msg}')
                
                # 如果有校验错误，记录并跳过这一行
                if validation_errors:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行数据校验失败: {"; ".join(validation_errors)}'
                    error_messages.append(error_msg)
                    logger.warning(f'导入长富QC报表数据校验失败: {error_msg}')
                    continue
                
                # 设置用户信息
                data['user'] = request.user
                data['username'] = request.user.username
                
                # 创建记录 - 使用ChangfuQCReport模型
                ChangfuQCReport.objects.create(**data)
                imported_count += 1
                
            except Exception as e:
                error_count += 1
                error_msg = f'第 {index + 2} 行导入失败: {str(e)}'
                error_messages.append(error_msg)
                logger.error(f'导入长富QC报表失败: {error_msg}', exc_info=True)
        
        # 记录操作日志
        from home.models import UserOperationLog
        UserOperationLog.log_operation(
            request, 'CREATE', 'changfu', None,
            f'批量导入Excel数据: 成功{imported_count}条, 失败{error_count}条'
        )
        
        result = {
            'status': 'success',
            'message': f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条空行，失败 {error_count} 条',
            'imported_count': imported_count,
            'error_count': error_count,
            'skipped_count': skipped_count
        }
        
        logger.info(f'📊 导入统计: 成功 {imported_count} 条，跳过 {skipped_count} 条，失败 {error_count} 条')
        
        if error_messages:
            result['error_messages'] = error_messages[:10]  # 只返回前10条错误信息
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f'导入长富QC报表Excel失败: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'导入失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def xinghui_report_import_excel(request):
    """兴辉QC报表Excel导入功能"""
    from home.excel_import_utils import import_xinghui_report_data
    from home.models import XinghuiQCReport
    return import_xinghui_report_data(request, XinghuiQCReport, '兴辉QC报表', 'xinghui')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def xinghui2_report_import_excel(request):
    """兴辉二线QC报表Excel导入功能"""
    from home.excel_import_utils import import_xinghui_report_data
    from home.models import Xinghui2QCReport
    return import_xinghui_report_data(request, Xinghui2QCReport, '兴辉二线QC报表', 'xinghui2')

@login_required
@permission_required('qc_report_edit')
def yuantong_report_download_template(request):
    """下载远通QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志：记录用户信息和下载请求
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],  # 限制长度
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载远通QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "远通QC报表导入模板"
        
        # 定义表头（按照历史记录页面的字段顺序，不包括操作人）
        # 字段顺序与 yuantong_report_history.html 中的表头顺序一致
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品型号',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '过滤时间(秒)',  # filter_time
            '水黏度(mPa.s)',  # water_viscosity
            '饼厚(mm)',  # cake_thickness
            '饼密度(g/cm3)',  # wet_cake_density
            '远通饼密度(g/cm3)',  # yuantong_cake_density
            '长富饼密度(g/cm3)',  # changfu_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M',  # sieving_14m
            '+30M',  # sieving_30m
            '+40M',  # sieving_40m
            '+80M',  # sieving_80m
            '+100M',  # sieving_100m
            '+150M',  # sieving_150m
            '+200M',  # sieving_200m
            '+325M',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油率(%)',  # oil_absorption
            '吸水率(%)',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨数',  # tons
            'LOT批号',  # batch_number
            '备注',  # remarks
            '班组',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（用于提示用户填写格式）
        example_row = [
            '2024-01-01',  # 日期格式示例
            '08:00',  # 时间格式示例
            '10.5',  # 干燥后原土水分（数字）
            '2.3',  # 入窑前碱含量（数字）
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品名称
            '5.5',  # 远通渗透率（数字）
            '6.0',  # 长富渗透率（数字）
            '120',  # 过滤时间（数字，秒）
            '1.002',  # 水黏度（数字，最多4位小数）
            '5.0',  # 饼厚（数字，mm）
            '1.5',  # 饼密度（数字）
            '1.6',  # 远通饼密度（数字）
            '1.7',  # 长富饼密度（数字）
            '1.8',  # 振实密度（数字）
            '5.0',  # +14M（数字）
            '10.0',  # +30M（数字）
            '15.0',  # +40M（数字）
            '20.0',  # +80M（数字）
            '25.0',  # +100M（文本或数字）
            '10.0',  # +150M（文本或数字）
            '5.0',  # +200M（文本或数字）
            '2.0',  # +325M（文本或数字）
            '50',  # Fe离子（数字）
            '30',  # Ca离子（数字）
            '20',  # Al离子（数字）
            '95.5',  # 白度（数字）
            '2.5',  # 涡值（数字，cm）
            '2.5',  # 气味（数字）
            '150',  # 电导值（数字）
            '7.0',  # pH（数字）
            '25.5',  # 吸油量（数字）
            '30.0',  # 吸水量（数字）
            '0.5',  # 水分（数字）
            '100',  # 袋数（数字）
            '包装类型示例',  # 包装类型
            '10.500',  # 吨（数字，最多3位小数）
            '批次号示例',  # 批号
            '备注信息',  # 备注（文本）
            '早班',  # 班次
        ]
        ws.append(example_row)
        
        # 设置示例行样式（浅灰色背景，提示这是示例）
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据。'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品名称
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 15,  # 过滤时间
            'J': 15,  # 水黏度
            'K': 12,  # 饼厚
            'L': 15,  # 饼密度
            'M': 18,  # 远通饼密度
            'N': 18,  # 长富饼密度
            'O': 15,  # 振实密度
            'P': 12,  # +14M
            'Q': 12,  # +30M
            'R': 12,  # +40M
            'S': 12,  # +80M
            'T': 12,  # +100M
            'U': 12,  # +150M
            'V': 12,  # +200M
            'W': 12,  # +325M
            'X': 12,  # Fe离子
            'Y': 12,  # Ca离子
            'Z': 12,  # Al离子
            'AA': 10,  # 白度
            'AB': 12,  # 涡值
            'AC': 10,  # 气味
            'AD': 15,  # 电导值
            'AE': 10,  # pH
            'AF': 12,  # 吸油量
            'AG': 12,  # 吸水量
            'AH': 12,  # 水分
            'AI': 10,  # 袋数
            'AJ': 15,  # 包装类型
            'AK': 12,  # 吨
            'AL': 15,  # 批号
            'AM': 30,  # 备注
            'AN': 10,  # 班次
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "远通QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        # 保存到响应
        wb.save(response)
        
        # 记录下载成功日志
        logger.info(f'[下载模板] 成功生成并返回远通QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        # 记录下载失败日志
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成远通QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def yuantong_report_import_excel(request):
    """远通QC报表Excel导入功能"""
    from home.excel_import_utils import import_yuantong_report_data
    from home.models import YuantongQCReport
    return import_yuantong_report_data(request, YuantongQCReport, '远通QC报表', 'yuantong')

@login_required
@permission_required('qc_report_edit')
def yuantong2_report_download_template(request):
    """下载远通二线QC报表Excel导入模板（完全参照远通QC报表导入模板）"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志：记录用户信息和下载请求
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],  # 限制长度
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载远通二线QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "远通二线QC报表导入模板"
        
        # 定义表头（按照历史记录页面的字段顺序，不包括操作人）
        # 字段顺序与 yuantong_report_history.html 中的表头顺序一致（完全参照远通QC报表导入模板）
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品型号',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '过滤时间(秒)',  # filter_time
            '水黏度(mPa.s)',  # water_viscosity
            '饼厚(mm)',  # cake_thickness
            '饼密度(g/cm3)',  # wet_cake_density
            '远通饼密度(g/cm3)',  # yuantong_cake_density
            '长富饼密度(g/cm3)',  # changfu_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M',  # sieving_14m
            '+30M',  # sieving_30m
            '+40M',  # sieving_40m
            '+80M',  # sieving_80m
            '+100M',  # sieving_100m
            '+150M',  # sieving_150m
            '+200M',  # sieving_200m
            '+325M',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油率(%)',  # oil_absorption
            '吸水率(%)',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨数',  # tons
            'LOT批号',  # batch_number
            '备注',  # remarks
            '班组',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（用于提示用户填写格式）
        example_row = [
            '2024-01-01',  # 日期格式示例
            '08:00',  # 时间格式示例
            '10.5',  # 干燥后原土水分（数字）
            '2.3',  # 入窑前碱含量（数字）
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品名称
            '5.5',  # 远通渗透率（数字）
            '6.0',  # 长富渗透率（数字）
            '120',  # 过滤时间（数字，秒）
            '1.002',  # 水黏度（数字，最多4位小数）
            '5.0',  # 饼厚（数字，mm）
            '1.5',  # 饼密度（数字）
            '1.6',  # 远通饼密度（数字）
            '1.7',  # 长富饼密度（数字）
            '1.8',  # 振实密度（数字）
            '5.0',  # +14M（数字）
            '10.0',  # +30M（数字）
            '15.0',  # +40M（数字）
            '20.0',  # +80M（数字）
            '25.0',  # +100M（文本或数字）
            '10.0',  # +150M（文本或数字）
            '5.0',  # +200M（文本或数字）
            '2.0',  # +325M（文本或数字）
            '50',  # Fe离子（数字）
            '30',  # Ca离子（数字）
            '20',  # Al离子（数字）
            '95.5',  # 白度（数字）
            '2.5',  # 涡值（数字，cm）
            '2.5',  # 气味（数字）
            '150',  # 电导值（数字）
            '7.0',  # pH（数字）
            '25.5',  # 吸油量（数字）
            '30.0',  # 吸水量（数字）
            '0.5',  # 水分（数字）
            '100',  # 袋数（数字）
            '包装类型示例',  # 包装类型
            '10.500',  # 吨（数字，最多3位小数）
            '批次号示例',  # 批号
            '备注信息',  # 备注（文本）
            '早班',  # 班次
        ]
        ws.append(example_row)
        
        # 设置示例行样式（浅灰色背景，提示这是示例）
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据。'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品名称
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 15,  # 过滤时间
            'J': 15,  # 水黏度
            'K': 12,  # 饼厚
            'L': 15,  # 饼密度
            'M': 18,  # 远通饼密度
            'N': 18,  # 长富饼密度
            'O': 15,  # 振实密度
            'P': 12,  # +14M
            'Q': 12,  # +30M
            'R': 12,  # +40M
            'S': 12,  # +80M
            'T': 12,  # +100M
            'U': 12,  # +150M
            'V': 12,  # +200M
            'W': 12,  # +325M
            'X': 12,  # Fe离子
            'Y': 12,  # Ca离子
            'Z': 12,  # Al离子
            'AA': 10,  # 白度
            'AB': 12,  # 涡值
            'AC': 10,  # 气味
            'AD': 15,  # 电导值
            'AE': 10,  # pH
            'AF': 12,  # 吸油量
            'AG': 12,  # 吸水量
            'AH': 12,  # 水分
            'AI': 10,  # 袋数
            'AJ': 15,  # 包装类型
            'AK': 12,  # 吨
            'AL': 15,  # 批号
            'AM': 30,  # 备注
            'AN': 10,  # 班次
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "远通二线QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        # 保存到响应
        wb.save(response)
        
        # 记录下载成功日志
        logger.info(f'[下载模板] 成功生成并返回远通二线QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        # 记录下载失败日志
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成远通二线QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def yuantong2_report_import_excel(request):
    """远通二线QC报表Excel导入功能"""
    from home.excel_import_utils import import_yuantong_report_data
    from home.models import Yuantong2QCReport
    return import_yuantong_report_data(request, Yuantong2QCReport, '远通二线QC报表', 'yuantong2')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def dongtai_report_download_template(request):
    """下载东泰QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志：记录用户信息和下载请求
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],  # 限制长度
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载东泰QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "东泰QC报表导入模板"
        
        # 定义表头（按照历史记录页面的字段顺序和名称，去掉东泰特有字段）
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品型号',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '过滤时间(秒)',  # filter_time
            '水黏度(mPa.s)',  # water_viscosity
            '饼厚(mm)',  # cake_thickness
            '饼密度(g/cm3)',  # wet_cake_density
            '远通饼密度(g/cm3)',  # yuantong_cake_density
            '长富饼密度(g/cm3)',  # changfu_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M',  # sieving_14m
            '+30M',  # sieving_30m
            '+40M',  # sieving_40m
            '+80M',  # sieving_80m
            '+100M',  # sieving_100m
            '+150M',  # sieving_150m
            '+200M',  # sieving_200m
            '+325M',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油率(%)',  # oil_absorption
            '吸水率(%)',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨数',  # tons
            'LOT批号',  # batch_number
            '备注',  # remarks
            '班组',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（用于提示用户填写格式，按照新的字段顺序）
        example_row = [
            '2024-01-01',  # 日期格式示例
            '08:00',  # 时间格式示例
            '10.5',  # 干燥后原土水分（数字）
            '2.3',  # 入窑前碱含量（数字）
            '助溶剂比例',  # 助溶剂添加比例
            '产品型号示例',  # 产品型号
            '5.5',  # 远通渗透率（数字）
            '6.0',  # 长富渗透率（数字）
            '120',  # 过滤时间（数字，秒）
            '1.0023',  # 水黏度（数字，最多4位小数）
            '5.0',  # 饼厚（数字）
            '1.5',  # 饼密度（数字）
            '1.52',  # 远通饼密度（数字）
            '1.50',  # 长富饼密度（数字）
            '1.8',  # 振实密度（数字）
            '5.0',  # +14M（数字）
            '10.0',  # +30M（数字）
            '15.0',  # +40M（数字）
            '20.0',  # +80M（数字）
            '25.0',  # +100M（文本或数字）
            '10.0',  # +150M（文本或数字）
            '5.0',  # +200M（文本或数字）
            '2.0',  # +325M（文本或数字）
            '50',  # Fe离子（数字）
            '30',  # Ca离子（数字）
            '20',  # Al离子（数字）
            '95.5',  # 白度（数字）
            '涡值示例',  # 涡值（文本）
            '2.5',  # 气味（数字）
            '150',  # 电导值（数字）
            '7.0',  # pH（数字）
            '25.5',  # 吸油率（数字）
            '30.0',  # 吸水率（数字）
            '0.5',  # 水分（数字）
            '100',  # 袋数（数字）
            '包装类型示例',  # 包装类型
            '10.500',  # 吨数（数字，最多3位小数）
            '批次号示例',  # LOT批号
            '备注信息',  # 备注（文本）
            '早班',  # 班组
        ]
        ws.append(example_row)
        
        # 设置示例行样式（浅灰色背景，提示这是示例）
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽（按照新的字段顺序，去掉东泰特有字段）
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品型号
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 12,  # 过滤时间
            'J': 15,  # 水黏度
            'K': 12,  # 饼厚
            'L': 15,  # 饼密度
            'M': 18,  # 远通饼密度
            'N': 18,  # 长富饼密度
            'O': 15,  # 振实密度
            'P': 12,  # +14M
            'Q': 12,  # +30M
            'R': 12,  # +40M
            'S': 12,  # +80M
            'T': 12,  # +100M
            'U': 12,  # +150M
            'V': 12,  # +200M
            'W': 12,  # +325M
            'X': 12,  # Fe离子
            'Y': 12,  # Ca离子
            'Z': 12,  # Al离子
            'AA': 10,  # 白度
            'AB': 12,  # 涡值
            'AC': 10,  # 气味
            'AD': 15,  # 电导值
            'AE': 10,  # pH
            'AF': 12,  # 吸油率
            'AG': 12,  # 吸水率
            'AH': 12,  # 水分
            'AI': 10,  # 袋数
            'AJ': 15,  # 包装类型
            'AK': 12,  # 吨数
            'AL': 15,  # LOT批号
            'AM': 30,  # 备注
            'AN': 10,  # 班组
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "东泰QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        # 使用 attachment 强制浏览器弹出保存对话框
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        # 添加额外的响应头，确保浏览器弹出保存对话框
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        # 保存到响应
        wb.save(response)
        
        # 记录下载成功日志
        logger.info(f'[下载模板] 成功生成并返回东泰QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        # 记录下载失败日志
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成东泰QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
def dayuan_report_download_template(request):
    """下载大塬QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志：记录用户信息和下载请求
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],  # 限制长度
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载大塬QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "大塬QC报表导入模板"
        
        # 定义表头（严格按照QC_REPORT_FIELD_MAPPING的顺序，与导出Excel完全一致）
        # 注意：
        # 1. username字段不包含在导入模板中（系统自动填充）
        # 2. 排除大塬不使用的字段：xinghui_permeability, yuantong_cake_density, changfu_cake_density, filter_time, water_viscosity, cake_thickness
        # 3. 字段顺序必须与导出Excel的字段顺序完全一致（按照QC_REPORT_FIELD_MAPPING的顺序）
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品名称',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            # 跳过 xinghui_permeability (大塬不使用)
            '饼密度(g/cm3)',  # wet_cake_density
            # 跳过 yuantong_cake_density (大塬不使用)
            # 跳过 changfu_cake_density (大塬不使用)
            # 跳过 filter_time (大塬不使用)
            # 跳过 water_viscosity (大塬不使用)
            # 跳过 cake_thickness (大塬不使用)
            '振实密度(g/cm3)',  # bulk_density
            '+14M (%)',  # sieving_14m
            '+30M (%)',  # sieving_30m
            '+40M (%)',  # sieving_40m
            '+80M (%)',  # sieving_80m
            '+100M (%)',  # sieving_100m
            '+150M (%)',  # sieving_150m
            '+200M (%)',  # sieving_200m
            '+325M (%)',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油量',  # oil_absorption
            '吸水量',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨',  # tons
            '批号',  # batch_number
            '备注',  # remarks
            '班次',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（用于提示用户填写格式，按照headers的顺序）
        example_row = [
            '2024-01-01',  # 日期格式示例
            '08:00',  # 时间格式示例
            '10.5',  # 干燥后原土水分（数字）
            '2.3',  # 入窑前碱含量（数字）
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品名称
            '5.5',  # 远通渗透率（数字）
            '6.0',  # 长富渗透率（数字）
            '1.5',  # 饼密度（数字）
            '1.8',  # 振实密度（数字）
            '5.0',  # +14M（数字）
            '10.0',  # +30M（数字）
            '15.0',  # +40M（数字）
            '20.0',  # +80M（数字）
            '25.0',  # +100M（文本或数字）
            '10.0',  # +150M（文本或数字）
            '5.0',  # +200M（文本或数字）
            '2.0',  # +325M（文本或数字）
            '50',  # Fe离子（数字）
            '30',  # Ca离子（数字）
            '20',  # Al离子（数字）
            '95.5',  # 白度（数字）
            '涡值示例',  # 涡值（文本或数字）
            '2.5',  # 气味（数字）
            '150',  # 电导值（数字）
            '7.0',  # pH（数字）
            '25.5',  # 吸油量（数字）
            '30.0',  # 吸水量（数字）
            '0.5',  # 水分（数字）
            '100',  # 袋数（数字）
            '包装类型示例',  # 包装类型
            '10.500',  # 吨（数字，最多3位小数）
            '批次号示例',  # 批号
            '备注信息',  # 备注（文本）
            '早班',  # 班次
        ]
        ws.append(example_row)
        
        # 设置示例行样式（浅灰色背景，提示这是示例）
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽（按照新的字段顺序，排除不包含的字段）
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品名称
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 15,  # 饼密度
            'J': 15,  # 振实密度
            'K': 12,  # +14M
            'L': 12,  # +30M
            'M': 12,  # +40M
            'N': 12,  # +80M
            'O': 12,  # +100M
            'P': 12,  # +150M
            'Q': 12,  # +200M
            'R': 12,  # +325M
            'S': 12,  # Fe离子
            'T': 12,  # Ca离子
            'U': 12,  # Al离子
            'V': 10,  # 白度
            'W': 12,  # 涡值
            'X': 10,  # 气味
            'Y': 15,  # 电导值
            'Z': 10,  # pH
            'AA': 12,  # 吸油量
            'AB': 12,  # 吸水量
            'AC': 12,  # 水分
            'AD': 10,  # 袋数
            'AE': 15,  # 包装类型
            'AF': 12,  # 吨
            'AG': 15,  # 批号
            'AH': 30,  # 备注
            'AI': 10,  # 班次
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "大塬QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        # 使用 attachment 强制浏览器弹出保存对话框
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        # 添加额外的响应头，确保浏览器弹出保存对话框
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        # 保存到响应
        wb.save(response)
        
        # 记录下载成功日志
        logger.info(f'[下载模板] 成功生成并返回大塬QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        # 记录下载失败日志
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成大塬QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
def changfu_report_download_template(request):
    """下载长富QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志：记录用户信息和下载请求
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],  # 限制长度
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载长富QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "长富QC报表导入模板"
        
        # 定义表头（按照历史记录字段顺序，去掉操作人字段）
        # 历史记录顺序：['username', 'date', 'time', 'moisture_after_drying', 'alkali_content', 'flux', 'product_name', 
        #                'permeability', 'permeability_long', 'wet_cake_density', 'bulk_density', 
        #                'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
        #                'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'brightness',
        #                'swirl', 'odor', 'conductance', 'ph', 'oil_absorption', 'water_absorption',
        #                'moisture', 'bags', 'packaging', 'tons', 'batch_number', 'remarks', 'shift']
        # 模板顺序（去掉username，与历史记录完全一致）：
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品名称',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '饼密度(g/cm3)',  # wet_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M (%)',  # sieving_14m
            '+30M (%)',  # sieving_30m
            '+40M (%)',  # sieving_40m
            '+80M (%)',  # sieving_80m
            '+100M (%)',  # sieving_100m
            '+150M (%)',  # sieving_150m
            '+200M (%)',  # sieving_200m
            '+325M (%)',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油量',  # oil_absorption
            '吸水量',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨',  # tons
            '批号',  # batch_number
            '备注',  # remarks
            '班次',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（用于提示用户填写格式，按照历史记录字段顺序，不包含过滤时间、水黏度、饼厚）
        example_row = [
            '2024-01-01',  # 日期格式示例
            '08:00',  # 时间格式示例
            '10.5',  # 干燥后原土水分（数字）
            '2.3',  # 入窑前碱含量（数字）
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品名称
            '5.5',  # 远通渗透率（数字）
            '6.0',  # 长富渗透率（数字）
            '1.5',  # 饼密度（数字）
            '1.8',  # 振实密度（数字）
            '5.0',  # +14M（数字）
            '10.0',  # +30M（数字）
            '15.0',  # +40M（数字）
            '20.0',  # +80M（数字）
            '25.0',  # +100M（文本或数字）
            '10.0',  # +150M（文本或数字）
            '5.0',  # +200M（文本或数字）
            '2.0',  # +325M（文本或数字）
            '50',  # Fe离子（数字）
            '30',  # Ca离子（数字）
            '20',  # Al离子（数字）
            '95.5',  # 白度（数字）
            '2.5',  # 涡值（数字）
            '2.5',  # 气味（数字）
            '150',  # 电导值（数字）
            '7.0',  # pH（数字）
            '25.5',  # 吸油量（数字）
            '30.0',  # 吸水量（数字）
            '0.5',  # 水分（数字）
            '100',  # 袋数（数字）
            '包装类型示例',  # 包装类型
            '10.5000',  # 吨（数字，最多4位小数）
            '批次号示例',  # 批号
            '备注信息',  # 备注（文本）
            '早班',  # 班次
        ]
        ws.append(example_row)
        
        # 设置示例行样式（浅灰色背景，提示这是示例）
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽（按照历史记录字段顺序，不包含过滤时间、水黏度、饼厚）
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品名称
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 15,  # 饼密度
            'J': 15,  # 振实密度
            'K': 12,  # +14M
            'L': 12,  # +30M
            'M': 12,  # +40M
            'N': 12,  # +80M
            'O': 12,  # +100M
            'P': 12,  # +150M
            'Q': 12,  # +200M
            'R': 12,  # +325M
            'S': 12,  # Fe离子
            'T': 12,  # Ca离子
            'U': 12,  # Al离子
            'V': 10,  # 白度
            'W': 12,  # 涡值
            'X': 10,  # 气味
            'Y': 15,  # 电导值
            'Z': 10,  # pH
            'AA': 12,  # 吸油量
            'AB': 12,  # 吸水量
            'AC': 12,  # 水分
            'AD': 10,  # 袋数
            'AE': 15,  # 包装类型
            'AF': 12,  # 吨
            'AG': 15,  # 批号
            'AH': 30,  # 备注
            'AI': 10,  # 班次
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "长富QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        # 使用 attachment 强制浏览器弹出保存对话框，用户可以选择保存位置
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        
        # 添加额外的响应头，确保浏览器弹出保存对话框
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        # 保存到响应
        wb.save(response)
        
        # 记录下载成功日志
        logger.info(f'[下载模板] 成功生成并返回长富QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        # 记录下载失败日志
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成长富QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
def xinghui_report_download_template(request):
    """下载兴辉QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载兴辉QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "兴辉QC报表导入模板"
        
        # 定义表头（按照历史记录页面的字段顺序，排除操作人字段）
        # 字段顺序与 xinghui_report_history.html 中的表头顺序一致
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品型号',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '兴辉渗透率(Darcy)',  # xinghui_permeability
            '饼密度(g/cm3)',  # wet_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M',  # sieving_14m
            '+30M',  # sieving_30m
            '+40M',  # sieving_40m
            '+80M',  # sieving_80m
            '+100M',  # sieving_100m
            '+150M',  # sieving_150m
            '+200M',  # sieving_200m
            '+325M',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油率(%)',  # oil_absorption
            '吸水率(%)',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨数',  # tons
            'LOT批号',  # batch_number
            '备注',  # remarks
            '班组',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（按照新的字段顺序）
        example_row = [
            '2024-01-01',  # 日期
            '08:00',  # 时间
            '10.5',  # 干燥后原土水分
            '2.3',  # 入窑前碱含量
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品型号
            '5.5',  # 远通渗透率
            '6.0',  # 长富渗透率
            '6.5',  # 兴辉渗透率
            '1.5',  # 饼密度
            '1.8',  # 振实密度
            '5.0',  # +14M
            '10.0',  # +30M
            '15.0',  # +40M
            '20.0',  # +80M
            '25.0',  # +100M
            '10.0',  # +150M
            '5.0',  # +200M
            '2.0',  # +325M
            '50',  # Fe离子
            '30',  # Ca离子
            '20',  # Al离子
            '95.5',  # 白度
            '2.5',  # 涡值
            '2.5',  # 气味
            '150',  # 电导值
            '7.0',  # pH
            '25.5',  # 吸油率
            '30.0',  # 吸水率
            '0.5',  # 水分
            '100',  # 袋数
            '包装类型示例',  # 包装类型
            '10.500',  # 吨数
            '批次号示例',  # LOT批号
            '备注信息',  # 备注
            '早班',  # 班组
        ]
        ws.append(example_row)
        
        # 设置示例行样式
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽（按照新的字段顺序）
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品型号
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 18,  # 兴辉渗透率
            'J': 15,  # 饼密度
            'K': 15,  # 振实密度
            'L': 12,  # +14M
            'M': 12,  # +30M
            'N': 12,  # +40M
            'O': 12,  # +80M
            'P': 12,  # +100M
            'Q': 12,  # +150M
            'R': 12,  # +200M
            'S': 12,  # +325M
            'T': 12,  # Fe离子
            'U': 12,  # Ca离子
            'V': 12,  # Al离子
            'W': 10,  # 白度
            'X': 12,  # 涡值
            'Y': 10,  # 气味
            'Z': 15,  # 电导值
            'AA': 10,  # pH
            'AB': 12,  # 吸油率
            'AC': 12,  # 吸水率
            'AD': 12,  # 水分
            'AE': 10,  # 袋数
            'AF': 15,  # 包装类型
            'AG': 12,  # 吨数
            'AH': 15,  # LOT批号
            'AI': 30,  # 备注
            'AJ': 10,  # 班组
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "兴辉QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        wb.save(response)
        
        logger.info(f'[下载模板] 成功生成并返回兴辉QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成兴辉QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
def xinghui2_report_download_template(request):
    """下载兴辉二线QC报表Excel导入模板"""
    logger = logging.getLogger(__name__)
    try:
        # 记录下载日志
        user_info = {
            'username': request.user.username if request.user.is_authenticated else 'Anonymous',
            'user_id': request.user.id if request.user.is_authenticated else None,
            'ip_address': request.META.get('REMOTE_ADDR', 'Unknown'),
            'user_agent': request.META.get('HTTP_USER_AGENT', 'Unknown')[:200],
            'timestamp': str(timezone.now()),
        }
        logger.info(f'[下载模板] 用户请求下载兴辉二线QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 时间: {user_info["timestamp"]}')
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "兴辉二线QC报表导入模板"
        
        # 定义表头（按照历史记录页面的字段顺序，排除操作人字段）
        # 字段顺序与 xinghui2_report_history.html 中的表头顺序一致
        headers = [
            '日期',  # date
            '时间',  # time
            '干燥后原土水分(%)',  # moisture_after_drying
            '入窑前碱含量(%)',  # alkali_content
            '助溶剂添加比例',  # flux
            '产品型号',  # product_name
            '远通渗透率(Darcy)',  # permeability
            '长富渗透率(Darcy)',  # permeability_long
            '兴辉渗透率(Darcy)',  # xinghui_permeability
            '饼密度(g/cm3)',  # wet_cake_density
            '振实密度(g/cm3)',  # bulk_density
            '+14M',  # sieving_14m
            '+30M',  # sieving_30m
            '+40M',  # sieving_40m
            '+80M',  # sieving_80m
            '+100M',  # sieving_100m
            '+150M',  # sieving_150m
            '+200M',  # sieving_200m
            '+325M',  # sieving_325m
            'Fe离子',  # fe_ion
            'Ca离子',  # ca_ion
            'Al离子',  # al_ion
            '白度',  # brightness
            '涡值(cm)',  # swirl
            '气味',  # odor
            '电导值(ms/cm)',  # conductance
            'pH',  # ph
            '吸油率(%)',  # oil_absorption
            '吸水率(%)',  # water_absorption
            '水分(%)',  # moisture
            '袋数',  # bags
            '包装类型',  # packaging
            '吨数',  # tons
            'LOT批号',  # batch_number
            '备注',  # remarks
            '班组',  # shift
        ]
        
        # 写入表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 添加示例数据行（按照新的字段顺序）
        example_row = [
            '2024-01-01',  # 日期
            '08:00',  # 时间
            '10.5',  # 干燥后原土水分
            '2.3',  # 入窑前碱含量
            '助溶剂比例',  # 助溶剂添加比例
            '产品名称示例',  # 产品型号
            '5.5',  # 远通渗透率
            '6.0',  # 长富渗透率
            '6.5',  # 兴辉渗透率
            '1.5',  # 饼密度
            '1.8',  # 振实密度
            '5.0',  # +14M
            '10.0',  # +30M
            '15.0',  # +40M
            '20.0',  # +80M
            '25.0',  # +100M
            '10.0',  # +150M
            '5.0',  # +200M
            '2.0',  # +325M
            '50',  # Fe离子
            '30',  # Ca离子
            '20',  # Al离子
            '95.5',  # 白度
            '2.5',  # 涡值
            '2.5',  # 气味
            '150',  # 电导值
            '7.0',  # pH
            '25.5',  # 吸油率
            '30.0',  # 吸水率
            '0.5',  # 水分
            '100',  # 袋数
            '包装类型示例',  # 包装类型
            '10.500',  # 吨数
            '批次号示例',  # LOT批号
            '备注信息',  # 备注
            '早班',  # 班组
        ]
        ws.append(example_row)
        
        # 设置示例行样式
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for cell in ws[2]:
            cell.fill = example_fill
            cell.border = border
            cell.font = Font(size=10, color="666666")
        
        # 添加说明行
        ws.append(['说明：请删除示例行后填写数据'])
        ws.merge_cells(f'A{ws.max_row}:{get_column_letter(len(headers))}{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(italic=True, color="FF0000", size=10)
        ws[f'A{ws.max_row}'].alignment = Alignment(horizontal="left")
        
        # 设置列宽（按照新的字段顺序）
        column_widths = {
            'A': 12,  # 日期
            'B': 10,  # 时间
            'C': 18,  # 干燥后原土水分
            'D': 15,  # 入窑前碱含量
            'E': 15,  # 助溶剂添加比例
            'F': 20,  # 产品型号
            'G': 18,  # 远通渗透率
            'H': 18,  # 长富渗透率
            'I': 18,  # 兴辉渗透率
            'J': 15,  # 饼密度
            'K': 15,  # 振实密度
            'L': 12,  # +14M
            'M': 12,  # +30M
            'N': 12,  # +40M
            'O': 12,  # +80M
            'P': 12,  # +100M
            'Q': 12,  # +150M
            'R': 12,  # +200M
            'S': 12,  # +325M
            'T': 12,  # Fe离子
            'U': 12,  # Ca离子
            'V': 12,  # Al离子
            'W': 10,  # 白度
            'X': 12,  # 涡值
            'Y': 10,  # 气味
            'Z': 15,  # 电导值
            'AA': 10,  # pH
            'AB': 12,  # 吸油率
            'AC': 12,  # 吸水率
            'AD': 12,  # 水分
            'AE': 10,  # 袋数
            'AF': 15,  # 包装类型
            'AG': 12,  # 吨数
            'AH': 15,  # LOT批号
            'AI': 30,  # 备注
            'AJ': 10,  # 班组
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # 冻结表头行
        ws.freeze_panes = 'A2'
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        import urllib.parse
        filename = "兴辉二线QC报表导入模板.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        response['X-Content-Type-Options'] = 'nosniff'
        
        wb.save(response)
        
        logger.info(f'[下载模板] 成功生成并返回兴辉二线QC报表导入模板 - 用户: {user_info["username"]}, IP: {user_info["ip_address"]}, 文件大小: {len(response.content)} 字节')
        
        return response
        
    except Exception as e:
        username = request.user.username if request.user.is_authenticated else 'Anonymous'
        ip_address = request.META.get('REMOTE_ADDR', 'Unknown')
        logger.error(f'[下载模板] 生成兴辉二线QC报表导入模板失败 - 用户: {username}, IP: {ip_address}, 错误: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'生成模板失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def dongtai_report_import_excel(request):
    """东泰QC报表Excel导入功能 - 使用严格的表头匹配"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': '仅支持POST请求'}, status=405)
    
    try:
        # 检查是否有上传的文件
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': '请选择要导入的Excel文件'}, status=400)
        
        excel_file = request.FILES['excel_file']
        
        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': '仅支持Excel文件格式(.xlsx, .xls)'}, status=400)
        
        # 定义标准的表头（必须与模板完全一致，按照历史记录页面顺序和名称）
        expected_headers = [
            '日期',
            '时间',
            '干燥后原土水分(%)',
            '入窑前碱含量(%)',
            '助溶剂添加比例',
            '产品型号',
            '远通渗透率(Darcy)',
            '长富渗透率(Darcy)',
            '过滤时间(秒)',
            '水黏度(mPa.s)',
            '饼厚(mm)',
            '饼密度(g/cm3)',
            '远通饼密度(g/cm3)',
            '长富饼密度(g/cm3)',
            '振实密度(g/cm3)',
            '+14M',
            '+30M',
            '+40M',
            '+80M',
            '+100M',
            '+150M',
            '+200M',
            '+325M',
            'Fe离子',
            'Ca离子',
            'Al离子',
            '白度',
            '涡值(cm)',
            '气味',
            '电导值(ms/cm)',
            'pH',
            '吸油率(%)',
            '吸水率(%)',
            '水分(%)',
            '袋数',
            '包装类型',
            '吨数',
            'LOT批号',
            '备注',
            '班组',
        ]
        
        # 定义表头到数据库字段的映射（严格一对一，字段名称与历史记录页面一致）
        header_to_field = {
            '日期': 'date',
            '时间': 'time',
            '干燥后原土水分(%)': 'moisture_after_drying',
            '入窑前碱含量(%)': 'alkali_content',
            '助溶剂添加比例': 'flux',
            '产品型号': 'product_name',
            '远通渗透率(Darcy)': 'permeability',
            '长富渗透率(Darcy)': 'permeability_long',
            '过滤时间(秒)': 'filter_time',
            '水黏度(mPa.s)': 'water_viscosity',
            '饼厚(mm)': 'cake_thickness',
            '饼密度(g/cm3)': 'wet_cake_density',
            '远通饼密度(g/cm3)': 'yuantong_cake_density',
            '长富饼密度(g/cm3)': 'changfu_cake_density',
            '振实密度(g/cm3)': 'bulk_density',
            '+14M': 'sieving_14m',
            '+30M': 'sieving_30m',
            '+40M': 'sieving_40m',
            '+80M': 'sieving_80m',
            '+100M': 'sieving_100m',
            '+150M': 'sieving_150m',
            '+200M': 'sieving_200m',
            '+325M': 'sieving_325m',
            'Fe离子': 'fe_ion',
            'Ca离子': 'ca_ion',
            'Al离子': 'al_ion',
            '白度': 'brightness',
            '涡值(cm)': 'swirl',
            '气味': 'odor',
            '电导值(ms/cm)': 'conductance',
            'pH': 'ph',
            '吸油率(%)': 'oil_absorption',
            '吸水率(%)': 'water_absorption',
            '水分(%)': 'moisture',
            '袋数': 'bags',
            '包装类型': 'packaging',
            '吨数': 'tons',
            'LOT批号': 'batch_number',
            '备注': 'remarks',
            '班组': 'shift',
        }
        
        # 读取Excel文件（统一按字符串读取，避免 pandas/openpyxl 自动解析日期时间导致 "The string did not match the expected pattern"）
        try:
            import pandas as pd
            df = pd.read_excel(excel_file, sheet_name=0, header=0, dtype=str)
            df = df.dropna(how='all')  # 删除完全空白的行
            # 将 nan 统一为 None，便于后续判断
            df = df.where(pd.notna(df), None)
            # 获取实际表头
            actual_headers = [str(h).strip() if h is not None and pd.notna(h) else '' for h in df.columns.tolist()]
            
        except ImportError:
            # 如果没有pandas，使用openpyxl
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # 读取表头
            actual_headers = []
            for cell in ws[1]:
                actual_headers.append(str(cell.value).strip() if cell.value else '')
            
            # 读取数据行
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    data_rows.append([cell for cell in row])
            
            # 转换为DataFrame格式
            import pandas as pd
            df = pd.DataFrame(data_rows, columns=actual_headers)
            df = df.dropna(how='all')
        
        # 严格检查表头是否匹配
        if len(actual_headers) != len(expected_headers):
            return JsonResponse({
                'status': 'error',
                'message': f'表头列数不匹配！期望{len(expected_headers)}列，实际{len(actual_headers)}列。请使用模板文件填写数据。'
            }, status=400)
        
        # 检查每个表头是否完全匹配
        header_errors = []
        for i, (expected, actual) in enumerate(zip(expected_headers, actual_headers)):
            if expected != actual:
                header_errors.append(f'第{i+1}列：期望"{expected}"，实际"{actual}"')
        
        if header_errors:
            error_msg = '表头不匹配！请使用模板文件填写数据。\n\n不匹配的列：\n' + '\n'.join(header_errors[:10])
            if len(header_errors) > 10:
                error_msg += f'\n... 还有{len(header_errors)-10}列不匹配'
            return JsonResponse({'status': 'error', 'message': error_msg}, status=400)
        
        # 表头匹配成功，开始处理数据
        # 由于表头已经严格匹配，直接使用header_to_field映射，不需要模糊匹配
        imported_count = 0
        error_count = 0
        error_messages = []
        skipped_count = 0
        
        # 重命名DataFrame列名为数据库字段名（严格一对一映射）
        df.rename(columns=header_to_field, inplace=True)
        
        # 处理每一行数据（使用严格的表头匹配，直接通过字段名获取值）
        import pandas as pd
        for index, row in df.iterrows():
            try:
                # 跳过示例行和说明行（如果存在）
                if index == 0 and '说明' in str(row.get('date', '')):
                    continue
                
                # 直接通过字段名获取值（表头已经严格匹配并重命名）
                def get_value(field_name):
                    """直接从DataFrame行中获取字段值（dtype=str 时空白会变成 'nan' 字符串，视为空）"""
                    try:
                        val = row.get(field_name)
                        if pd.isna(val):
                            return None
                        if isinstance(val, str) and val.strip().lower() == 'nan':
                            return None
                        return val
                    except (KeyError, AttributeError):
                        return None
                
                # 检查是否为空行
                date_val = get_value('date')
                product_name_val = get_value('product_name')
                
                has_any_data = False
                if date_val and pd.notna(date_val):
                    has_any_data = True
                if product_name_val and pd.notna(product_name_val) and str(product_name_val).strip():
                    has_any_data = True
                
                if not has_any_data:
                    # 检查其他关键字段
                    for key in ['shift', 'packaging', 'bags', 'batch_number', 'moisture_after_drying', 
                               'alkali_content', 'permeability', 'permeability_long', 'wet_cake_density']:
                        val = get_value(key)
                        if val and pd.notna(val):
                            has_any_data = True
                            break
                    
                    if not has_any_data:
                        skipped_count += 1
                        continue
                
                data = {}
                
                # 数据校验错误收集
                validation_errors = []
                
                # 处理日期和时间（支持多种格式及 Excel 序列，避免 "The string did not match the expected pattern"）
                if date_val and pd.notna(date_val) and str(date_val).strip():
                    date_val_str = str(date_val).strip()
                    parsed_date = None
                    if isinstance(date_val, str) or (isinstance(date_val, (int, float)) and not hasattr(date_val, 'date')):
                        # 常见字符串格式
                        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'):
                            try:
                                s = (date_val_str[:19] if ' ' in date_val_str else date_val_str[:10])
                                parsed_date = datetime.strptime(s, fmt).date()
                                break
                            except ValueError:
                                continue
                        if parsed_date is None and len(date_val_str) <= 8 and date_val_str.replace('.', '', 1).isdigit():
                            # Excel 日期序列（纯数字，如 45321 或 45321.0）
                            try:
                                serial = float(date_val_str)
                                parsed_date = (datetime(1899, 12, 30) + timedelta(days=int(serial))).date()
                            except (ValueError, OverflowError):
                                pass
                    if parsed_date is None and hasattr(date_val, 'date'):
                        d = getattr(date_val, 'date', None)
                        parsed_date = d() if callable(d) else d
                    if parsed_date is not None:
                        data['date'] = parsed_date
                    else:
                        validation_errors.append(f'字段"日期"值"{date_val}"日期格式错误，无法解析（支持：YYYY-MM-DD、YYYY/MM/DD 或 Excel 日期序列）')
                        data['date'] = None
                else:
                    data['date'] = date.today()
                
                time_val = get_value('time')
                if time_val and pd.notna(time_val) and str(time_val).strip():
                    time_str = str(time_val).strip()
                    parsed_time = None
                    try:
                        if ':' in time_str:
                            for fmt in ('%H:%M', '%H:%M:%S', '%H:%M:%S.%f'):
                                try:
                                    s = time_str[:12] if fmt == '%H:%M:%S.%f' else (time_str[:8] if fmt == '%H:%M:%S' else time_str[:5])
                                    parsed_time = datetime.strptime(s, fmt).time()
                                    break
                                except ValueError:
                                    continue
                        if parsed_time is None:
                            try:
                                frac = float(time_str)
                                if 0 <= frac < 1:
                                    total_seconds = int(frac * 24 * 3600)
                                    parsed_time = time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
                            except (ValueError, TypeError):
                                pass
                        if parsed_time is None and hasattr(time_val, 'time'):
                            t = getattr(time_val, 'time', None)
                            parsed_time = t() if callable(t) else t
                        elif parsed_time is None and isinstance(time_val, (int, float)):
                            total_seconds = int(float(time_val) * 24 * 3600)
                            parsed_time = time(total_seconds // 3600, (total_seconds % 3600) // 60, total_seconds % 60)
                    except Exception as e:
                        validation_errors.append(f'字段"时间"值"{time_val}"处理失败: {str(e)}')
                    if parsed_time is not None:
                        data['time'] = parsed_time
                    else:
                        validation_errors.append(f'字段"时间"值"{time_val}"时间格式错误（支持：HH:MM、HH:MM:SS 或 Excel 时间小数）')
                        data['time'] = None
                else:
                    data['time'] = time(0, 0)
                
                # 处理字符串字段
                string_fields = ['shift', 'product_name', 'packaging', 'batch_number', 'flux', 'remarks', 'material_type', 'swirl']
                for field in string_fields:
                    val = get_value(field)
                    if val and pd.notna(val):
                        data[field] = str(val)
                    else:
                        if field == 'material_type':
                            data[field] = '助熔煅烧品'  # 默认值
                        else:
                            data[field] = ''
                
                # 处理入窑前碱含量
                alkali_val = get_value('alkali_content')
                if alkali_val and pd.notna(alkali_val):
                    try:
                        data['alkali_content'] = float(alkali_val)
                    except:
                        validation_errors.append(f'字段"入窑前碱含量"值"{alkali_val}"无法转换为数字类型')
                        data['alkali_content'] = None
                else:
                    data['alkali_content'] = None
                
                # 处理数字字段（已移除东泰特有字段）
                numeric_fields = [
                    'moisture_after_drying',
                    'permeability', 'permeability_long', 'filter_time', 'water_viscosity', 'cake_thickness',
                    'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density', 'bulk_density', 'brightness', 'odor',
                    'conductance', 'ph', 'moisture', 'bags', 'tons', 'fe_ion', 'ca_ion',
                    'al_ion', 'oil_absorption', 'water_absorption', 'sieving_14m', 'sieving_30m',
                    'sieving_40m', 'sieving_80m'
                ]
                
                # 字段中文名称映射（用于错误提示）
                field_names = {
                    'moisture_after_drying': '干燥后原土水分',
                    'permeability': '远通渗透率',
                    'permeability_long': '长富渗透率',
                    'filter_time': '过滤时间',
                    'water_viscosity': '水黏度',
                    'cake_thickness': '饼厚',
                    'wet_cake_density': '饼密度',
                    'yuantong_cake_density': '远通饼密度',
                    'changfu_cake_density': '长富饼密度',
                    'bulk_density': '振实密度',
                    'brightness': '白度',
                    'odor': '气味',
                    'conductance': '电导值',
                    'ph': 'pH',
                    'moisture': '水分',
                    'bags': '袋数',
                    'tons': '吨数',
                    'fe_ion': 'Fe离子',
                    'ca_ion': 'Ca离子',
                    'al_ion': 'Al离子',
                    'oil_absorption': '吸油率(%)',
                    'water_absorption': '吸水率(%)',
                    'sieving_14m': '+14M',
                    'sieving_30m': '+30M',
                    'sieving_40m': '+40M',
                    'sieving_80m': '+80M',
                }
                
                for field in numeric_fields:
                    val = get_value(field)
                    if val is not None and pd.notna(val):
                        try:
                            if isinstance(val, str):
                                val = val.strip()
                                if val == '' or val.isspace():
                                    data[field] = None
                                else:
                                    float_val = float(val)
                                    if pd.isna(float_val) or (float_val != float_val or abs(float_val) == float('inf')):
                                        data[field] = None
                                    else:
                                        data[field] = float_val
                            elif isinstance(val, (int, float)):
                                if pd.isna(val) or (val != val or abs(val) == float('inf')):
                                    data[field] = None
                                else:
                                    data[field] = float(val)
                            else:
                                data[field] = float(val)
                        except (ValueError, TypeError) as e:
                            # 数字字段转换失败，记录错误
                            field_display_name = field_names.get(field, field)
                            validation_errors.append(f'字段"{field_display_name}"值"{val}"无法转换为数字类型（必须是数字）')
                            data[field] = None
                    else:
                        data[field] = None
                
                # 处理筛分字段（可能是字符串）
                sieving_fields = ['sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m']
                for field in sieving_fields:
                    val = get_value(field)
                    if val and pd.notna(val):
                        data[field] = str(val)
                    else:
                        data[field] = ''
                
                # 再次检查是否为空行
                has_valid_data = False
                if data.get('date'):
                    has_valid_data = True
                if data.get('product_name') and str(data['product_name']).strip():
                    has_valid_data = True
                if data.get('shift') and str(data['shift']).strip():
                    has_valid_data = True
                if data.get('packaging') and str(data['packaging']).strip():
                    has_valid_data = True
                for field in ['bags', 'moisture_after_drying', 'alkali_content', 'permeability', 
                             'permeability_long', 'wet_cake_density', 'brightness', 'swirl']:
                    if data.get(field) is not None:
                        has_valid_data = True
                        break
                
                if not has_valid_data:
                    skipped_count += 1
                    continue
                
                data['user'] = request.user
                data['username'] = request.user.username
                
                # 数据校验：根据数据库字段类型进行校验
                # 导入模型类
                from home.models import DongtaiQCReport
                
                # 字段中文名称映射（用于错误提示，与历史记录页面一致）
                field_display_names = {
                    'date': '日期',
                    'time': '时间',
                    'shift': '班组',
                    'product_name': '产品型号',
                    'packaging': '包装类型',
                    'batch_number': 'LOT批号',
                    'material_type': '物料类型',
                    'moisture_after_drying': '干燥后原土水分(%)',
                    'alkali_content': '入窑前碱含量(%)',
                    'flux': '助溶剂添加比例',
                    'permeability': '远通渗透率(Darcy)',
                    'permeability_long': '长富渗透率(Darcy)',
                    'filter_time': '过滤时间(秒)',
                    'water_viscosity': '水黏度(mPa.s)',
                    'cake_thickness': '饼厚(mm)',
                    'wet_cake_density': '饼密度(g/cm3)',
                    'yuantong_cake_density': '远通饼密度(g/cm3)',
                    'changfu_cake_density': '长富饼密度(g/cm3)',
                    'bulk_density': '振实密度(g/cm3)',
                    'brightness': '白度',
                    'swirl': '涡值(cm)',
                    'odor': '气味',
                    'conductance': '电导值(ms/cm)',
                    'ph': 'pH',
                    'moisture': '水分(%)',
                    'bags': '袋数',
                    'tons': '吨数',
                    'sieving_14m': '+14M',
                    'sieving_30m': '+30M',
                    'sieving_40m': '+40M',
                    'sieving_80m': '+80M',
                    'sieving_100m': '+100M',
                    'sieving_150m': '+150M',
                    'sieving_200m': '+200M',
                    'sieving_325m': '+325M',
                    'fe_ion': 'Fe离子',
                    'ca_ion': 'Ca离子',
                    'al_ion': 'Al离子',
                    'oil_absorption': '吸油率(%)',
                    'water_absorption': '吸水率(%)',
                    'remarks': '备注',
                }
                
                # 对所有字段进行基于模型定义的校验
                for field_name, field_value in data.items():
                    # 跳过user和username字段（这些是系统字段）
                    if field_name in ['user', 'username']:
                        continue
                    
                    # 获取字段显示名称
                    field_display_name = field_display_names.get(field_name, field_name)
                    
                    # 使用模型字段定义进行校验
                    is_valid, error_msg = validate_field_by_model(
                        DongtaiQCReport, 
                        field_name, 
                        field_value, 
                        field_display_name
                    )
                    
                    if not is_valid:
                        validation_errors.append(error_msg)
                
                # 如果有任何校验错误，跳过这条数据，不导入
                if validation_errors:
                    error_count += 1
                    error_msg = f'第 {index + 2} 行数据校验失败，已跳过导入: ' + '; '.join(validation_errors)
                    error_messages.append(error_msg)
                    logger.warning(f'导入东泰QC报表数据校验失败: {error_msg}')
                    continue  # 跳过这条数据，不导入
                
                # 创建记录 - 使用DongtaiQCReport模型
                try:
                    DongtaiQCReport.objects.create(**data)
                    imported_count += 1
                except Exception as db_error:
                    # 数据库层面的错误（如字段类型不匹配、约束违反等）
                    error_count += 1
                    error_detail = str(db_error)
                    if 'max_length' in error_detail.lower() or 'too long' in error_detail.lower():
                        error_msg = f'第 {index + 2} 行导入失败: 字段值超过最大长度限制 - {error_detail}'
                    elif 'decimal' in error_detail.lower() or 'precision' in error_detail.lower():
                        error_msg = f'第 {index + 2} 行导入失败: 数值精度超出限制 - {error_detail}'
                    elif 'invalid' in error_detail.lower() or 'choice' in error_detail.lower():
                        error_msg = f'第 {index + 2} 行导入失败: 字段值无效 - {error_detail}'
                    else:
                        error_msg = f'第 {index + 2} 行导入失败: {error_detail}'
                    error_messages.append(error_msg)
                    logger.error(f'导入东泰QC报表失败: {error_msg}', exc_info=True)
                
            except Exception as e:
                error_count += 1
                error_msg = f'第 {index + 2} 行导入失败: {str(e)}'
                error_messages.append(error_msg)
                logger.error(f'导入东泰QC报表失败: {error_msg}', exc_info=True)
        
        # 记录操作日志
        from home.models import UserOperationLog
        UserOperationLog.log_operation(
            request, 'CREATE', 'dongtai', None,
            f'批量导入Excel数据: 成功{imported_count}条, 失败{error_count}条'
        )
        
        result = {
            'status': 'success',
            'message': f'导入完成！成功导入 {imported_count} 条数据，跳过 {skipped_count} 条空行，失败 {error_count} 条',
            'imported_count': imported_count,
            'error_count': error_count,
            'skipped_count': skipped_count
        }
        
        logger.info(f'📊 导入统计: 成功 {imported_count} 条，跳过 {skipped_count} 条，失败 {error_count} 条')
        
        if error_messages:
            result['error_messages'] = error_messages[:10]
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f'导入东泰QC报表Excel失败: {str(e)}', exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'导入失败: {str(e)}'}, status=500)

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def dayuan_report_edit(request, report_id):
    """大塬QC报表编辑页面"""
    try:
        report = DayuanQCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '大塬QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'permeability', 'permeability_long',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption'
                ]
                for field in ['date', 'time', 'moisture_after_drying', 'alkali_content', 'flux',
                              'product_name', 'permeability', 'permeability_long',
                              'wet_cake_density', 'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m',
                              'sieving_80m', 'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion',
                              'al_ion', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'oil_absorption',
                              'water_absorption', 'moisture', 'bags', 'packaging', 'tons', 'batch_number',
                              'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/dayuan_report/history/'
                })
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'dayuan'
        }
        return render(request, 'production/dayuan_report_edit.html', context)
    except DayuanQCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/dayuan_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/dayuan_report/history/'
        })

@login_required
@permission_required('qc_report_view')
def dayuan_report_history(request):
    """大塬QC报表历史记录页面"""
    return render(request, 'production/dayuan_report_history.html')


@login_required
@permission_required('qc_report_view')
def dongtai_report(request):
    """东泰报表页面"""
    return render(request, 'production/dongtai_report.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def dongtai_report_edit(request, report_id):
    """东泰QC报表编辑页面"""
    try:
        report = DongtaiQCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '东泰QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'flux', 'permeability', 'permeability_long',
                    'dongtai_permeability_coefficient', 'dongtai_sample_weight', 'dongtai_filter_area',
                    'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
                    'filter_time', 'water_viscosity', 'cake_thickness'
                ]
                for field in ['date', 'time', 'material_type', 'dongtai_permeability_coefficient', 'dongtai_sample_weight', 'dongtai_filter_area',
                              'moisture_after_drying', 'alkali_content', 'flux', 'product_name', 'permeability', 'permeability_long', 
                              'filter_time', 'water_viscosity', 'cake_thickness', 'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density',
                              'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                              'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'brightness', 'swirl', 'odor', 
                              'conductance', 'ph', 'oil_absorption', 'water_absorption', 'moisture', 'bags', 'packaging', 
                              'tons', 'batch_number', 'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/dongtai_report/history/'
                })
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'dongtai'
        }
        return render(request, 'production/dongtai_report_edit.html', context)
    except DongtaiQCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/dongtai_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/dongtai_report/history/'
        })


@login_required
@permission_required('qc_report_view')
def dongtai_report_history(request):
    """东泰QC报表历史记录页面"""
    return render(request, 'production/dongtai_report_history.html')


@login_required
@permission_required('qc_report_view')
def yuantong_report(request):
    """远通QC报表页面"""
    return render(request, 'production/yuantong_report.html')

@login_required
@permission_required('qc_report_view')
def yuantong_report_history(request):
    """远通QC报表历史记录页面"""
    return render(request, 'production/yuantong_report_history.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def yuantong_report_edit(request, report_id):
    """远通QC报表编辑页面"""
    try:
        report = YuantongQCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '远通QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'flux', 'permeability', 'permeability_long',
                    'yuantong_permeability_coefficient', 'yuantong_sample_weight', 'yuantong_filter_area',
                    'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
                    'filter_time', 'water_viscosity', 'cake_thickness'
                ]
                for field in ['date', 'time', 'material_type', 'yuantong_permeability_coefficient', 'yuantong_sample_weight', 'yuantong_filter_area',
                              'moisture_after_drying', 'alkali_content', 'flux', 'product_name', 'permeability', 'permeability_long', 
                              'filter_time', 'water_viscosity', 'cake_thickness', 'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density',
                              'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                              'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'brightness', 'swirl', 'odor', 
                              'conductance', 'ph', 'oil_absorption', 'water_absorption', 'moisture', 'bags', 'packaging', 
                              'tons', 'batch_number', 'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        is_expired = False
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                is_expired = True
        
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'yuantong',
            'is_expired': is_expired,
            'edit_limit': edit_limit
        }
        return render(request, 'production/yuantong_report_edit.html', context)
    except YuantongQCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/yuantong_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/yuantong_report/history/'
        })


@login_required
@permission_required('qc_report_view')
def yuantong2_report(request):
    """远通二线QC报表页面"""
    return render(request, 'production/yuantong2_report.html')

@login_required
@permission_required('qc_report_view')
def yuantong2_report_history(request):
    """远通二线QC报表历史记录页面"""
    return render(request, 'production/yuantong2_report_history.html')

@login_required
def yuantong2_report_edit(request, report_id):
    """远通二线QC报表编辑页面"""
    try:
        report = Yuantong2QCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '远通二线QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            # 处理AJAX JSON请求
            try:
                import json
                from django.http import JsonResponse
                
                # 检查请求体是否为空
                if not request.body:
                    return JsonResponse({
                        'status': 'error',
                        'message': '请求体为空'
                    }, status=400)
                
                # 解析JSON数据
                data = json.loads(request.body)
                
                # 检查编辑权限（基于日期限制）
                from datetime import datetime, timedelta
                
                # 获取编辑时间限制参数
                edit_limit = 7  # 默认7天
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                
                # 检查是否在编辑期限内
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    
                    if days_diff > edit_limit:
                        return JsonResponse({
                            'status': 'error',
                            'message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'
                        }, status=403)
                
                # 定义NOT NULL字段（数据库约束要求不能为None）
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                
                # 更新报表数据
                for field in ['date', 'time', 'moisture_after_drying', 'alkali_content', 'flux',
                             'product_name', 'permeability', 'permeability_long',
                             'wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density', 'bulk_density', 
                             'filter_time', 'water_viscosity', 'cake_thickness',
                             'sieving_14m', 'sieving_30m', 'sieving_40m',
                             'sieving_80m', 'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion',
                             'al_ion', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'oil_absorption',
                             'water_absorption', 'moisture', 'bags', 'packaging', 'tons', 'batch_number',
                             'remarks', 'shift', 'material_type', 'yuantong_permeability_coefficient', 
                             'yuantong_sample_weight', 'yuantong_filter_area']:
                    if field in data:
                        value = data.get(field)
                        # 特殊处理NOT NULL字段，确保不为None（数据库约束要求）
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                
                # 确保所有NOT NULL字段总是有值（即使不在数据中）
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                
                report.save()
                
                return JsonResponse({
                    'status': 'success',
                    'message': '报表更新成功！'
                })
                
            except json.JSONDecodeError as e:
                return JsonResponse({
                    'status': 'error',
                    'message': f'请求数据格式错误: {str(e)}'
                }, status=400)
            except Exception as e:
                return JsonResponse({
                    'status': 'error',
                    'message': f'保存失败：{str(e)}'
                }, status=500)
        
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        
        # 获取编辑时间限制参数
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/yuantong2_report/history/'
                })
        
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'yuantong2'
        }
        return render(request, 'production/yuantong2_report_edit.html', context)
        
    except Yuantong2QCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/yuantong2_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/yuantong2_report/history/'
        })

@login_required
@permission_required('qc_report_view')
def xinghui_report(request):
    """大塬报表页面"""
    return render(request, 'production/xinghui_report.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def xinghui_report_edit(request, report_id):
    """兴辉QC报表编辑页面"""
    try:
        report = XinghuiQCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '兴辉QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'permeability', 'permeability_long', 'xinghui_permeability',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption'
                ]
                for field in ['date', 'time', 'moisture_after_drying', 'alkali_content', 'flux',
                              'product_name', 'permeability', 'permeability_long', 'xinghui_permeability',
                              'wet_cake_density', 'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m',
                              'sieving_80m', 'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion',
                              'al_ion', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'oil_absorption',
                              'water_absorption', 'moisture', 'bags', 'packaging', 'tons', 'batch_number',
                              'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/xinghui_report/history/'
                })
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'xinghui'
        }
        return render(request, 'production/xinghui_report_edit.html', context)
    except XinghuiQCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/xinghui_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/xinghui_report/history/'
        })

@login_required
@permission_required('qc_report_view')
def xinghui_report_history(request):
    """大塬QC报表历史记录页面"""
    return render(request, 'production/xinghui_report_history.html')


@login_required
@permission_required('qc_report_view')
def xinghui2_report(request):
    """大塬报表页面"""
    return render(request, 'production/xinghui2_report.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def xinghui2_report_edit(request, report_id):
    """兴辉二线QC报表编辑页面"""
    try:
        report = Xinghui2QCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '兴辉二线QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'permeability', 'permeability_long', 'xinghui_permeability',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption'
                ]
                for field in ['date', 'time', 'moisture_after_drying', 'alkali_content', 'flux',
                              'product_name', 'permeability', 'permeability_long', 'xinghui_permeability',
                              'wet_cake_density', 'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m',
                              'sieving_80m', 'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion',
                              'al_ion', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'oil_absorption',
                              'water_absorption', 'moisture', 'bags', 'packaging', 'tons', 'batch_number',
                              'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/xinghui2_report/history/'
                })
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'xinghui2'
        }
        return render(request, 'production/xinghui2_report_edit.html', context)
    except Xinghui2QCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/xinghui2_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/xinghui2_report/history/'
        })

@login_required
@permission_required('qc_report_view')
def xinghui2_report_history(request):
    """大塬QC报表历史记录页面"""
    return render(request, 'production/xinghui2_report_history.html')


@login_required
@permission_required('qc_report_view')
def changfu_report(request):
    """大塬报表页面"""
    return render(request, 'production/changfu_report.html')

@login_required
@csrf_exempt
@permission_required('qc_report_edit')
def changfu_report_edit(request, report_id):
    """changfuQC报表编辑页面"""
    try:
        report = ChangfuQCReport.objects.get(id=report_id)
        
        # 检查编辑权限
        from home.utils import can_edit_report
        if not can_edit_report(request.user, report, '长富QC报表'):
            return JsonResponse({'status': 'error', 'message': '您没有权限编辑此报表'}, status=403)
        
        if request.method == 'POST':
            try:
                import json
                from django.http import JsonResponse
                if not request.body:
                    return JsonResponse({'status': 'error','message': '请求体为空'}, status=400)
                data = json.loads(request.body)
                from datetime import datetime, timedelta
                edit_limit = 7
                try:
                    from home.models import Parameter
                    param = Parameter.objects.filter(id='report_edit_limit').first()
                    if param and param.value:
                        edit_limit = int(param.value)
                except:
                    pass
                if report.date:
                    report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
                    days_diff = (datetime.now() - report_date).days
                    if days_diff > edit_limit:
                        return JsonResponse({'status': 'error','message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。'}, status=403)
                not_null_fields = ['batch_number', 'odor', 'packaging', 'product_name', 'shift', 'remarks']
                number_fields = [
                    'moisture_after_drying', 'alkali_content', 'permeability', 'permeability_long',
                    'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'conductance', 'ph', 'moisture',
                    'bags', 'tons', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m', 'sieving_100m', 'sieving_150m',
                    'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption'
                ]
                for field in ['date', 'time', 'moisture_after_drying', 'alkali_content', 'flux',
                              'product_name', 'permeability', 'permeability_long',
                              'wet_cake_density', 'bulk_density', 'sieving_14m', 'sieving_30m', 'sieving_40m',
                              'sieving_80m', 'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m', 'fe_ion', 'ca_ion',
                              'al_ion', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'oil_absorption',
                              'water_absorption', 'moisture', 'bags', 'packaging', 'tons', 'batch_number',
                              'remarks', 'shift']:
                    if field in data:
                        value = data.get(field)
                        if field in not_null_fields and (value == '' or value is None):
                            value = ''
                        elif field in number_fields and (value == '' or value is None):
                            value = None
                        elif value == '' or value is None:
                            value = None
                        setattr(report, field, value)
                for field in not_null_fields:
                    if field not in data or data.get(field) is None:
                        setattr(report, field, '')
                report.save()
                return JsonResponse({'status': 'success','message': '报表更新成功！'})
            except json.JSONDecodeError as e:
                return JsonResponse({'status': 'error','message': f'请求数据格式错误: {str(e)}'}, status=400)
            except Exception as e:
                return JsonResponse({'status': 'error','message': f'保存失败：{str(e)}'}, status=500)
        # GET请求 - 显示编辑表单
        # 检查编辑权限（基于日期限制）
        from datetime import datetime, timedelta
        edit_limit = 7  # 默认7天
        try:
            from home.models import Parameter
            param = Parameter.objects.filter(id='report_edit_limit').first()
            if param and param.value:
                edit_limit = int(param.value)
        except:
            pass
        # 检查是否在编辑期限内
        if report.date:
            report_date = datetime.strptime(str(report.date), '%Y-%m-%d')
            days_diff = (datetime.now() - report_date).days
            if days_diff > edit_limit:
                return render(request, 'error.html', {
                    'error_message': f'该记录已超过{edit_limit}天编辑期限，无法编辑。',
                    'back_url': '/changfu_report/history/'
                })
        context = {
            'report': report,
            'is_edit': True,
            'report_type': 'changfu'
        }
        return render(request, 'production/changfu_report_edit.html', context)
    except ChangfuQCReport.DoesNotExist:
        return render(request, 'error.html', {
            'error_message': '记录不存在',
            'back_url': '/changfu_report/history/'
        })
    except Exception as e:
        return render(request, 'error.html', {
            'error_message': f'加载编辑页面失败：{str(e)}',
            'back_url': '/changfu_report/history/'
        })

@login_required
@permission_required('qc_report_view')
def changfu_report_history(request):
    """大塬QC报表历史记录页面"""
    return render(request, 'production/changfu_report_history.html')

@method_decorator(login_required, name='dispatch')
class DayuanQCReportAPI(BaseQCReportAPI):
    """大塬QC报表API"""
    model_class = DayuanQCReport
    report_name = "大塬QC报表"
    history_template = 'production/dayuan_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '大塬QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '大塬QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'wet_cake_density': report.wet_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理大塬报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'moisture_after_drying',
            'alkali_content', 'flux', 'permeability', 'permeability_long', 'wet_cake_density',
            'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        # processed: processed
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed


@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class DongtaiQCReportAPI(BaseQCReportAPI):
    """东泰QC报表API"""
    model_class = DongtaiQCReport
    report_name = "东泰QC报表"
    history_template = 'production/dongtai_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '东泰QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '东泰QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'material_type': report.material_type,
            'dongtai_permeability_coefficient': report.dongtai_permeability_coefficient,
            'dongtai_sample_weight': report.dongtai_sample_weight,
            'dongtai_filter_area': report.dongtai_filter_area,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'filter_time' : report.filter_time,
            'water_viscosity' : report.water_viscosity,
            'cake_thickness' : report.cake_thickness,
            'wet_cake_density': report.wet_cake_density,
            'yuantong_cake_density': report.yuantong_cake_density,
            'changfu_cake_density': report.changfu_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理东泰报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'material_type',
            'dongtai_permeability_coefficient', 'dongtai_sample_weight', 'dongtai_filter_area',
            'moisture_after_drying', 'alkali_content', 'flux', 'permeability', 'permeability_long', 
            'filter_time','water_viscosity', 'cake_thickness','wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density',
            'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed

    

@login_required
@permission_required('qc_report_view')
def export_dongtai_report_excel(request):
    """导出东泰QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, DongtaiQCReport, "东泰", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)


@login_required
def export_dongtai_yesterday_production(request):
    """导出东泰昨日产量统计Excel - 使用与大塬相同的通用函数"""
    return export_production_excel(request, DongtaiQCReport, "东泰", "昨日")



@login_required
def export_dongtai_today_production(request):
    """导出东泰今日产量统计Excel - 使用与大塬相同的通用函数"""
    return export_production_excel(request, DongtaiQCReport, "东泰", "今日")




@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class YuantongQCReportAPI(BaseQCReportAPI):
    """远通QC报表API"""
    model_class = YuantongQCReport
    report_name = "远通QC报表"
    history_template = 'production/yuantong_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '远通QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '远通QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'material_type': report.material_type,
            'yuantong_permeability_coefficient': report.yuantong_permeability_coefficient,
            'yuantong_sample_weight': report.yuantong_sample_weight,
            'yuantong_filter_area': report.yuantong_filter_area,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'filter_time' : report.filter_time,
            'water_viscosity' : report.water_viscosity,
            'cake_thickness' : report.cake_thickness,
            'wet_cake_density': report.wet_cake_density,
            'yuantong_cake_density': report.yuantong_cake_density,
            'changfu_cake_density': report.changfu_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理远通报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'material_type',
            'yuantong_permeability_coefficient', 'yuantong_sample_weight', 'yuantong_filter_area',
            'moisture_after_drying', 'alkali_content', 'flux', 'permeability', 'permeability_long', 
            'filter_time','water_viscosity', 'cake_thickness','wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density',
            'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 为可选字段提供默认值（如果用户没有提供）
        if 'batch_number' not in processed:
            processed['batch_number'] = ''
        
        if 'remarks' not in processed:
            processed['remarks'] = ''
        
        if 'odor' not in processed:
            processed['odor'] = None

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class Yuantong2QCReportAPI(BaseQCReportAPI):
    """远通二线QC报表API"""
    model_class = Yuantong2QCReport
    report_name = "远通二线QC报表"
    history_template = 'production/yuantong2_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '远通二线QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '远通二线QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'material_type': report.material_type,
            'yuantong_permeability_coefficient': report.yuantong_permeability_coefficient,
            'yuantong_sample_weight': report.yuantong_sample_weight,
            'yuantong_filter_area': report.yuantong_filter_area,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'filter_time' : report.filter_time,
            'water_viscosity' : report.water_viscosity,
            'cake_thickness' : report.cake_thickness,
            'wet_cake_density': report.wet_cake_density,
            'yuantong_cake_density': report.yuantong_cake_density,
            'changfu_cake_density': report.changfu_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': timezone.localtime(report.created_at).strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'created_at') and report.created_at else '',
            'updated_at': timezone.localtime(report.updated_at).strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理远通二线报表需要的字段，顺序和远通一致
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'material_type',
            'yuantong_permeability_coefficient', 'yuantong_sample_weight', 'yuantong_filter_area',
            'moisture_after_drying', 'alkali_content', 'flux', 'permeability', 'permeability_long', 
            'filter_time','water_viscosity', 'cake_thickness','wet_cake_density', 'yuantong_cake_density', 'changfu_cake_density',
            'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed

@login_required
@permission_required('qc_report_view')
def export_yuantong_report_excel(request):
    """导出远通QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, YuantongQCReport, "远通", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
def export_yuantong_yesterday_production(request):
    """导出远通昨日产量统计Excel"""
    return export_production_excel(request, YuantongQCReport, "远通", "昨日")

@login_required
def export_yuantong_today_production(request):
    """导出远通今日产量统计Excel"""
    return export_production_excel(request, YuantongQCReport, "远通", "今日")   



@login_required
@permission_required('qc_report_view')
def export_yuantong2_report_excel(request):
    """导出远通二线QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, Yuantong2QCReport, "远通二线", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
def export_yuantong2_yesterday_production(request):
    """导出远通二线昨日产量统计Excel"""
    return export_production_excel(request, Yuantong2QCReport, "远通二线", "昨日")

@login_required
def export_yuantong2_today_production(request):
    """导出远通二线今日产量统计Excel"""
    return export_production_excel(request, Yuantong2QCReport, "远通二线", "今日")


@login_required
@permission_required('qc_report_view')
def export_dayuan_report_excel(request):
    """导出大塬QC报表Excel - 使用通用函数"""
    return export_qc_report_excel_universal(request, DayuanQCReport, "大塬", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
def export_dayuan_yesterday_production(request):
    """导出大塬昨日产量统计Excel"""
    return export_production_excel(request, DayuanQCReport, "大塬", "昨日")

@login_required
def export_dayuan_today_production(request):
    """导出大塬今日产量统计Excel"""
    return export_production_excel(request, DayuanQCReport, "大塬", "今日")


@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class XinghuiQCReportAPI(BaseQCReportAPI):
    """兴辉QC报表API"""
    model_class = XinghuiQCReport
    report_name = "兴辉QC报表"
    history_template = 'production/xinghui_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '兴辉QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '兴辉QC报表')

    def calculate_yesterday_production(self, request):
        """统计昨日产量 - 兴辉版本使用4字段分组（不包含备注）"""
        from datetime import date, timedelta
        
        try:
            # 获取昨天的日期
            yesterday = date.today() - timedelta(days=1)
            
            # 查询昨日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=yesterday,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number')
            
            # 按4个字段分组累加产量（兴辉版本不包含备注）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键 - 兴辉版本不包含备注
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': yesterday.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': yesterday.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def calculate_today_production(self, request):
        """统计今日产量 - 兴辉版本使用4字段分组（不包含备注）"""
        from datetime import date
        
        try:
            # 获取今天的日期
            today = date.today()
            
            # 查询今日的数据，先获取原始数据，然后进行智能累加
            raw_reports = self.model_class.objects.filter(
                date=today,
                tons__isnull=False  # 确保吨数不为空
            ).values(
                'shift',  # 班组
                'product_name',  # 产品型号
                'packaging',  # 包装类型
                'batch_number',  # 批号
                'tons'  # 产量
            ).order_by('shift', 'product_name', 'packaging', 'batch_number')
            
            # 按4个字段分组累加产量（兴辉版本不包含备注）
            grouped_production = {}
            for report in raw_reports:
                # 创建分组键 - 兴辉版本不包含备注
                group_key = (
                    report['shift'] or '未设置',
                    report['product_name'] or '未设置',
                    report['packaging'] or '未设置',
                    report['batch_number'] or '未设置'
                )
                
                if group_key not in grouped_production:
                    grouped_production[group_key] = {
                        'shift': group_key[0],
                        'product_name': group_key[1],
                        'packaging': group_key[2],
                        'batch_number': group_key[3],
                        'total_tons': 0,
                        'count': 0
                    }
                
                # 累加产量
                try:
                    if report['tons'] is not None:
                        grouped_production[group_key]['total_tons'] += float(report['tons'])
                        grouped_production[group_key]['count'] += 1
                except (ValueError, TypeError):
                    continue
            
            # 转换为列表格式
            production_stats = list(grouped_production.values())
            
            # 格式化数据
            result_data = []
            for stat in production_stats:
                result_data.append({
                    'shift': stat['shift'] if stat['shift'] else '未设置',
                    'product_name': stat['product_name'] if stat['product_name'] else '未设置',
                    'packaging': stat['packaging'] if stat['packaging'] else '未设置',
                    'batch_number': stat['batch_number'] if stat['batch_number'] else '未设置',
                    'total_tons': float(stat['total_tons']) if stat['total_tons'] else 0,
                    'date': today.strftime('%Y-%m-%d')
                })
            
            return JsonResponse({
                'status': 'success',
                'data': result_data,
                'date': today.strftime('%Y-%m-%d'),
                'total_groups': len(result_data)
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'统计失败：{str(e)}'
            }, status=500)

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'xinghui_permeability': report.xinghui_permeability,
            'wet_cake_density': report.wet_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report,
                                                                                     'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report,
                                                                                     'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理兴辉报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'moisture_after_drying',
            'alkali_content', 'flux', 'permeability', 'permeability_long', 'xinghui_permeability',
            'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed

@login_required
@permission_required('qc_report_view')
def export_xinghui_report_excel(request):
    """导出兴辉QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, XinghuiQCReport, "兴辉", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
@permission_required('qc_report_view')
def export_xinghui_yesterday_production(request):
    """导出兴辉昨日产量统计Excel - 使用与大塬相同的通用函数，但不包含备注字段"""
    return export_xinghui_production_excel(request, XinghuiQCReport, "兴辉", "昨日")

@login_required
@permission_required('qc_report_view')
def export_xinghui_today_production(request):
    """导出兴辉今日产量统计Excel - 使用与大塬相同的通用函数，但不包含备注字段"""
    return export_xinghui_production_excel(request, XinghuiQCReport, "兴辉", "今日")

def export_xinghui_production_excel(request, model_class, report_name, period):
    """导出兴辉产量统计Excel - 不包含备注字段的4字段版本"""
    logger = logging.getLogger(__name__)
    
    # 添加详细的调试信息
    logger.info(f"=== 开始导出{report_name}{period}产量统计（4字段版本） ===")
    logger.info(f"请求用户: {request.user.username}")
    logger.info(f"请求方法: {request.method}")
    logger.info(f"模型类: {model_class.__name__}")
    logger.info(f"报表名称: {report_name}")
    logger.info(f"统计周期: {period}")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        logger.info("✅ openpyxl库导入成功")
        
        # 确定日期范围
        if period == "昨日":
            target_date = (datetime.now() - timedelta(days=1)).date()
            start_date = target_date
            end_date = target_date
        else:  # 今日
            target_date = datetime.now().date()
            start_date = target_date
            end_date = target_date
        
        logger.info(f"📅 目标日期: {target_date}")
        logger.info(f"📅 查询开始日期: {start_date}")
        logger.info(f"📅 查询结束日期: {end_date}")
        
        # 构建查询 - 兴辉版本不包含备注字段
        query = Q(date__gte=start_date) & Q(date__lte=end_date)
        logger.info(f"🔍 构建查询条件: {query}")
        
        # 执行查询 - 兴辉版本按4个字段排序
        reports = model_class.objects.filter(query).order_by('shift', 'product_name', 'packaging', 'batch_number')
        logger.info(f"📊 查询到{reports.count()}条记录")
        
        if not reports.exists():
            logger.warning(f"⚠️ {report_name}{period}没有找到产量数据")
            return HttpResponse(f"{report_name}{period}没有找到产量数据", content_type='text/plain')
        
        # 按4个字段分组累加产量（兴辉版本不包含备注）
        logger.info("📊 开始按4个字段分组累加产量（兴辉版本）...")
        grouped_production = {}
        for report in reports:
            # 创建分组键 - 兴辉版本不包含备注
            group_key = (
                report.shift or '未设置',
                report.product_name or '未设置',
                report.packaging or '未设置',
                report.batch_number or '未设置'
            )
            
            if group_key not in grouped_production:
                grouped_production[group_key] = {
                    'shift': group_key[0],
                    'product_name': group_key[1],
                    'packaging': group_key[2],
                    'batch_number': group_key[3],
                    'total_tons': 0,
                    'count': 0
                }
            
            # 累加产量
            try:
                if report.tons is not None:
                    grouped_production[group_key]['total_tons'] += float(report.tons)
                    grouped_production[group_key]['count'] += 1
            except (ValueError, TypeError):
                logger.warning(f"    ⚠️ 产量值转换失败: {report.tons}")
                continue
        
        logger.info(f"📊 分组累加完成，共{len(grouped_production)}个唯一组合")
        
        # 按班组分组数据（用于Excel显示）
        grouped_data = {}
        for group_key, production_data in grouped_production.items():
            shift = production_data['shift']
            if shift not in grouped_data:
                grouped_data[shift] = []
            grouped_data[shift].append(production_data)
        
        logger.info(f"📋 按班组分组完成，共{len(grouped_data)}个班组")
        for shift, shift_reports in grouped_data.items():
            logger.info(f"  - 班组 '{shift}': {len(shift_reports)}条记录")
        
        # 创建Excel工作簿
        logger.info("📊 开始创建Excel工作簿...")
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name}{period}产量统计"
        logger.info(f"✅ Excel工作表创建成功，标题: {ws.title}")
        
        # 设置表头 - 兴辉版本6列：班组、产品、包装、批号、产量、班组产量
        headers = ['班组', '产品型号', '包装类型', '批号', '产量(吨)', '班组产量(吨)']
        ws.append(headers)
        logger.info(f"✅ 表头设置完成: {headers}")
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        logger.info("✅ 表头样式设置完成")
        
        # 写入数据
        row_num = 2
        total_tons = 0
        
        # 按班组分组写入数据
        for shift, shift_reports in grouped_data.items():
            shift_total = sum(float(item['total_tons']) for item in shift_reports)
            
            for i, item in enumerate(shift_reports):
                # 写入数据行
                row_data = [
                    item['shift'],
                    item['product_name'],
                    item['packaging'],
                    item['batch_number'],
                    float(item['total_tons']),
                    shift_total
                ]
                
                ws.append(row_data)
                
                # 设置单元格样式
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # 设置数字格式
                    if col_num in [5, 6]:  # 产量列
                        cell.number_format = '0.000'
                    
                    # 设置对齐方式
                    if col_num in [1, 5, 6]:  # 班组、产量列居中
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:  # 其他列左对齐
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 如果是班组第一行，设置班组列和班组产量列的合并
                if i == 0:
                    # 班组列合并
                    if len(shift_reports) > 1:
                        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(shift_reports) - 1, end_column=1)
                        # 设置合并后的单元格样式
                        merged_cell = ws.cell(row=row_num, column=1)
                        merged_cell.font = Font(bold=True, color="1976D2")
                        merged_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # 班组产量列合并
                    if len(shift_reports) > 1:
                        ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num + len(shift_reports) - 1, end_column=6)
                        # 设置合并后的单元格样式
                        merged_cell = ws.cell(row=row_num, column=6)
                        merged_cell.font = Font(bold=True, color="1976D2")
                        merged_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                
                row_num += 1
            
            total_tons += shift_total
        
        logger.info(f"✅ 数据写入完成，共{row_num - 2}行数据")
        
        # 添加总计行
        total_row = ['总计', '', '', '', total_tons, total_tons]
        ws.append(total_row)
        
        # 设置总计行样式
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, 7):  # 兴辉版本6列
            cell = ws.cell(row=row_num, column=col_num, value=total_row[col_num - 1])
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = total_alignment
            
            # 设置数字格式
            if col_num in [5, 6]:
                cell.number_format = '0.00'
        
        logger.info("✅ 总计行设置完成")
        
        # 设置列宽
        column_widths = [12, 15, 12, 15, 12, 15]  # 兴辉版本6列
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        logger.info("✅ 列宽设置完成")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为所有数据单元格设置边框
        for row in range(1, row_num + 1):
            for col in range(1, 7):  # 兴辉版本6列
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        logger.info("✅ 边框设置完成")
        
        # 生成文件名
        filename = f"{report_name}{period}产量统计_{target_date}.xlsx"
        logger.info(f"📁 文件名: {filename}")
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 保存到响应
        wb.save(response)
        logger.info(f"✅ Excel文件生成成功，大小: {len(response.content)} 字节")
        
        return response
        
    except Exception as e:
        error_msg = f"导出失败: {str(e)}"
        logger.error(f"❌ {error_msg}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        return HttpResponse(error_msg, content_type='text/plain')

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class ChangfuQCReportAPI(BaseQCReportAPI):
    """长富QC报表API"""
    model_class = ChangfuQCReport
    report_name = "长富QC报表"
    history_template = 'production/changfu_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '长富QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '长富QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'wet_cake_density': report.wet_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report, 'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理长富报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'moisture_after_drying',
            'alkali_content', 'flux', 'permeability', 'permeability_long', 'wet_cake_density',
            'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value
        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed
@login_required
@permission_required('qc_report_view')
def export_changfu_report_excel(request):
    """导出长富QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, ChangfuQCReport, "长富", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
def export_changfu_yesterday_production(request):
    """导出长富昨日产量统计Excel"""
    return export_production_excel(request, ChangfuQCReport, "长富", "昨日")

@login_required
def export_changfu_today_production(request):
    """导出长富今日产量统计Excel"""
    return export_production_excel(request, ChangfuQCReport, "长富", "今日")

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class Xinghui2QCReportAPI(BaseQCReportAPI):
    """兴辉二线QC报表API"""
    model_class = Xinghui2QCReport
    report_name = "兴辉二线QC报表"
    history_template = 'production/xinghui2_report_history.html'
    field_mapping = QC_REPORT_FIELD_MAPPING

    def _check_edit_permission(self, report, current_user):
        """检查编辑权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_edit_report
        return can_edit_report(current_user, report, '兴辉二线QC报表')

    def _check_delete_permission(self, report, current_user):
        """检查删除权限 - 使用基于公司、部门的权限控制"""
        from home.utils import can_delete_report
        return can_delete_report(current_user, report, '兴辉二线QC报表')

    def _serialize_report(self, report, current_user=None):
        user_info = get_user_info(report.username)
        can_edit = self._check_edit_permission(report, current_user)
        can_delete = self._check_delete_permission(report, current_user)
        return {
            'id': report.id,
            'date': report.date.strftime('%Y-%m-%d') if report.date else '',
            'time': report.time.strftime('%H:%M') if report.time else '',
            'shift': report.shift,
            'product_name': report.product_name,
            'packaging': report.packaging,
            'moisture_after_drying': report.moisture_after_drying,
            'alkali_content': report.alkali_content,
            'flux': report.flux,
            'permeability': report.permeability,
            'permeability_long': report.permeability_long,
            'xinghui_permeability': report.xinghui_permeability,
            'wet_cake_density': report.wet_cake_density,
            'bulk_density': report.bulk_density,
            'brightness': report.brightness,
            'swirl': report.swirl,
            'odor': report.odor,
            'conductance': report.conductance,
            'ph': report.ph,
            'moisture': report.moisture,
            'bags': report.bags,
            'tons': report.tons,
            'fe_ion': report.fe_ion,
            'ca_ion': report.ca_ion,
            'al_ion': report.al_ion,
            'oil_absorption': report.oil_absorption,
            'water_absorption': report.water_absorption,
            'remarks': report.remarks,
            'batch_number': report.batch_number,
            'sieving_14m': report.sieving_14m,
            'sieving_30m': report.sieving_30m,
            'sieving_40m': report.sieving_40m,
            'sieving_80m': report.sieving_80m,
            'sieving_100m': report.sieving_100m,
            'sieving_150m': report.sieving_150m,
            'sieving_200m': report.sieving_200m,
            'sieving_325m': report.sieving_325m,
            'created_at': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report,
                                                                                     'created_at') and report.created_at else '',
            'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(report,
                                                                                     'updated_at') and report.updated_at else '',
            'username': user_info.get('name', report.username),
            'original_username': report.username,
            'can_edit': can_edit,
            'can_delete': can_delete,
            'permission_reason': self._get_permission_reason(report, current_user)
        }

    def _process_input_data(self, data, request):
        # 只处理兴辉报表需要的字段
        fields = [
            'date', 'time', 'shift', 'product_name', 'packaging', 'moisture_after_drying',
            'alkali_content', 'flux', 'permeability', 'permeability_long', 'xinghui_permeability',
            'wet_cake_density', 'bulk_density', 'brightness', 'swirl', 'odor', 'conductance', 'ph', 'moisture',
            'bags', 'tons', 'fe_ion', 'ca_ion', 'al_ion', 'oil_absorption', 'water_absorption',
            'remarks', 'batch_number', 'sieving_14m', 'sieving_30m', 'sieving_40m', 'sieving_80m',
            'sieving_100m', 'sieving_150m', 'sieving_200m', 'sieving_325m'
        ]
        processed = {}
        for field in fields:
            if field in data:
                value = data[field]
                # 处理空字符串，转换为None
                if value == '':
                    processed[field] = None
                else:
                    processed[field] = value

        # 处理日期和时间
        if 'date' in processed and processed['date']:
            processed['date'] = datetime.strptime(processed['date'], '%Y-%m-%d').date()
        if 'time' in processed and processed['time']:
            processed['time'] = datetime.strptime(processed['time'], '%H:%M').time()
        # 用户信息
        processed['user'] = request.user
        processed['username'] = request.user.username
        return processed


@login_required
@permission_required('qc_report_view')
def export_xinghui2_report_excel(request):
    """导出兴辉二线QC报表Excel - 使用大塬格式"""
    return export_qc_report_excel_universal(request, Xinghui2QCReport, "兴辉二线", QC_REPORT_FIELD_MAPPING, use_formatted_style=True)

@login_required
@permission_required('qc_report_view')
def export_xinghui2_yesterday_production(request):
    """导出兴辉二线昨日产量统计Excel - 使用与兴辉相同的通用函数，但不包含备注字段"""
    return export_xinghui2_production_excel(request, Xinghui2QCReport, "兴辉二线", "昨日")

@login_required
@permission_required('qc_report_view')
def export_xinghui2_today_production(request):
    """导出兴辉二线今日产量统计Excel - 使用与兴辉相同的通用函数，但不包含备注字段"""
    return export_xinghui2_production_excel(request, Xinghui2QCReport, "兴辉二线", "今日")

def export_xinghui2_production_excel(request, model_class, report_name, period):
    """导出兴辉二线产量统计Excel - 不包含备注字段的4字段版本"""
    logger = logging.getLogger(__name__)
    
    # 添加详细的调试信息
    logger.info(f"=== 开始导出{report_name}{period}产量统计（4字段版本） ===")
    logger.info(f"请求用户: {request.user.username}")
    logger.info(f"请求方法: {request.method}")
    logger.info(f"模型类: {model_class.__name__}")
    logger.info(f"报表名称: {report_name}")
    logger.info(f"统计周期: {period}")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        logger.info("✅ openpyxl库导入成功")
        
        # 确定日期范围
        if period == "昨日":
            target_date = (datetime.now() - timedelta(days=1)).date()
            start_date = target_date
            end_date = target_date
        else:  # 今日
            target_date = datetime.now().date()
            start_date = target_date
            end_date = target_date
        
        logger.info(f"📅 目标日期: {target_date}")
        logger.info(f"📅 查询开始日期: {start_date}")
        logger.info(f"📅 查询结束日期: {end_date}")
        
        # 构建查询 - 兴辉二线版本不包含备注字段
        query = Q(date__gte=start_date) & Q(date__lte=end_date)
        logger.info(f"🔍 构建查询条件: {query}")
        
        # 执行查询 - 兴辉二线版本按4个字段排序
        reports = model_class.objects.filter(query).order_by('shift', 'product_name', 'packaging', 'batch_number')
        logger.info(f"📊 查询到{reports.count()}条记录")
        
        if not reports.exists():
            logger.warning(f"⚠️ {report_name}{period}没有找到产量数据")
            return HttpResponse(f"{report_name}{period}没有找到产量数据", content_type='text/plain')
        
        # 按4个字段分组累加产量（兴辉二线版本不包含备注）
        logger.info("📊 开始按4个字段分组累加产量（兴辉二线版本）...")
        grouped_production = {}
        for report in reports:
            # 创建分组键 - 兴辉二线版本不包含备注
            group_key = (
                report.shift or '未设置',
                report.product_name or '未设置',
                report.packaging or '未设置',
                report.batch_number or '未设置'
            )
            
            if group_key not in grouped_production:
                grouped_production[group_key] = {
                    'shift': group_key[0],
                    'product_name': group_key[1],
                    'packaging': group_key[2],
                    'batch_number': group_key[3],
                    'total_tons': 0,
                    'count': 0
                }
            
            # 累加产量
            try:
                if report.tons is not None:
                    grouped_production[group_key]['total_tons'] += float(report.tons)
                    grouped_production[group_key]['count'] += 1
            except (ValueError, TypeError):
                logger.warning(f"    ⚠️ 产量值转换失败: {report.tons}")
                continue
        
        logger.info(f"📊 分组累加完成，共{len(grouped_production)}个唯一组合")
        
        # 按班组分组数据（用于Excel显示）
        grouped_data = {}
        for group_key, production_data in grouped_production.items():
            shift = production_data['shift']
            if shift not in grouped_data:
                grouped_data[shift] = []
            grouped_data[shift].append(production_data)
        
        logger.info(f"📋 按班组分组完成，共{len(grouped_data)}个班组")
        for shift, shift_reports in grouped_data.items():
            logger.info(f"  - 班组 '{shift}': {len(shift_reports)}条记录")
        
        # 创建Excel工作簿
        logger.info("📊 开始创建Excel工作簿...")
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name}{period}产量统计"
        logger.info(f"✅ Excel工作表创建成功，标题: {ws.title}")
        
        # 设置表头 - 兴辉二线版本6列：班组、产品、包装、批号、产量、班组产量
        headers = ['班组', '产品型号', '包装类型', '批号', '产量(吨)', '班组产量(吨)']
        ws.append(headers)
        logger.info(f"✅ 表头设置完成: {headers}")
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        logger.info("✅ 表头样式设置完成")
        
        # 写入数据
        row_num = 2
        total_tons = 0
        
        # 按班组分组写入数据
        for shift, shift_reports in grouped_data.items():
            shift_total = sum(float(item['total_tons']) for item in shift_reports)
            
            for i, item in enumerate(shift_reports):
                # 写入数据行
                row_data = [
                    item['shift'],
                    item['product_name'],
                    item['packaging'],
                    item['batch_number'],
                    float(item['total_tons']),
                    shift_total
                ]
                
                ws.append(row_data)
                
                # 设置单元格样式
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # 设置数字格式
                    if col_num in [5, 6]:  # 产量列
                        cell.number_format = '0.000'
                    
                    # 设置对齐方式
                    if col_num in [1, 5, 6]:  # 班组、产量列居中
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:  # 其他列左对齐
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 如果是班组第一行，设置班组列和班组产量列的合并
                if i == 0:
                    # 班组列合并
                    if len(shift_reports) > 1:
                        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(shift_reports) - 1, end_column=1)
                        # 设置合并后的单元格样式
                        merged_cell = ws.cell(row=row_num, column=1)
                        merged_cell.font = Font(bold=True, color="1976D2")
                        merged_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    
                    # 班组产量列合并
                    if len(shift_reports) > 1:
                        ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num + len(shift_reports) - 1, end_column=6)
                        # 设置合并后的单元格样式
                        merged_cell = ws.cell(row=row_num, column=6)
                        merged_cell.font = Font(bold=True, color="1976D2")
                        merged_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                
                row_num += 1
            
            total_tons += shift_total
        
        logger.info(f"✅ 数据写入完成，共{row_num - 2}行数据")
        
        # 添加总计行
        total_row = ['总计', '', '', '', total_tons, total_tons]
        ws.append(total_row)
        
        # 设置总计行样式
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        total_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num in range(1, 7):  # 兴辉二线版本6列
            cell = ws.cell(row=row_num, column=col_num, value=total_row[col_num - 1])
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = total_alignment
            
            # 设置数字格式
            if col_num in [5, 6]:
                cell.number_format = '0.00'
        
        logger.info("✅ 总计行设置完成")
        
        # 设置列宽
        column_widths = [12, 15, 12, 15, 12, 15]  # 兴辉二线版本6列
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        logger.info("✅ 列宽设置完成")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为所有数据单元格设置边框
        for row in range(1, row_num + 1):
            for col in range(1, 7):  # 兴辉二线版本6列
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        logger.info("✅ 边框设置完成")
        
        # 生成文件名
        filename = f"{report_name}{period}产量统计_{target_date}.xlsx"
        logger.info(f"📁 文件名: {filename}")
        
        # 创建HTTP响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 保存到响应
        wb.save(response)
        logger.info(f"✅ Excel文件生成成功，大小: {len(response.content)} 字节")
        
        return response
        
    except Exception as e:
        error_msg = f"导出失败: {str(e)}"
        logger.error(f"❌ {error_msg}")
        logger.error(f"详细错误信息: {traceback.format_exc()}")
        
        return HttpResponse(error_msg, content_type='text/plain', status=500)
