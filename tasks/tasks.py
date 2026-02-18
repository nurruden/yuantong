"""
定时任务模块
用于处理每日大塬QC报表发送等定时任务
"""
import os
import logging
import requests
import tempfile
from datetime import datetime, timedelta, date
from celery import shared_task
from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse
from home.models import DayuanQCReport, DongtaiQCReport, ChangfuQCReport, XinghuiQCReport, Xinghui2QCReport, YuantongQCReport, Yuantong2QCReport
from tasks.models import TaskLog, QCReportSchedule
from home.utils.user_helpers import get_user_info
from home.utils.excel_export import export_qc_report_excel_universal
from home.config import QC_REPORT_FIELD_MAPPING
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@shared_task(bind=True)
def send_daily_dayuan_report(self):
    """
    每日发送大塬QC报表给GaoBieKeLe
    每天早晨8点执行
    """
    task_id = self.request.id
    task_name = "每日大塬QC报表发送"
    
    # 记录任务开始
    TaskLog.objects.create(
        task_name=task_name,
        status='running',
        message='任务开始执行'
    )
    
    try:
        # 获取昨日日期
        yesterday = date.today() - timedelta(days=1)
        
        # 获取昨日的大塬QC报表数据
        reports = DayuanQCReport.objects.filter(date=yesterday).order_by('-created_at')
        
        if not reports.exists():
            message = f"昨日({yesterday})没有大塬QC报表数据"
            logger.warning(message)
            TaskLog.objects.create(
                task_name=task_name,
                status='success',
                message=message
            )
            return message
        
        # 生成Excel报表
        excel_file_path = generate_dayuan_excel_report(reports, yesterday)
        
        # 发送企业微信消息（包含Excel文件）
        result = send_wechat_message_with_file(excel_file_path, yesterday)
        
        # 记录任务成功
        TaskLog.objects.create(
            task_name=task_name,
            status='success',
            message=f"成功发送昨日大塬QC报表，共{reports.count()}条记录"
        )
        
        logger.info(f"每日大塬QC报表发送任务完成: {result}")
        return result
        
    except Exception as e:
        error_msg = f"发送大塬QC报表失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        
        # 记录任务失败
        TaskLog.objects.create(
            task_name=task_name,
            status='failed',
            message=error_msg
        )
        
        raise self.retry(exc=e, countdown=300, max_retries=3)


@shared_task(bind=True)
def send_qc_report_by_schedule(self, report_type):
    """
    根据配置发送指定类型的QC报表
    支持同一报表类型发送给多个用户
    """
    task_name = f"定时发送{report_type}QC报表"
    logger.info(f"[{timezone.now()}] 任务 '{task_name}' 开始执行...")
    
    try:
        # 获取所有启用的配置
        schedules = QCReportSchedule.objects.filter(report_type=report_type, is_enabled=True)
        
        if not schedules.exists():
            logger.warning(f"未找到{report_type}的启用配置，跳过发送")
            return f"未找到{report_type}的启用配置"
        
        # 记录任务开始
        TaskLog.objects.create(
            task_name=task_name,
            status='running',
            message=f'任务开始执行，共{schedules.count()}个接收人'
        )
        
        # 获取昨日数据
        yesterday = date.today() - timedelta(days=1)
        
        # 根据报表类型获取对应的模型
        model_mapping = {
            'dayuan': DayuanQCReport,
            'dongtai': DongtaiQCReport,
            'changfu': ChangfuQCReport,
            'xinghui': XinghuiQCReport,
            'xinghui2': Xinghui2QCReport,
            'yuantong': YuantongQCReport,
            'yuantong2': Yuantong2QCReport,
        }
        
        if report_type not in model_mapping:
            raise ValueError(f"不支持的报表类型: {report_type}")
        
        model_class = model_mapping[report_type]
        reports = model_class.objects.filter(date=yesterday)
        
        # 处理无数据情况
        if not reports.exists():
            no_data_message = f"📊 {schedules.first().get_report_type_display()} - {yesterday.strftime('%Y年%m月%d日')}\n\n⚠️ 未找到昨日{schedules.first().get_report_type_display()}数据。"
            
            # 向所有配置的接收人发送无数据消息
            for schedule in schedules:
                if schedule.send_text:
                    send_wechat_message_to_user(no_data_message, yesterday, schedule.recipient_userid)
            
            TaskLog.objects.create(
                task_name=task_name,
                status='success',
                message=f"无数据，已通知{schedules.count()}个接收人"
            )
            return f"无数据，已通知{schedules.count()}个接收人"
        
        # 生成报表数据（只生成一次，所有接收人共享）
        text_message = None
        excel_file_path = None
        
        if any(schedule.send_text for schedule in schedules):
            text_message = format_qc_report_data(reports, yesterday, schedules.first().get_report_type_display(), None)
        
        if any(schedule.send_excel for schedule in schedules):
            excel_file_path = generate_qc_excel_report(reports, yesterday, schedules.first().get_report_type_display())
        
        # 向每个配置的接收人发送报表
        success_count = 0
        failed_count = 0
        results = []
        
        for schedule in schedules:
            try:
                # 发送文本消息
                if schedule.send_text and text_message:
                    send_wechat_message_to_user(text_message, yesterday, schedule.recipient_userid)
                
                # 发送Excel文件
                if schedule.send_excel and excel_file_path:
                    send_wechat_message_with_file_to_user(excel_file_path, yesterday, schedule.recipient_userid)
                
                success_count += 1
                results.append(f"✅ {schedule.recipient_name}")
                logger.info(f"成功发送给 {schedule.recipient_name}")
                
            except Exception as e:
                failed_count += 1
                results.append(f"❌ {schedule.recipient_name}: {str(e)}")
                logger.error(f"发送给 {schedule.recipient_name} 失败: {str(e)}")
        
        # 清理临时文件
        if excel_file_path:
            try:
                os.unlink(excel_file_path)
            except:
                pass
        
        # 记录任务结果
        result_message = f"成功发送给{success_count}人，失败{failed_count}人"
        TaskLog.objects.create(
            task_name=task_name,
            status='success' if failed_count == 0 else 'failed',
            message=f"{result_message}\n详情: {'; '.join(results)}"
        )
        
        logger.info(f"[{timezone.now()}] {result_message}")
        return result_message
        
    except Exception as e:
        logger.error(f"任务 '{task_name}' 执行失败: {str(e)}", exc_info=True)
        TaskLog.objects.create(
            task_name=task_name,
            status='failed',
            message=f"任务执行失败: {str(e)}"
        )
        raise self.retry(exc=e, countdown=300, max_retries=3)


def generate_qc_excel_report(reports, report_date, report_name):
    """
    生成QC报表Excel文件 - 通用版本
    """
    try:
        from django.contrib.auth.models import User
        from django.db.models import Q
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # 使用与历史记录页面相同的字段映射
        field_mapping = QC_REPORT_FIELD_MAPPING
        
        # 检查每个字段是否在所有行中均为空值（与历史记录页面逻辑一致）
        field_values = {field: [] for field in field_mapping.keys()}
        
        for report in reports:
            # 获取用户真实姓名（与历史记录页面逻辑一致）
            user_display_name = '-'
            if report.user:
                user_display_name = report.user.first_name or report.user.username
            elif report.username:
                try:
                    user = User.objects.get(username=report.username)
                    user_display_name = user.first_name or user.username
                except User.DoesNotExist:
                    user_display_name = report.username

            # 收集每个字段的值（与历史记录页面逻辑一致）
            for field in field_mapping.keys():
                if field == 'username':
                    field_values[field].append(user_display_name)
                elif field == 'date':
                    field_values[field].append(report.date.strftime('%Y-%m-%d') if report.date else '')
                elif field == 'time':
                    field_values[field].append(report.time.strftime('%H:%M') if report.time else '')
                else:
                    value = getattr(report, field, None)
                    field_values[field].append(str(value) if value is not None else '')

        # 过滤掉所有行均为空值的字段（与历史记录页面逻辑一致）
        non_empty_fields = []
        for field, values in field_values.items():
            # 检查是否有任何非空值
            if any(str(value).strip() for value in values if value is not None and str(value).strip()):
                non_empty_fields.append(field)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_name} QC历史记录"

        # 写入表头（只包含非空字段）
        headers = [field_mapping[field] for field in non_empty_fields]
        ws.append(headers)

        # 写入数据（只包含非空字段）
        for i in range(len(reports)):
            row = [field_values[field][i] for field in non_empty_fields]
            ws.append(row)

        # 设置表头样式（与历史记录页面格式完全一致）
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置数据行样式（与历史记录页面格式完全一致）
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row in ws.iter_rows(min_row=2):
            for col_idx, cell in enumerate(row):
                cell.font = data_font
                cell.alignment = data_alignment
                
                # 为吨数字段设置特殊格式（小数点后三位）
                if col_idx < len(non_empty_fields):
                    field_name = non_empty_fields[col_idx]
                    if field_name == 'tons':
                        cell.number_format = '0.000'

        # 设置列宽（与历史记录页面格式完全一致）
        column_widths = {
            'Date日期': 13.25,
            'IPKP CODE包装类型': 24,
            '操作人': 14,
            'LOT批号/日期': 13.5,
        }
        
        # 应用列宽设置（与历史记录页面逻辑完全一致）
        for i, field in enumerate(non_empty_fields, 1):
            column_letter = get_column_letter(i)
            header_name = field_mapping[field]
            
            # 检查是否有特定的列宽设置
            if header_name in column_widths:
                ws.column_dimensions[column_letter].width = column_widths[header_name]
                logger.info(f"设置列 {header_name} 宽度为 {column_widths[header_name]}")
            else:
                # 其余字段设置为8.3（与历史记录页面一致）
                ws.column_dimensions[column_letter].width = 8.3
                logger.info(f"设置列 {header_name} 宽度为 8.3")

        # 设置边框（与历史记录页面格式完全一致）
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # 保存文件
        wb.save(temp_file.name)
        
        logger.info(f"Excel报表生成成功: {temp_file.name}")
        return temp_file.name
        
    except Exception as e:
        logger.error(f"生成Excel报表失败: {str(e)}", exc_info=True)
        raise e


def format_qc_report_data(reports, report_date, report_name, custom_template=None):
    """
    格式化QC报表数据为消息格式 - 通用版本
    """
    if custom_template:
        # 使用自定义模板
        template = custom_template
    else:
        # 使用默认模板
        template = f"""📊 {report_name} - {report_date.strftime('%Y年%m月%d日')}

📈 数据统计:
• 记录数量: {reports.count()}条
• 总袋数: {sum(report.bags or 0 for report in reports):.0f}袋
• 总吨数: {sum(float(report.tons or 0) for report in reports):.3f}吨

📋 产品明细:
{get_product_summary(reports)}

⏰ 发送时间: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"""
    
    return template


def get_product_summary(reports):
    """获取产品汇总信息"""
    product_stats = {}
    for report in reports:
        product = report.product_name or '未知产品'
        if product not in product_stats:
            product_stats[product] = {'count': 0, 'bags': 0, 'tons': 0}
        product_stats[product]['count'] += 1
        product_stats[product]['bags'] += report.bags or 0
        product_stats[product]['tons'] += float(report.tons or 0)
    
    summary_lines = []
    for product, stats in product_stats.items():
        summary_lines.append(f"• {product}: {stats['count']}条, {stats['bags']:.0f}袋, {stats['tons']:.3f}吨")
    
    return '\n'.join(summary_lines) if summary_lines else "• 无产品数据"


def generate_dayuan_excel_report(reports, report_date):
    """
    生成大塬QC报表Excel文件 - 使用与历史记录页面相同的格式
    """
    try:
        from django.contrib.auth.models import User
        from django.db.models import Q
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        
        # 使用与历史记录页面相同的字段映射
        field_mapping = QC_REPORT_FIELD_MAPPING
        
        # 检查每个字段是否在所有行中均为空值（与历史记录页面逻辑一致）
        field_values = {field: [] for field in field_mapping.keys()}
        
        for report in reports:
            # 获取用户真实姓名（与历史记录页面逻辑一致）
            user_display_name = '-'
            if report.user:
                user_display_name = report.user.first_name or report.user.username
            elif report.username:
                try:
                    user = User.objects.get(username=report.username)
                    user_display_name = user.first_name or user.username
                except User.DoesNotExist:
                    user_display_name = report.username

            # 收集每个字段的值（与历史记录页面逻辑一致）
            for field in field_mapping.keys():
                if field == 'username':
                    field_values[field].append(user_display_name)
                elif field == 'date':
                    field_values[field].append(report.date.strftime('%Y-%m-%d') if report.date else '')
                elif field == 'time':
                    field_values[field].append(report.time.strftime('%H:%M') if report.time else '')
                else:
                    value = getattr(report, field, None)
                    field_values[field].append(str(value) if value is not None else '')

        # 过滤掉所有行均为空值的字段（与历史记录页面逻辑一致）
        non_empty_fields = []
        for field, values in field_values.items():
            # 检查是否有任何非空值
            if any(str(value).strip() for value in values if value is not None and str(value).strip()):
                non_empty_fields.append(field)

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"大塬 QC历史记录"

        # 写入表头（只包含非空字段）
        headers = [field_mapping[field] for field in non_empty_fields]
        ws.append(headers)

        # 写入数据（只包含非空字段）
        for i in range(len(reports)):
            row = [field_values[field][i] for field in non_empty_fields]
            ws.append(row)

        # 设置表头样式（与历史记录页面格式一致）
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置数据行样式（与历史记录页面格式一致）
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row in ws.iter_rows(min_row=2):
            for col_idx, cell in enumerate(row):
                cell.font = data_font
                cell.alignment = data_alignment
                
                # 为吨数字段设置特殊格式（小数点后三位）
                if col_idx < len(non_empty_fields):
                    field_name = non_empty_fields[col_idx]
                    if field_name == 'tons':
                        cell.number_format = '0.000'

        # 设置列宽（与历史记录页面格式完全一致）
        column_widths = {
            'Date日期': 13.25,
            'IPKP CODE包装类型': 24,
            '操作人': 14,
            'LOT批号/日期': 13.5,
        }
        
        # 应用列宽设置（与历史记录页面逻辑完全一致）
        for i, field in enumerate(non_empty_fields, 1):
            column_letter = get_column_letter(i)
            header_name = field_mapping[field]
            
            # 检查是否有特定的列宽设置
            if header_name in column_widths:
                ws.column_dimensions[column_letter].width = column_widths[header_name]
                logger.info(f"设置列 {header_name} 宽度为 {column_widths[header_name]}")
            else:
                # 其余字段设置为8.3（与历史记录页面一致）
                ws.column_dimensions[column_letter].width = 8.3
                logger.info(f"设置列 {header_name} 宽度为 8.3")

        # 设置边框（与历史记录页面格式完全一致）
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # 保存文件
        wb.save(temp_file.name)
        
        logger.info(f"Excel报表生成成功: {temp_file.name}")
        return temp_file.name
        
    except Exception as e:
        logger.error(f"生成Excel报表失败: {str(e)}", exc_info=True)
        raise e


def format_dayuan_report_data(reports, report_date):
    """
    格式化大塬QC报表数据为消息格式
    """
    total_reports = reports.count()
    
    # 统计信息
    total_bags = sum(report.bags or 0 for report in reports)
    total_tons = sum(float(report.tons or 0) for report in reports)
    
    # 按产品分组统计
    product_stats = {}
    for report in reports:
        product = report.product_name or '未知产品'
        if product not in product_stats:
            product_stats[product] = {
                'count': 0,
                'bags': 0,
                'tons': 0
            }
        product_stats[product]['count'] += 1
        product_stats[product]['bags'] += report.bags or 0
        product_stats[product]['tons'] += float(report.tons or 0)
    
    # 构建消息内容
    message = f"""📊 大塬QC报表 - {report_date.strftime('%Y年%m月%d日')}

📈 总体统计：
• 报表数量：{total_reports}条
• 总袋数：{total_bags:.0f}袋
• 总吨数：{total_tons:.3f}吨

📋 产品明细："""
    
    for product, stats in product_stats.items():
        message += f"""
• {product}：{stats['count']}条记录，{stats['bags']:.0f}袋，{stats['tons']:.3f}吨"""
    
    # 添加质量指标统计
    quality_indicators = calculate_quality_indicators(reports)
    if quality_indicators:
        message += f"""

🔬 质量指标统计：
• 平均白度：{quality_indicators.get('avg_brightness', 0):.2f}
• 平均pH值：{quality_indicators.get('avg_ph', 0):.2f}
• 平均水分：{quality_indicators.get('avg_moisture', 0):.2f}%"""
    
    message += f"""

⏰ 发送时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
📱 来源：远通信息化系统"""
    
    return message


def calculate_quality_indicators(reports):
    """
    计算质量指标统计
    """
    indicators = {}
    
    # 计算平均值
    brightness_values = [r.brightness for r in reports if r.brightness is not None]
    ph_values = [r.ph for r in reports if r.ph is not None]
    moisture_values = [r.moisture for r in reports if r.moisture is not None]
    
    if brightness_values:
        indicators['avg_brightness'] = sum(brightness_values) / len(brightness_values)
    if ph_values:
        indicators['avg_ph'] = sum(ph_values) / len(ph_values)
    if moisture_values:
        indicators['avg_moisture'] = sum(moisture_values) / len(moisture_values)
    
    return indicators


def send_wechat_message_with_file(excel_file_path, report_date):
    """
    发送企业微信消息（包含Excel文件）给GaoBieKeLe（兼容旧版本）
    """
    return send_wechat_message_with_file_to_user(excel_file_path, report_date, "GaoBieKeLe")


def send_wechat_message_with_file_to_user(excel_file_path, report_date, recipient_userid):
    """
    发送企业微信消息（包含Excel文件）给指定用户
    """
    try:
        # 获取企业微信配置
        corp_id = os.environ.get('WECHAT_CORP_ID')
        corp_secret = os.environ.get('WECHAT_APP_SECRET')
        
        if not corp_id or not corp_secret:
            raise Exception("缺少企业微信配置")
        
        # 获取access_token
        token_url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={corp_id}&corpsecret={corp_secret}'
        token_resp = requests.get(token_url, timeout=10)
        token_data = token_resp.json()
        
        if token_data.get('errcode') != 0:
            raise Exception(f"获取access_token失败: {token_data}")
        
        access_token = token_data.get('access_token')
        
        # 上传文件到企业微信
        upload_url = f'https://qyapi.weixin.qq.com/cgi-bin/media/upload?access_token={access_token}&type=file'
        
        with open(excel_file_path, 'rb') as f:
            files = {'media': (f'QC报表_{report_date.strftime("%Y%m%d")}.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            upload_resp = requests.post(upload_url, files=files, timeout=30)
            upload_data = upload_resp.json()
        
        if upload_data.get('errcode') != 0:
            raise Exception(f"上传文件失败: {upload_data}")
        
        media_id = upload_data.get('media_id')
        
        # 发送文件消息给指定用户
        message_url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
        
        # 先发送文本消息
        text_message = f"""📊 QC报表 - {report_date.strftime('%Y年%m月%d日')}

📈 昨日QC报表已生成，请查收Excel文件。

⏰ 发送时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
📱 来源：远通信息化系统"""
        
        text_data = {
            "touser": recipient_userid,
            "msgtype": "text",
            "agentid": os.environ.get('WECHAT_AGENT_ID', '1000016'),
            "text": {
                "content": text_message
            }
        }
        
        text_response = requests.post(message_url, json=text_data, timeout=10)
        text_result = text_response.json()
        
        if text_result.get('errcode') != 0:
            logger.warning(f"发送文本消息给{recipient_userid}失败: {text_result}")
        
        # 发送文件消息
        file_data = {
            "touser": recipient_userid,
            "msgtype": "file",
            "agentid": os.environ.get('WECHAT_AGENT_ID', '1000016'),
            "file": {
                "media_id": media_id
            }
        }
        
        file_response = requests.post(message_url, json=file_data, timeout=10)
        file_result = file_response.json()
        
        if file_result.get('errcode') != 0:
            raise Exception(f"发送文件消息失败: {file_result}")
        
        logger.info(f"成功发送QC报表Excel文件给{recipient_userid}: {report_date}")
        return f"Excel文件发送成功，errcode: {file_result.get('errcode')}"
        
    except Exception as e:
        logger.error(f"发送企业微信文件消息给{recipient_userid}失败: {str(e)}", exc_info=True)
        raise e


def send_wechat_message(message_content, report_date):
    """
    发送企业微信消息给GaoBieKeLe（兼容旧版本）
    """
    return send_wechat_message_to_user(message_content, report_date, "GaoBieKeLe")


def send_wechat_message_to_user(message_content, report_date, recipient_userid):
    """
    发送企业微信消息给指定用户
    """
    try:
        # 获取企业微信配置
        corp_id = os.environ.get('WECHAT_CORP_ID')
        corp_secret = os.environ.get('WECHAT_APP_SECRET')
        
        if not corp_id or not corp_secret:
            raise Exception("缺少企业微信配置")
        
        # 获取access_token
        token_url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={corp_id}&corpsecret={corp_secret}'
        token_resp = requests.get(token_url, timeout=10)
        token_data = token_resp.json()
        
        if token_data.get('errcode') != 0:
            raise Exception(f"获取access_token失败: {token_data}")
        
        access_token = token_data.get('access_token')
        
        # 发送消息给指定用户
        message_url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
        
        message_data = {
            "touser": recipient_userid,
            "msgtype": "text",
            "agentid": os.environ.get('WECHAT_AGENT_ID', '1000016'),
            "text": {
                "content": message_content
            }
        }
        
        response = requests.post(message_url, json=message_data, timeout=10)
        result = response.json()
        
        if result.get('errcode') != 0:
            raise Exception(f"发送消息失败: {result}")
        
        logger.info(f"成功发送消息给{recipient_userid}: {report_date}")
        return f"消息发送成功，errcode: {result.get('errcode')}"
        
    except Exception as e:
        logger.error(f"发送企业微信消息给{recipient_userid}失败: {str(e)}", exc_info=True)
        raise e


